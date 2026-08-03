"""Tests for experiments/recodeagent/collect.py: developer-test/coverage output
parsers (pure, fixture-string based), the independent stub/completeness scan,
``full``-variant trajectory reconstruction from CLI stdout, ablation/baseagent
exact trajectory rollup from recodeagent_calls.jsonl, build/test/coverage
evaluation (via an injected fake command runner -- never a real toolchain),
per-run collection (missing-vs-zero semantics), matrix-wide collection
(never-attempted vs. not-terminal vs. measured), and raw_runs/failures output.
"""
from __future__ import annotations

import json
import os
from pathlib import Path

import pytest

from experiments.recodeagent import collect as COL
from experiments.recodeagent import common as C
from experiments.recodeagent import run as R
from experiments.recodeagent.common import ExecResult, Measurement, Status


@pytest.mark.skipif(os.name != "posix", reason="requires POSIX special files")
def test_copy_evaluation_tree_skips_runtime_special_files(tmp_path: Path):
    source = tmp_path / "source"
    source.mkdir()
    (source / "lib.rs").write_text("pub fn f() {}\n", encoding="utf-8")
    (source / "internal-link").symlink_to("lib.rs")
    (source / "dangling-link").symlink_to("missing")
    (source / "external-link").symlink_to(tmp_path / "outside")
    os.mkfifo(source / "runtime.pipe")

    destination = tmp_path / "destination"
    COL.copy_evaluation_tree(source, destination)

    assert (destination / "lib.rs").is_file()
    assert (destination / "internal-link").is_symlink()
    assert not (destination / "dangling-link").is_symlink()
    assert not (destination / "external-link").is_symlink()
    assert not (destination / "runtime.pipe").exists()


# --------------------------------------------------------------------------- #
# Fake command runner -- collect.py never spawns a real subprocess in tests
# --------------------------------------------------------------------------- #
class FakeRunner:
    def __init__(self, *, default_returncode=0, default_stdout="", default_stderr=""):
        self.calls: list[dict] = []
        self.default_returncode = default_returncode
        self.default_stdout = default_stdout
        self.default_stderr = default_stderr
        self.script: dict[tuple, ExecResult] = {}

    def script_for(self, argv0: str, result: ExecResult) -> None:
        self.script[argv0] = result

    def __call__(self, argv, *, cwd, timeout=None):
        self.calls.append({"argv": list(argv), "cwd": str(cwd), "timeout": timeout})
        if argv and argv[0] in self.script:
            return self.script[argv[0]]
        return ExecResult(argv=list(argv), returncode=self.default_returncode, stdout=self.default_stdout,
                          stderr=self.default_stderr, duration_s=0.01, timed_out=False,
                          started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), cwd=str(cwd))


def _exec(*, returncode=0, stdout="", stderr="", timed_out=False, error="") -> ExecResult:
    return ExecResult(argv=["x"], returncode=returncode, stdout=stdout, stderr=stderr, duration_s=0.01,
                      timed_out=timed_out, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), error=error)


# --------------------------------------------------------------------------- #
# Developer-test output parsers
# --------------------------------------------------------------------------- #
def test_parse_cargo_test_output_single_binary():
    out = "running 3 tests\n...\ntest result: ok. 2 passed; 1 failed; 0 ignored; 0 measured\n"
    parsed = COL.parse_cargo_test_output(out, "")
    assert parsed == {"total": 3, "passed": 2, "failed": 1}


def test_parse_cargo_test_output_sums_multiple_binaries():
    out = ("test result: ok. 4 passed; 0 failed; 0 ignored; 0 measured\n"
          "test result: FAILED. 1 passed; 2 failed; 0 ignored; 0 measured\n")
    parsed = COL.parse_cargo_test_output(out, "")
    assert parsed == {"total": 7, "passed": 5, "failed": 2}


def test_parse_cargo_test_output_none_when_unrecognized():
    assert COL.parse_cargo_test_output("compiling...\nerror[E0433]\n", "") is None


def test_parse_python_unittest_output_ok():
    out = "...\n----------------------------------------------------------------------\nRan 12 tests in 0.03s\n\nOK\n"
    parsed = COL.parse_python_unittest_output(out, "")
    assert parsed == {"total": 12, "passed": 12, "failed": 0}


def test_parse_python_unittest_output_failed():
    out = ("....F.E\n----------------------------------------------------------------------\n"
          "Ran 7 tests in 0.01s\n\nFAILED (failures=1, errors=1)\n")
    parsed = COL.parse_python_unittest_output(out, "")
    assert parsed == {"total": 7, "passed": 5, "failed": 2}


def test_parse_python_unittest_output_none_when_unrecognized():
    assert COL.parse_python_unittest_output("Traceback (most recent call last):\n", "") is None


def test_parse_jest_output_ok():
    out = "Test Suites: 2 passed, 2 total\nTests:       9 passed, 9 total\nSnapshots:   0 total\n"
    parsed = COL.parse_jest_output(out, "")
    assert parsed == {"total": 9, "passed": 9, "failed": 0}


def test_parse_jest_output_with_failures_and_skips():
    out = "Tests:       2 failed, 1 skipped, 7 passed, 10 total\n"
    parsed = COL.parse_jest_output(out, "")
    assert parsed == {"total": 10, "passed": 7, "failed": 2}


def test_parse_jest_output_none_when_unrecognized():
    assert COL.parse_jest_output("npm ERR! missing script: test\n", "") is None


def test_parse_node_tap_output_ok():
    parsed = COL.parse_node_tap_output("# pass 5\n# fail 0\n", "")
    assert parsed == {"total": 5, "passed": 5, "failed": 0}


def test_parse_node_tap_output_none_when_unrecognized():
    assert COL.parse_node_tap_output("nothing to see here", "") is None


def test_parse_pytest_output_all_passed():
    parsed = COL.parse_pytest_output("collected 5 items\n....\n5 passed in 0.12s\n", "")
    assert parsed == {"total": 5, "passed": 5, "failed": 0}


def test_parse_pytest_output_passed_and_failed():
    parsed = COL.parse_pytest_output("....F..\n3 passed, 2 failed in 1.00s\n", "")
    assert parsed == {"total": 5, "passed": 3, "failed": 2}


def test_parse_pytest_output_errors_count_as_failed():
    parsed = COL.parse_pytest_output("1 failed, 1 error in 0.05s\n", "")
    assert parsed == {"total": 2, "passed": 0, "failed": 2}


def test_parse_pytest_output_skipped_excluded_from_total():
    parsed = COL.parse_pytest_output("3 passed, 2 skipped in 0.20s\n", "")
    assert parsed == {"total": 3, "passed": 3, "failed": 0}


def test_parse_pytest_output_no_tests_ran_is_a_real_zero_not_unavailable():
    """Regression: real pytest prints its OWN duration suffix even on the
    zero-tests-collected line (``no tests ran in 0.00s``), so this must be
    recognized as a real, reportable {"total": 0, ...}, not misparsed into
    None (which collect.py would then surface as Status.UNAVAILABLE) just
    because a duration happened to be present."""
    out = "============================= no tests ran in 0.00s ==============================\n"
    assert COL.parse_pytest_output(out, "") == {"total": 0, "passed": 0, "failed": 0}


def test_parse_pytest_output_none_when_unrecognized():
    assert COL.parse_pytest_output("ImportError: no module named foo\n", "") is None


def test_parse_test_output_dispatches_by_tool():
    out = "test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n"
    assert COL.parse_test_output("crust", out, "") == {"total": 1, "passed": 1, "failed": 0}
    assert COL.parse_test_output("oxidizer", out, "") == {"total": 1, "passed": 1, "failed": 0}


def test_parse_test_output_unknown_tool_returns_none():
    assert COL.parse_test_output("unknown_tool", "anything", "") is None


def test_parse_test_output_dataset_spec_override():
    out = "Ran 1 tests in 0.01s\n\nOK\n"
    # crust normally maps to cargo_test -- override to python_unittest explicitly.
    parsed = COL.parse_test_output("crust", out, "", dataset_spec={"test_output_format": "python_unittest"})
    assert parsed == {"total": 1, "passed": 1, "failed": 0}


def test_parse_test_output_pytest_override_for_alphatrans_independent_oracle():
    """AlphaTrans's OWN unit_test_cmd parser id differs from the pytest-based
    parser its INDEPENDENT oracle check needs -- collect.py selects the
    latter via an explicit dataset_spec override, never a global remap."""
    out = "4 passed in 0.08s\n"
    parsed = COL.parse_test_output("alphatrans", out, "", dataset_spec={"test_output_format": "pytest"})
    assert parsed == {"total": 4, "passed": 4, "failed": 0}


# --------------------------------------------------------------------------- #
# Coverage output parsers
# --------------------------------------------------------------------------- #
def test_parse_coverage_py_json_ok():
    text = json.dumps({"totals": {"percent_covered": 87.5}})
    assert COL.parse_coverage_py_json(text) == 87.5


def test_parse_coverage_py_json_malformed_returns_none():
    assert COL.parse_coverage_py_json("not json") is None
    assert COL.parse_coverage_py_json(json.dumps({"nope": True})) is None


def test_parse_tarpaulin_json_ok():
    assert COL.parse_tarpaulin_json(json.dumps({"coverage": 42.0})) == 42.0


def test_parse_tarpaulin_json_malformed_returns_none():
    assert COL.parse_tarpaulin_json("{broken") is None


def test_parse_istanbul_summary_json_ok():
    text = json.dumps({"total": {"lines": {"pct": 73.2}}})
    assert COL.parse_istanbul_summary_json(text) == 73.2


def test_parse_istanbul_summary_json_missing_keys_returns_none():
    assert COL.parse_istanbul_summary_json(json.dumps({"total": {}})) is None


def test_crust_paper_coverage_pair_unions_developer_and_generated_lines(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    target = tmp_path / "target"
    _write(scaffold / "src" / "bin" / "developer.rs", "PRISTINE TEST")
    _write(target / "src" / "bin" / "developer.rs", "MUTATED TEST")
    _write(target / "src" / "bin" / "generated.rs", "GENERATED TEST")
    _write(target / "src" / "lib.rs", "pub fn f() {}\n")

    class TarpaulinRunner:
        def __call__(self, argv, *, cwd, timeout=None):
            binary = argv[argv.index("--bin") + 1]
            out_dir = Path(argv[argv.index("--output-dir") + 1])
            line = 1 if binary == "developer" else 2
            payload = {
                "files": [{
                    "path": ["/", "tmp", "target", "src", "lib.rs"],
                    "coverable": 2,
                    "traces": [{"line": line, "stats": {"Line": 1}}],
                }]
            }
            _write(out_dir / "tarpaulin-report.json", json.dumps(payload))
            if binary == "developer":
                assert (Path(cwd) / "src" / "bin" / "developer.rs").read_text() == "PRISTINE TEST"
            return _exec(returncode=0)

    before, after = COL.crust_paper_coverage_pair(
        target, scaffold, timeout=30, runner=TarpaulinRunner()
    )

    assert before.is_measured and before.value == pytest.approx(50.0)
    assert after.is_measured and after.value == pytest.approx(100.0)
    assert "1 CodeWeaver-added" in after.reason


def test_crust_codeweaver_coverage_preserves_generated_tests_in_developer_binary(
    tmp_path: Path,
):
    scaffold = tmp_path / "scaffold"
    target = tmp_path / "target"
    _write(scaffold / "src" / "bin" / "developer.rs", "fn main() {}\n")
    _write(
        target / "src" / "bin" / "developer.rs",
        "fn main() {}\n#[test]\nfn test_generated() {}\n",
    )
    _write(target / "src" / "lib.rs", "pub fn f() {}\n")

    class TarpaulinRunner:
        def __call__(self, argv, *, cwd, timeout=None):
            if argv[:2] == ["cargo", "test"]:
                assert any(
                    arg.startswith("__codeweaver_generated_")
                    for arg in argv
                )
                return _exec(stdout="test_generated: test\n")
            binary = argv[argv.index("--bin") + 1]
            out_dir = Path(argv[argv.index("--output-dir") + 1])
            line = 1 if binary == "developer" else 2
            _write(
                out_dir / "tarpaulin-report.json",
                json.dumps({
                    "files": [{
                        "path": ["/", "tmp", "target", "src", "lib.rs"],
                        "coverable": 2,
                        "traces": [{"line": line, "stats": {"Line": 1}}],
                    }]
                }),
            )
            return _exec(
                stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n"
            )

    before, after = COL.crust_paper_coverage_pair(
        target,
        scaffold,
        timeout=30,
        runner=TarpaulinRunner(),
        generated_tests=[("src/bin/developer.rs", "test_generated")],
    )

    assert before.value == pytest.approx(50.0)
    assert after.value == pytest.approx(100.0)
    assert "1 classified CodeWeaver-authored" in after.reason


def test_coverage_py_production_percentage_excludes_test_trees(tmp_path: Path):
    report = {
        "files": {
            "package/core.py": {"summary": {"covered_lines": 6, "num_statements": 10}},
            "package/__init__.py": {"summary": {"covered_lines": 2, "num_statements": 2}},
            "verified_test/test_core.py": {"summary": {"covered_lines": 100, "num_statements": 100}},
            "agent_test/generated.py": {"summary": {"covered_lines": 100, "num_statements": 100}},
            "tests/test_own.py": {"summary": {"covered_lines": 100, "num_statements": 100}},
            "setup.py": {"summary": {"covered_lines": 1, "num_statements": 1}},
        }
    }
    path = tmp_path / "coverage.json"
    path.write_text(json.dumps(report), encoding="utf-8")
    assert COL._coverage_py_production_percentage(path) == pytest.approx(100 * 8 / 12)


def test_alphatrans_codeweaver_coverage_runs_selected_generated_nodeids(
    tmp_path: Path,
):
    target = tmp_path / "target"
    ref = tmp_path / "reference"
    _write(target / "pkg" / "__init__.py", "from .core import add\n")
    _write(target / "pkg" / "core.py", "def add(a, b):\n    return a + b\n")
    _write(
        target / "tests" / "test_generated.py",
        "def test_generated():\n    assert True\n",
    )
    _write(ref / "verified_test" / "CoreTest.py", "def test_add():\n    assert True\n")

    class CoverageRunner:
        def __init__(self):
            self.coverage_runs: list[list[str]] = []

        def __call__(self, argv, *, cwd, timeout=None):
            if "--collect-only" in argv:
                return _exec(stdout="tests/test_generated.py::test_generated\n")
            if argv[:3] == ["python", "-m", "coverage"] and argv[3] == "run":
                self.coverage_runs.append(list(argv))
                return _exec(
                    stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n"
                )
            if argv[:3] == ["python", "-m", "coverage"] and argv[3] == "json":
                output = Path(argv[argv.index("-o") + 1])
                covered = 8 if "combined" in output.name else 5
                _write(
                    output,
                    json.dumps({
                        "files": {
                            "pkg/core.py": {
                                "summary": {
                                    "covered_lines": covered,
                                    "num_statements": 10,
                                }
                            }
                        }
                    }),
                )
                return _exec()
            raise AssertionError(argv)

    runner = CoverageRunner()
    before, after = COL.alphatrans_codeweaver_coverage_pair(
        target,
        ref,
        [("tests/test_generated.py", "test_generated")],
        timeout=30,
        runner=runner,
    )

    assert before.value == pytest.approx(50.0)
    assert after.value == pytest.approx(80.0)
    assert any(
        "tests/test_generated.py::test_generated" in argv
        for argv in runner.coverage_runs
    )


def test_oxidizer_paper_coverage_pair_unions_developer_and_generated_lines(tmp_path: Path):
    target = tmp_path / "target"
    ref = tmp_path / "reference"
    _write(target / "Cargo.toml", "[package]\nname='demo'\nversion='0.1.0'\n")
    _write(target / "src" / "lib.rs", "pub fn translated() {}\n")
    _write(ref / "rust" / "tests" / "demo_test.rs", "#[test]\nfn test_demo() {}\n")
    _write(ref / "rust" / "tests" / "DemoTest_generated.rs", "#[test]\nfn generated_demo() {}\n")

    class TarpaulinRunner:
        def __call__(self, argv, *, cwd, timeout=None):
            test_name = argv[argv.index("--test") + 1]
            out_dir = Path(argv[argv.index("--output-dir") + 1])
            line = 1 if test_name == "demo_test" else 2
            payload = {
                "files": [
                    {
                        "path": ["/", "tmp", "target", "src", "lib.rs"],
                        "coverable": 2,
                        "traces": [{"line": line, "stats": {"Line": 1}}],
                    },
                    {
                        "path": ["/", "tmp", "target", "tests", f"{test_name}.rs"],
                        "coverable": 100,
                        "traces": [{"line": 1, "stats": {"Line": 1}}],
                    },
                ]
            }
            _write(out_dir / "tarpaulin-report.json", json.dumps(payload))
            return _exec(returncode=0)

    before, after = COL.oxidizer_paper_coverage_pair(
        target, ref, name_mapping=None, timeout=30, runner=TarpaulinRunner(),
    )

    assert before.value == pytest.approx(50.0)
    assert after.value == pytest.approx(100.0)
    assert "1 official generated" in after.reason


def test_oxidizer_codeweaver_coverage_runs_only_selected_generated_test(
    tmp_path: Path,
):
    target = tmp_path / "target"
    ref = tmp_path / "reference"
    _write(target / "Cargo.toml", "[package]\nname='demo'\nversion='0.1.0'\n")
    _write(target / "src" / "lib.rs", "pub fn translated() {}\n")
    _write(
        target / "tests" / "own.rs",
        "#[test]\nfn translated_test() {}\n#[test]\nfn generated_test() {}\n",
    )
    _write(
        ref / "rust" / "tests" / "demo_test.rs",
        "#[test]\nfn test_demo() {}\n",
    )

    class TarpaulinRunner:
        def __call__(self, argv, *, cwd, timeout=None):
            if argv[:2] == ["cargo", "test"]:
                return _exec(stdout="translated_test: test\ngenerated_test: test\n")
            test_target = argv[argv.index("--test") + 1]
            out_dir = Path(argv[argv.index("--output-dir") + 1])
            generated = test_target == "own"
            if generated:
                separator = argv.index("--")
                assert argv[separator + 1:separator + 3] == [
                    "--skip", "translated_test",
                ]
                assert "--exact" not in argv
            line = 2 if generated else 1
            _write(
                out_dir / "tarpaulin-report.json",
                json.dumps({
                    "files": [{
                        "path": ["/", "tmp", "target", "src", "lib.rs"],
                        "coverable": 2,
                        "traces": [{"line": line, "stats": {"Line": 1}}],
                    }]
                }),
            )
            return _exec(
                stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n"
            )

    before, after = COL.oxidizer_codeweaver_coverage_pair(
        target,
        ref,
        [("tests/own.rs", "generated_test")],
        name_mapping=None,
        timeout=30,
        runner=TarpaulinRunner(),
    )

    assert before.value == pytest.approx(50.0)
    assert after.value == pytest.approx(100.0)
    assert "1 classified CodeWeaver-authored" in after.reason


def test_skel_paper_coverage_pair_accumulates_generated_c8_data(tmp_path: Path):
    if COL._skel_js_parser() is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "future_project", verified_js_names=["test_trivial"],
    )
    _write(
        ref / "javascript" / "FeatureTest_generated.js",
        "const { add } = require('./source.js');\nif (add(1, 2) !== 3) process.exit(1);\n",
    )
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };\n")

    class C8Runner:
        def __call__(self, argv, *, cwd, timeout=None):
            report_dir = Path(next(arg.split("=", 1)[1] for arg in argv if arg.startswith("--reports-dir=")))
            script = argv[-1]
            pct = 80.0 if "FeatureTest_generated" in script else 50.0
            _write(
                report_dir / "coverage-summary.json",
                json.dumps({"total": {"lines": {"pct": pct}}}),
            )
            return _exec(returncode=0)

    before, after = COL.skel_paper_coverage_pair(
        target, ref, timeout=30, runner=C8Runner(),
    )

    assert before.value == pytest.approx(50.0)
    assert after.value == pytest.approx(80.0)
    assert "1 official generated" in after.reason


def test_skel_evaluator_suppresses_top_level_translated_test_runner(
    tmp_path: Path,
):
    if COL._skel_js_parser() is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed")
    path = tmp_path / "index.js"
    _write(
        path,
        "function testOwn() { return true; }\n"
        "const production = 1;\n"
        "testOwn();\n",
    )

    COL._suppress_skel_top_level_test_calls(path)

    text = path.read_text(encoding="utf-8")
    assert "function testOwn()" in text
    assert text.count("testOwn") == 1
    assert "const production = 1" in text


def test_skel_codeweaver_coverage_uses_full_target_tree_and_generated_harness(
    tmp_path: Path,
):
    if COL._skel_js_parser() is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "future_project",
        verified_js_names=["test_trivial"],
    )
    target = tmp_path / "target"
    _write(
        target / "index.js",
        "const add = require('./math.js').add;\n"
        "function testGenerated() { return add(1, 2) === 3; }\n"
        "testGenerated();\n"
        "module.exports = { add };\n",
    )
    _write(target / "math.js", "exports.add = (a, b) => a + b;\n")

    class C8Runner:
        def __call__(self, argv, *, cwd, timeout=None):
            includes = {
                arg.split("=", 1)[1]
                for arg in argv if arg.startswith("--include=")
            }
            assert includes == {"index.js", "math.js"}
            staged_entry = Path(cwd) / "index.js"
            assert staged_entry.read_text(encoding="utf-8").count(
                "testGenerated"
            ) == 2  # declaration plus evaluator export, never top-level call
            report_dir = Path(next(
                arg.split("=", 1)[1]
                for arg in argv if arg.startswith("--reports-dir=")
            ))
            pct = (
                75.0
                if argv[-1].startswith("__codeweaver_generated_coverage_")
                else 40.0
            )
            _write(
                report_dir / "coverage-summary.json",
                json.dumps({"total": {"lines": {"pct": pct}}}),
            )
            return _exec()

    before, after = COL.skel_codeweaver_coverage_pair(
        target,
        ref,
        [("index.js", "testGenerated")],
        timeout=30,
        runner=C8Runner(),
    )

    assert before.value == pytest.approx(40.0)
    assert after.value == pytest.approx(75.0)
    assert "1 classified CodeWeaver-authored" in after.reason


# --------------------------------------------------------------------------- #
# Independent stub/completeness scan
# --------------------------------------------------------------------------- #
def test_scan_stub_markers_counts_rust_markers(tmp_path: Path):
    root = tmp_path / "target"
    root.mkdir()
    (root / "lib.rs").write_text("fn a() { todo!() }\nfn b() { unimplemented!() }\n", encoding="utf-8")
    (root / "clean.rs").write_text("fn c() -> i32 { 42 }\n", encoding="utf-8")
    m = COL.scan_stub_markers(root, "Rust")
    assert m.is_measured
    assert m.value["stub_marker_count"] == 2
    assert "lib.rs" in m.value["files_with_stubs"]
    assert "clean.rs" not in m.value["files_with_stubs"]


def test_scan_stub_markers_counts_python_markers(tmp_path: Path):
    root = tmp_path / "target"
    root.mkdir()
    (root / "mod.py").write_text("def f():\n    raise NotImplementedError\n", encoding="utf-8")
    m = COL.scan_stub_markers(root, "Python")
    assert m.is_measured
    assert m.value["stub_marker_count"] >= 1


def test_scan_stub_markers_zero_when_no_markers_present(tmp_path: Path):
    root = tmp_path / "target"
    root.mkdir()
    (root / "clean.py").write_text("def f():\n    return 1\n", encoding="utf-8")
    m = COL.scan_stub_markers(root, "Python")
    assert m.is_measured
    assert m.value["stub_marker_count"] == 0


def test_scan_stub_markers_unavailable_for_unknown_language(tmp_path: Path):
    m = COL.scan_stub_markers(tmp_path, "COBOL")
    assert m.status == Status.UNAVAILABLE


def test_scan_stub_markers_missing_when_root_absent(tmp_path: Path):
    m = COL.scan_stub_markers(tmp_path / "nope", "Rust")
    assert m.status == Status.MISSING


# --------------------------------------------------------------------------- #
# `full`-variant CLI stdout trajectory reconstruction
# --------------------------------------------------------------------------- #
CLI_STDOUT_SAMPLE = """\
[codeweaver] project=demo app_id=abc123 mock=False db=/x/pipeline/burr.db
[codeweaver] loaded state at startup: milestone_idx=0 history_len=0  (idx>0 => resumed, not restarted)
    m1  iter=1  passed=False
    m1  iter=2  passed=True
    m2  iter=1  passed=True
[codeweaver] finished at terminal: done=True milestone_idx=2
"""

# novalidator: CodeWeaver core's validate() skip branch appends passed=None to
# EVERY milestone's history entry (no genuine validator attestation exists),
# and unconditionally marks the milestone "passed" internally so the loop
# always advances -- hence exactly one entry per milestone, never a repair
# repeat, and every "passed" value is the literal string "None".
CLI_STDOUT_NOVALIDATOR_SAMPLE = """\
[codeweaver] project=demo app_id=abc123 mock=False db=/x/pipeline/burr.db
[codeweaver] loaded state at startup: milestone_idx=0 history_len=0  (idx>0 => resumed, not restarted)
    m1  iter=1  passed=None
    m2  iter=1  passed=None
[codeweaver] finished at terminal: done=True milestone_idx=2
"""


def test_parse_full_pipeline_stdout_reconstructs_history():
    parsed = COL.parse_full_pipeline_stdout(CLI_STDOUT_SAMPLE)
    assert parsed is not None
    assert parsed["done"] is True
    assert parsed["milestone_idx"] == 2
    assert parsed["history"] == [
        {"milestone": "m1", "iter": 1, "passed": False},
        {"milestone": "m1", "iter": 2, "passed": True},
        {"milestone": "m2", "iter": 1, "passed": True},
    ]


def test_parse_full_pipeline_stdout_none_when_unrelated_text():
    assert COL.parse_full_pipeline_stdout("hello\nworld\n") is None


def test_parse_full_pipeline_stdout_none_when_empty():
    assert COL.parse_full_pipeline_stdout("") is None
    assert COL.parse_full_pipeline_stdout(None) is None


def test_trajectory_from_full_pipeline_computes_loop_count_and_sec():
    traj = COL.trajectory_from_full_pipeline(CLI_STDOUT_SAMPLE, parity_ran=True)
    assert traj.precision == "lower_bound"
    assert traj.sec["translate"] == 3   # one per history entry
    assert traj.sec["validate"] == 3
    assert traj.sec["analyze"] == 1
    assert traj.sec["plan"] == 1
    assert traj.sec["parity"] == 1
    assert traj.lc == 1                # m1 looped once (2 entries - 1 distinct milestone credit)
    assert traj.tec == sum(traj.sec.values())
    assert traj.all_ == traj.tec
    assert traj.nc == sum(1 for v in traj.sec.values() if v > 0)


def test_trajectory_from_full_pipeline_no_parity_excludes_it():
    traj = COL.trajectory_from_full_pipeline(CLI_STDOUT_SAMPLE, parity_ran=False)
    assert traj.sec["parity"] == 0


def test_trajectory_from_full_pipeline_unavailable_without_stdout():
    traj = COL.trajectory_from_full_pipeline(None, parity_ran=False)
    assert traj.precision == "unavailable"
    assert traj.nc is None and traj.tec is None


# --------------------------------------------------------------------------- #
# Regression: stage-skip ablations (noanalyzer/noplanning/novalidator) now
# reconstruct their trajectory identically to `full` via
# trajectory_from_full_pipeline (CodeWeaver core's CODEWEAVER_SKIP_STAGES
# makes all four run the same real Burr CLI subprocess), but the
# deliberately-skipped stage must be excluded from nc/tec/sec exactly like a
# `kind="placeholder"` call already is for trajectory_from_calls -- a
# skipped stage only wrote a placeholder artifact, it never executed.
# --------------------------------------------------------------------------- #
def test_trajectory_from_full_pipeline_skipped_stage_excluded_from_sec_nc_tec():
    traj = COL.trajectory_from_full_pipeline(CLI_STDOUT_SAMPLE, parity_ran=True, skipped_stage="analyze")
    assert "analyze" not in traj.sec
    assert traj.sec["plan"] == 1
    assert traj.sec["translate"] == 3
    assert traj.sec["validate"] == 3
    # nc/tec must not count the skipped "analyze" stage as an executed node.
    assert traj.tec == sum(traj.sec.values())
    assert traj.nc == sum(1 for v in traj.sec.values() if v > 0)
    assert "analyze" in traj.reason
    assert traj.precision == "lower_bound"    # still real evidence-derived, just like `full`


def test_trajectory_from_full_pipeline_skipped_validate_stage_excluded():
    traj = COL.trajectory_from_full_pipeline(CLI_STDOUT_SAMPLE, parity_ran=True, skipped_stage="validate")
    assert "validate" not in traj.sec
    assert traj.sec["analyze"] == 1
    assert traj.sec["translate"] == 3
    assert traj.tec == sum(traj.sec.values())


def test_trajectory_from_full_pipeline_no_skipped_stage_keeps_all_five():
    # Default (None) behaves exactly as before this fix -- `full`'s own path.
    traj = COL.trajectory_from_full_pipeline(CLI_STDOUT_SAMPLE, parity_ran=True, skipped_stage=None)
    assert set(traj.sec) == {"analyze", "plan", "scope", "translate", "validate", "parity"}


# --------------------------------------------------------------------------- #
# Ablation/baseagent exact trajectory from recodeagent_calls.jsonl records
# --------------------------------------------------------------------------- #
def _ablation_calls(skip_stage: str) -> list[dict]:
    calls = []
    for stage in ("analyze", "scope", "plan", "translate", "validate"):
        if stage == skip_stage:
            calls.append({"stage": stage, "kind": "placeholder", "ok": True})
        else:
            calls.append({"stage": stage, "kind": "invoke", "ok": True})
    return calls


def test_trajectory_from_calls_ablation_excludes_placeholder_from_tec():
    calls = _ablation_calls("plan")
    traj = COL.trajectory_from_calls(calls)
    assert traj.precision == "exact"
    # Regression for review finding #3: nc/sec previously counted a
    # placeholder's stage as an executed node (nc was 5, "all stage slots"),
    # even though the "plan" stage was deliberately SKIPPED, not run. nc must
    # now reflect only stages that actually executed, exactly like tec/sec.
    assert traj.nc == 4          # placeholder ("plan") excluded from nc too -- NOT 5
    assert traj.tec == 4         # placeholder ("plan") excluded from actual executions
    assert traj.lc == 0
    assert traj.sec.get("plan") is None
    assert traj.sec["translate"] == 1


def test_trajectory_from_calls_placeholder_stage_never_counted_as_executed_node():
    """Regression for review finding #3: previously ``stages_seen`` (which
    feeds ``nc``) was computed from ALL calls, including ``kind="placeholder"``
    ones, so a deliberately-skipped stage (noanalyzer/noplanning/novalidator)
    silently inflated nc/sec as if it had actually run. A single placeholder
    call, with nothing else, must now report nc=0/tec=0 -- there is nothing
    that actually executed -- not nc=1."""
    calls = [{"stage": "validate", "kind": "placeholder", "ok": True}]
    traj = COL.trajectory_from_calls(calls)
    assert traj.nc == 0
    assert traj.tec == 0
    assert traj.sec == {}
    assert "validate" in traj.reason   # transparency: the skip is still noted in `reason`


def test_trajectory_from_calls_baseagent_single_node():
    calls = [{"stage": "baseagent", "kind": "raw", "ok": True}]
    traj = COL.trajectory_from_calls(calls)
    assert traj.nc == 1
    assert traj.tec == 1
    assert traj.sec == {"baseagent": 1}


def test_trajectory_from_calls_empty_is_unavailable():
    traj = COL.trajectory_from_calls([])
    assert traj.precision == "unavailable"


# --------------------------------------------------------------------------- #
# JSONL tool/token rollup (pipeline/logs/*.stdout.jsonl)
# --------------------------------------------------------------------------- #
def _jsonl_event_line(obj: dict) -> str:
    return json.dumps(obj) + "\n"


def test_collect_jsonl_tool_rollup_sums_across_role_files(tmp_path: Path):
    logs = tmp_path / "logs"
    logs.mkdir()
    (logs / "analyzer.stdout.jsonl").write_text(
        _jsonl_event_line({"type": "assistant.message"})
        + _jsonl_event_line({"type": "tool.execution_complete"})
        + _jsonl_event_line({"type": "tool.execution_complete"})
        + _jsonl_event_line({"type": "result", "exitCode": 0,
                            "usage": {"premiumRequests": 1, "sessionDurationMs": 500}}),
        encoding="utf-8",
    )
    (logs / "translator.stdout.jsonl").write_text(
        _jsonl_event_line({"type": "tool.execution_complete"})
        + _jsonl_event_line({"type": "result", "exitCode": 0,
                            "usage": {"premiumRequests": 2, "sessionDurationMs": 700}}),
        encoding="utf-8",
    )
    rollup, precision = COL.collect_jsonl_tool_rollup(logs)
    assert precision == "lower_bound"
    assert rollup["tool_invocations"] == 3
    assert rollup["assistant_turns"] == 1
    assert rollup["premium_requests"] == 3
    assert rollup["session_duration_ms"] == 1200
    assert "input_tokens" not in rollup  # never fabricated when absent


def test_collect_jsonl_tool_rollup_includes_tokens_when_present(tmp_path: Path):
    logs = tmp_path / "logs"
    logs.mkdir()
    (logs / "translator.stdout.jsonl").write_text(
        _jsonl_event_line({"type": "result", "exitCode": 0,
                          "usage": {"inputTokens": 100, "outputTokens": 50}}),
        encoding="utf-8",
    )
    rollup, _precision = COL.collect_jsonl_tool_rollup(logs)
    assert rollup["input_tokens"] == 100
    assert rollup["output_tokens"] == 50


def test_collect_jsonl_tool_rollup_unique_invocations_are_exact(tmp_path: Path):
    logs = tmp_path / "logs"
    logs.mkdir()
    (logs / "translator.123456-42.stdout.jsonl").write_text(
        _jsonl_event_line({
            "type": "assistant.message",
            "data": {"outputTokens": 50, "toolRequests": [{"name": "view"}]},
        })
        + _jsonl_event_line({
            "type": "session.usage_checkpoint",
            "data": {"totalNanoAiu": 1234, "totalPremiumRequests": 1},
        }),
        encoding="utf-8",
    )
    rollup, precision = COL.collect_jsonl_tool_rollup(logs)
    assert precision == "exact"
    assert rollup["output_tokens"] == 50
    assert rollup["nano_aiu"] == 1234
    assert rollup["tool_counts"] == {"view": 1}
    assert "input_tokens" not in rollup


def test_collect_jsonl_tool_rollup_unavailable_when_no_logs_dir(tmp_path: Path):
    rollup, precision = COL.collect_jsonl_tool_rollup(tmp_path / "nope")
    assert rollup == {}
    assert precision == "unavailable"


# --------------------------------------------------------------------------- #
# Build / test / coverage evaluation (fake runner only)
# --------------------------------------------------------------------------- #
def test_evaluate_build_not_applicable_when_no_cmd_configured(tmp_path: Path):
    m = COL.evaluate_build(tmp_path, [], timeout=None, runner=FakeRunner())
    assert m.status == Status.NOT_APPLICABLE


def test_evaluate_build_missing_when_target_dir_absent(tmp_path: Path):
    m = COL.evaluate_build(tmp_path / "nope", ["cargo", "build"], timeout=None, runner=FakeRunner())
    assert m.status == Status.MISSING


def test_evaluate_build_measured_success(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_returncode=0)
    m = COL.evaluate_build(target, ["cargo", "build"], timeout=None, runner=runner)
    assert m.is_measured
    assert m.value is True
    assert runner.calls[0]["argv"] == ["cargo", "build"]
    assert runner.calls[0]["cwd"] == str(target)


def test_evaluate_build_measured_failure_is_false_not_missing(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_returncode=1)
    m = COL.evaluate_build(target, ["cargo", "build"], timeout=None, runner=runner)
    assert m.is_measured
    assert m.value is False


def test_evaluate_build_timeout_is_error_status(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()

    def timeout_runner(argv, *, cwd, timeout=None):
        return _exec(timed_out=True)

    m = COL.evaluate_build(target, ["cargo", "build"], timeout=5, runner=timeout_runner)
    assert m.status == Status.ERROR


def test_evaluate_tests_not_applicable_when_no_cmd(tmp_path: Path):
    result = COL.evaluate_tests(tmp_path, [], "crust", timeout=None, runner=FakeRunner())
    assert result["total"].status == Status.NOT_APPLICABLE


def test_evaluate_tests_missing_when_target_absent(tmp_path: Path):
    result = COL.evaluate_tests(tmp_path / "nope", ["cargo", "test"], "crust", timeout=None, runner=FakeRunner())
    assert all(v.status == Status.MISSING for v in result.values())


def test_evaluate_tests_measured_when_parseable(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_stdout="test result: ok. 3 passed; 1 failed; 0 ignored; 0 measured\n")
    result = COL.evaluate_tests(target, ["cargo", "test"], "crust", timeout=None, runner=runner)
    assert result["total"].value == 4
    assert result["passed"].value == 3
    assert result["failed"].value == 1


def test_evaluate_tests_unavailable_when_output_unrecognized(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_stdout="???unparseable output???")
    result = COL.evaluate_tests(target, ["cargo", "test"], "crust", timeout=None, runner=runner)
    assert result["total"].status == Status.UNAVAILABLE


def test_evaluate_tests_error_on_exec_error(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()

    def erroring_runner(argv, *, cwd, timeout=None):
        return _exec(error="cargo: command not found")

    result = COL.evaluate_tests(target, ["cargo", "test"], "crust", timeout=None, runner=erroring_runner)
    assert all(v.status == Status.ERROR for v in result.values())


# --------------------------------------------------------------------------- #
# evaluate_tests: an unparseable/compile-failure output must (a) STAY
# Status.UNAVAILABLE for total/passed/failed (never a fabricated 0/"failed"
# value -- this was ALREADY true before this change; re-asserted here as a
# regression guard) and (b) now additionally surface the actual
# compiler-error TEXT in its reason string (this review-finding fix), so a
# reader/caller can tell "never compiled" apart from "compiled and every
# assertion genuinely failed" from the reason alone, not just the status.
# --------------------------------------------------------------------------- #
def test_evaluate_tests_unavailable_reason_includes_compiler_error_snippet_when_present(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_returncode=101, default_stdout="",
                        default_stderr="error[E0425]: cannot find function `NewLuhn` in this scope\n"
                                       " --> tests/checkdigit_test.rs:5:17\n")
    result = COL.evaluate_tests(target, ["cargo", "test"], "oxidizer", timeout=None, runner=runner)
    for key in ("total", "passed", "failed"):
        assert result[key].status == Status.UNAVAILABLE
        assert result[key].value is None      # never fabricated as measured/zero
        assert "likely a compile/import failure" in result[key].reason
        assert "not a behavioral test failure" in result[key].reason
        assert "cannot find function `NewLuhn`" in result[key].reason


def test_evaluate_tests_unavailable_reason_has_no_snippet_when_output_is_entirely_blank(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_returncode=139, default_stdout="", default_stderr="")   # e.g. a bare segfault
    result = COL.evaluate_tests(target, ["cargo", "test"], "crust", timeout=None, runner=runner)
    assert result["total"].status == Status.UNAVAILABLE
    assert result["total"].value is None
    assert "did not match a recognized" in result["total"].reason
    assert "likely a compile/import failure" not in result["total"].reason   # nothing to extract, none fabricated


def test_evaluate_tests_never_sets_failed_to_measured_expected_on_compile_failure(tmp_path: Path):
    """The user's core invariant, exercised through the PUBLIC evaluate_tests
    entry point directly (see also the ``_finalize_validated_tests``
    end-to-end version below, under the Oxidizer identifier-rewrite
    section): a suite that never compiled must never report
    ``failed.status == Status.MEASURED``, regardless of how "confident" the
    compiler error text looks."""
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_returncode=101, default_stdout="",
                        default_stderr="error: could not compile `checkdigit` due to previous error\n")
    result = COL.evaluate_tests(target, ["cargo", "test"], "oxidizer", timeout=None, runner=runner)
    assert result["failed"].status != Status.MEASURED
    assert result["failed"].value is None


# --------------------------------------------------------------------------- #
# extract_compiler_error_snippet: tool-agnostic best-effort compiler/
# interpreter error-line extraction (pure function, no I/O).
# --------------------------------------------------------------------------- #
def test_extract_compiler_error_snippet_rust_error_code_line():
    stderr = "error[E0425]: cannot find function `NewLuhn` in this scope\n --> tests/x.rs:5:17\n"
    assert COL.extract_compiler_error_snippet("", stderr) == \
        "error[E0425]: cannot find function `NewLuhn` in this scope"


def test_extract_compiler_error_snippet_rust_plain_error_line():
    stderr = "some preamble\nerror: could not compile `checkdigit` due to previous error\n"
    snippet = COL.extract_compiler_error_snippet("", stderr)
    assert snippet == "error: could not compile `checkdigit` due to previous error"


def test_extract_compiler_error_snippet_prefers_named_exception_over_traceback_preamble():
    """A bare 'Traceback (most recent call last):' line names no actual
    error by itself -- the concrete ``XxxError``/``XxxException`` line
    (however far below the preamble) is always the more informative,
    preferred snippet."""
    stdout = ("Traceback (most recent call last):\n"
             "  File \"mod.py\", line 3, in <module>\n"
             "    import foo\n"
             "ModuleNotFoundError: No module named 'foo'\n")
    snippet = COL.extract_compiler_error_snippet(stdout, "")
    assert snippet == "ModuleNotFoundError: No module named 'foo'"
    assert "Traceback" not in snippet


def test_extract_compiler_error_snippet_falls_back_to_traceback_preamble_when_no_named_exception():
    stdout = "Traceback (most recent call last):\n  File \"mod.py\", line 3\n"
    snippet = COL.extract_compiler_error_snippet(stdout, "")
    assert snippet is not None
    assert "Traceback (most recent call last):" in snippet


def test_extract_compiler_error_snippet_js_style_reference_error():
    stderr = "/x/index.js:10\nReferenceError: NewLuhn is not defined\n    at Object.<anonymous>\n"
    snippet = COL.extract_compiler_error_snippet("", stderr)
    assert snippet == "ReferenceError: NewLuhn is not defined"


def test_extract_compiler_error_snippet_prefers_last_specific_line_when_several_present():
    stderr = ("ImportError: cannot import name 'a'\n"
             "\n"
             "During handling of the above exception, another exception occurred:\n"
             "\n"
             "AttributeError: module 'b' has no attribute 'c'\n")
    snippet = COL.extract_compiler_error_snippet("", stderr)
    assert snippet == "AttributeError: module 'b' has no attribute 'c'"


def test_extract_compiler_error_snippet_falls_back_to_last_nonblank_line_when_unrecognized():
    stderr = "some unrelated diagnostic\nprocess exited abnormally\n\n"
    snippet = COL.extract_compiler_error_snippet("", stderr)
    assert snippet == "process exited abnormally"


def test_extract_compiler_error_snippet_none_when_both_streams_blank():
    assert COL.extract_compiler_error_snippet("", "") is None
    assert COL.extract_compiler_error_snippet("   \n  ", "\n\n") is None


def test_extract_compiler_error_snippet_truncates_long_lines():
    long_msg = "error: " + ("x" * 500)
    snippet = COL.extract_compiler_error_snippet("", long_msg)
    assert snippet is not None
    assert len(snippet) == 300
    assert snippet.endswith("...")


def test_compute_pass_rate_measured():
    r = COL.compute_pass_rate(Measurement.ok(10), Measurement.ok(7))
    assert r.is_measured
    assert r.value == pytest.approx(0.7)


def test_compute_pass_rate_zero_total_is_not_applicable():
    r = COL.compute_pass_rate(Measurement.ok(0), Measurement.ok(0))
    assert r.status == Status.NOT_APPLICABLE


def test_compute_pass_rate_missing_when_not_both_measured():
    r = COL.compute_pass_rate(Measurement.missing("x"), Measurement.ok(1))
    assert r.status == Status.MISSING


def test_evaluate_coverage_unavailable_when_not_configured(tmp_path: Path):
    m = COL.evaluate_coverage(tmp_path, [], None, timeout=None, runner=FakeRunner())
    assert m.status == Status.UNAVAILABLE


def test_evaluate_coverage_measured_when_configured_and_parseable(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_stdout=json.dumps({"totals": {"percent_covered": 55.5}}))
    m = COL.evaluate_coverage(target, ["coverage", "json"], "coverage_py_json", timeout=None, runner=runner)
    assert m.is_measured
    assert m.value == 55.5


def test_evaluate_coverage_unavailable_when_unparseable(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    runner = FakeRunner(default_stdout="not json at all")
    m = COL.evaluate_coverage(target, ["coverage", "json"], "coverage_py_json", timeout=None, runner=runner)
    assert m.status == Status.UNAVAILABLE


# --------------------------------------------------------------------------- #
# Target function/test counting (reuses manifest.py's heuristics)
# --------------------------------------------------------------------------- #
def test_target_function_counts_rust(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    (target / "lib.rs").write_text("fn a() {}\npub fn b() {}\n", encoding="utf-8")
    m = COL.target_function_counts(target, "Rust")
    assert m.is_measured
    assert m.value == 2


def test_target_function_counts_missing_when_absent(tmp_path: Path):
    m = COL.target_function_counts(tmp_path / "nope", "Rust")
    assert m.status == Status.MISSING


def test_target_function_counts_unavailable_unknown_language(tmp_path: Path):
    m = COL.target_function_counts(tmp_path, "COBOL")
    assert m.status == Status.UNAVAILABLE


def test_target_test_counts_rust(tmp_path: Path):
    target = tmp_path / "target"
    target.mkdir()
    (target / "lib.rs").write_text("#[test]\nfn t1() {}\n#[test]\nfn t2() {}\n", encoding="utf-8")
    m = COL.target_test_counts(target, "Rust")
    assert m.is_measured
    assert m.value == 2


# --------------------------------------------------------------------------- #
# Per-function / per-milestone validation
# --------------------------------------------------------------------------- #
def test_milestone_validation_full_variant_reconstructs_from_history(tmp_path: Path):
    result = COL.milestone_validation(tmp_path, "full", CLI_STDOUT_SAMPLE, None)
    assert result["granularity"] == "real"
    assert result["total"].value == 2      # m1, m2
    assert result["passed"].value == 2     # m1's LATEST attempt passed=True; m2 passed=True


def test_milestone_validation_full_variant_missing_without_history(tmp_path: Path):
    result = COL.milestone_validation(tmp_path, "full", None, None)
    assert result["total"].status == Status.MISSING


def test_milestone_validation_baseagent_single_synthetic_milestone():
    result = COL.milestone_validation(Path("."), "baseagent-condensed", None, True)
    assert result["granularity"] == "single-synthetic"
    assert result["total"].value == 1
    assert result["passed"].value == 1


def test_milestone_validation_baseagent_missing_without_final_call():
    result = COL.milestone_validation(Path("."), "baseagent-condensed", None, None)
    assert result["total"].status == Status.MISSING


# --------------------------------------------------------------------------- #
# Regression: noanalyzer/noplanning/novalidator now dispatch to the SAME
# real per-milestone-history reconstruction as `full` (CodeWeaver core's
# CODEWEAVER_SKIP_STAGES makes all four run the identical real Burr graph),
# never the old single-synthetic-milestone rollup, which is now reserved for
# baseagent-condensed/baseagent-concat only (see tests immediately above).
# --------------------------------------------------------------------------- #
def test_milestone_validation_noanalyzer_uses_real_granularity_not_single_synthetic(tmp_path: Path):
    result = COL.milestone_validation(tmp_path, "noanalyzer", CLI_STDOUT_SAMPLE, None, skipped_stage="analyze")
    assert result["granularity"] == "real"
    assert result["total"].value == 2
    assert result["passed"].value == 2     # noanalyzer's validator is real -- genuine pass/fail


def test_milestone_validation_noplanning_uses_real_granularity_not_single_synthetic(tmp_path: Path):
    result = COL.milestone_validation(tmp_path, "noplanning", CLI_STDOUT_SAMPLE, None, skipped_stage="plan")
    assert result["granularity"] == "real"
    assert result["total"].value == 2
    assert result["passed"].value == 2


def test_milestone_validation_novalidator_passed_is_missing_not_zero(tmp_path: Path):
    """Core regression: novalidator's validate() skip branch appends
    passed=None to EVERY milestone's history entry (no genuine validator
    attestation exists), NOT ok=True/False. Naively summing "v is True"
    would silently report a MEASURED passed=0, which misleadingly implies
    "every milestone was confirmed to fail" -- the honest fact is "no
    verdict exists either way". total must remain genuinely measured (the
    milestones themselves DID run); only passed must be missing."""
    result = COL.milestone_validation(tmp_path, "novalidator", CLI_STDOUT_NOVALIDATOR_SAMPLE, None,
                                      skipped_stage="validate")
    assert result["granularity"] == "real"
    assert result["total"].status == Status.MEASURED
    assert result["total"].value == 2                       # m1, m2 genuinely ran
    assert result["passed"].status == Status.MISSING        # NOT Measurement.ok(0)
    assert result["passed"].value is None


def test_milestone_validation_dispatch_is_data_driven_not_purely_variant_driven(tmp_path: Path):
    """The all-None-passed detection reacts to the PARSED evidence itself
    (every history entry's passed is None), not merely to
    skipped_stage=="validate" -- defensively correct even if a future/older
    artifact shape ever produced an all-None history under a different
    label. Passing skipped_stage=None must not suppress this."""
    result = COL.milestone_validation(tmp_path, "full", CLI_STDOUT_NOVALIDATOR_SAMPLE, None, skipped_stage=None)
    assert result["passed"].status == Status.MISSING
    assert result["total"].value == 2


# --------------------------------------------------------------------------- #
# Post-hoc independent evaluator: translated (self-graded) vs. independently
# VALIDATED developer tests / function validation. Nothing here ever
# touches a real toolchain -- ``SnapshotRunner`` records a snapshot of every
# file (path -> content) under the ephemeral temporary directory an adapter
# builds, captured WHILE that directory is still alive (i.e. from inside the
# injected command-runner callback, before the caller's own
# ``tempfile.TemporaryDirectory()`` context manager tears it down), so a test
# can assert exactly what was and was not copied into it -- e.g. that a
# reference implementation/Cargo manifest was never leaked in alongside the
# reference's own developer tests.
# --------------------------------------------------------------------------- #
class SnapshotRunner:
    def __init__(self, *, stdout: str = "", returncode: int = 0):
        self.stdout = stdout
        self.returncode = returncode
        self.calls: list[dict] = []

    def __call__(self, argv, *, cwd, timeout=None):
        cwd = Path(cwd)
        files: dict[str, str] = {}
        for p in sorted(cwd.rglob("*")):
            if p.is_file():
                files[p.relative_to(cwd).as_posix()] = p.read_text(encoding="utf-8", errors="replace")
        self.calls.append({"argv": list(argv), "cwd": str(cwd), "cwd_files": files})
        return ExecResult(argv=list(argv), returncode=self.returncode, stdout=self.stdout, stderr="",
                          duration_s=0.01, timed_out=False, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(),
                          cwd=str(cwd))


class PerFileRunner:
    """Fake command runner whose ExecResult varies by the LAST argv element
    (the filename) -- needed for SKEL's per-file multi-invocation function-
    harness adapter, where different generated test files must
    independently pass/fail within the SAME evaluate call. Neither
    ``FakeRunner`` (scripts by argv[0]/executable name only) nor
    ``SnapshotRunner`` (one fixed result for every call) can express "file A
    passes, file B fails" within a single invocation of the adapter under
    test, since both return the identical result regardless of which
    specific file is being invoked. Also snapshots ``cwd``'s file listing at
    call time (like ``SnapshotRunner``) so tests can assert exactly which
    files were staged before ANY invocation ran."""
    def __init__(self, *, results: dict[str, ExecResult], default: ExecResult | None = None):
        self.results = results
        self.default = default
        self.calls: list[dict] = []

    def __call__(self, argv, *, cwd, timeout=None):
        cwd = Path(cwd)
        files = sorted(p.name for p in cwd.iterdir() if p.is_file())
        key = argv[-1] if argv else ""
        self.calls.append({"argv": list(argv), "cwd": str(cwd), "cwd_files": files})
        if key in self.results:
            return self.results[key]
        if self.default is not None:
            return self.default
        return ExecResult(argv=list(argv), returncode=0, stdout="", stderr="", duration_s=0.01,
                          timed_out=False, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), cwd=str(cwd))


class CrustCargoRunner:
    """Fake command runner for CRUST end-to-end tests that must distinguish
    a plain ``cargo test`` invocation from a ``cargo run --bin <name>``
    binary-assertion-harness invocation (``crust_run_binary_test_harnesses``)
    -- routed on ``"run" in argv``, NOT on ``argv[-1]`` (unlike
    ``PerFileRunner``), because a real project's binary harness may be
    literally named "test" (e.g. CRUST's own ``libfor``'s
    ``src/bin/test.rs``), which would make ``argv[-1] == "test"`` for BOTH
    the ``cargo test`` call and a ``cargo run --bin test`` call and so
    collide under naive last-token keying."""
    def __init__(self, *, cargo_test_result: ExecResult, binary_results: dict[str, ExecResult] | None = None,
                default_binary_result: ExecResult | None = None):
        self.cargo_test_result = cargo_test_result
        self.binary_results = binary_results or {}
        self.default_binary_result = default_binary_result
        self.calls: list[list[str]] = []

    def __call__(self, argv, *, cwd, timeout=None):
        argv = list(argv)
        self.calls.append(argv)
        if "run" in argv:
            bin_name = argv[-1]
            if bin_name in self.binary_results:
                return self.binary_results[bin_name]
            if self.default_binary_result is not None:
                return self.default_binary_result
            return ExecResult(argv=argv, returncode=0, stdout="", stderr="", duration_s=0.01,
                              timed_out=False, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), cwd=str(cwd))
        return self.cargo_test_result


def _write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")



OXIDIZER_SPEC = {
    "label": "Oxidizer", "target_language": "Rust",
    "build_cmd": ["cargo", "build"], "unit_test_cmd": ["cargo", "test", "--manifest-path", "Cargo.toml"],
}
ALPHATRANS_SPEC = {"label": "AlphaTrans", "target_language": "Python"}


# --------------------------------------------------------------------------- #
# reference_project_dir: case-insensitive resolution under
# <root>/recodeagent_translations/data/tool_projects/{tool}/{project}
# --------------------------------------------------------------------------- #
def test_reference_project_dir_resolves_case_insensitively(tmp_path: Path):
    root = tmp_path / "results"
    proj = root / "recodeagent_translations" / "data" / "tool_projects" / "Oxidizer" / "MyProject"
    proj.mkdir(parents=True)
    assert COL.reference_project_dir(root, "oxidizer", "myproject") == proj


def test_reference_project_dir_none_when_root_missing():
    assert COL.reference_project_dir(None, "oxidizer", "p1") is None


def test_reference_project_dir_none_when_project_missing(tmp_path: Path):
    assert COL.reference_project_dir(tmp_path, "oxidizer", None) is None


def test_reference_project_dir_none_when_tool_dir_absent(tmp_path: Path):
    (tmp_path / "recodeagent_translations" / "data" / "tool_projects").mkdir(parents=True)
    assert COL.reference_project_dir(tmp_path, "oxidizer", "p1") is None


# --------------------------------------------------------------------------- #
# CRUST: contract-path discovery, oracle_integrity (pristine/mutated/
# not_copied), and pristine-scaffold-overlay evaluation
# --------------------------------------------------------------------------- #
def test_crust_contract_relpaths_collects_top_level_files_and_contract_dirs(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "Cargo.toml", "toml")
    _write(scaffold / "Cargo.lock", "lock")
    _write(scaffold / "src" / "bin" / "harness.rs", "harness")
    _write(scaffold / "tests" / "contract_test.rs", "test")
    _write(scaffold / "src" / "lib.rs", "translatable stub -- NOT part of the contract")
    rels = COL.crust_contract_relpaths(scaffold)
    assert set(rels) == {"Cargo.toml", "Cargo.lock", "src/bin/harness.rs", "tests/contract_test.rs"}


def test_crust_contract_relpaths_empty_when_no_contract_files_present(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    scaffold.mkdir()
    assert COL.crust_contract_relpaths(scaffold) == []


def test_crust_oracle_integrity_pristine_when_contract_matches(tmp_path: Path):
    scaffold, target = tmp_path / "scaffold", tmp_path / "target"
    _write(scaffold / "Cargo.toml", "same-content")
    _write(target / "Cargo.toml", "same-content")
    m = COL.crust_oracle_integrity(scaffold, target)
    assert m.value == "pristine"
    assert m.status == Status.MEASURED


def test_crust_oracle_integrity_mutated_when_contract_differs(tmp_path: Path):
    scaffold, target = tmp_path / "scaffold", tmp_path / "target"
    _write(scaffold / "Cargo.toml", "pristine-content")
    _write(target / "Cargo.toml", "agent-edited-content")
    m = COL.crust_oracle_integrity(scaffold, target)
    assert m.value == "mutated"
    assert m.status == Status.MEASURED
    assert "Cargo.toml" in m.reason


def test_crust_oracle_integrity_not_copied_when_target_missing_contract_path(tmp_path: Path):
    scaffold, target = tmp_path / "scaffold", tmp_path / "target"
    _write(scaffold / "Cargo.toml", "content")
    target.mkdir()
    m = COL.crust_oracle_integrity(scaffold, target)
    assert m.value == "not_copied"
    assert m.status == Status.MEASURED


def test_crust_oracle_integrity_not_applicable_when_no_scaffold(tmp_path: Path):
    m = COL.crust_oracle_integrity(tmp_path / "nope", tmp_path / "target")
    assert m.status == Status.NOT_APPLICABLE


def test_crust_oracle_integrity_unavailable_when_scaffold_has_no_contract_paths(tmp_path: Path):
    scaffold, target = tmp_path / "scaffold", tmp_path / "target"
    scaffold.mkdir()
    target.mkdir()
    m = COL.crust_oracle_integrity(scaffold, target)
    assert m.status == Status.UNAVAILABLE


def test_crust_oracle_integrity_missing_when_target_absent(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "Cargo.toml", "content")
    m = COL.crust_oracle_integrity(scaffold, tmp_path / "nope")
    assert m.status == Status.MISSING


def test_crust_validated_tests_eval_uses_pristine_content_despite_mutated_target(tmp_path: Path):
    """Regression: even when the run's OWN target copy of a CRUST contract
    path has been mutated (e.g. the translating agent edited Cargo.toml or
    the compiled test-harness binary), the independent-oracle evaluation
    must still run -- and must run against the PRISTINE scaffold's content
    restored over a TEMPORARY copy of the target, never the run's own
    (possibly-tampered) copy of those same paths."""
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE_CARGO_TOML")
    _write(scaffold / "src" / "bin" / "harness.rs",
          "PRISTINE_HARNESS\n#[test]\nfn t1() {}\n#[test]\nfn t2() {}\n")
    _write(target / "Cargo.toml", "MUTATED_CARGO_TOML")
    _write(target / "src" / "bin" / "harness.rs", "MUTATED_HARNESS")
    _write(target / "src" / "lib.rs", "OWN_TRANSLATION")

    # oracle_integrity must independently see the mutation...
    integrity = COL.crust_oracle_integrity(scaffold, target)
    assert integrity.value == "mutated"

    # ...yet the pristine-overlay evaluation must still run and use scaffold content.
    snapshot = SnapshotRunner(stdout="test result: ok. 2 passed; 0 failed; 0 ignored; 0 measured\n")
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=snapshot)
    assert result["executed"].is_measured
    assert result["executed"].value == 2
    # expected is a static #[test] count over the PRISTINE scaffold's own
    # contract paths -- unaffected by the target's own mutated/different
    # harness.rs content, and stays measured regardless of oracle_integrity.
    assert result["expected"].is_measured
    assert result["expected"].value == 2
    assert result["not_executed"].value == 0
    assert len(snapshot.calls) == 1
    files = snapshot.calls[0]["cwd_files"]
    assert files["Cargo.toml"] == "PRISTINE_CARGO_TOML"          # NOT the mutated target copy
    assert files["src/bin/harness.rs"] == "PRISTINE_HARNESS\n#[test]\nfn t1() {}\n#[test]\nfn t2() {}\n"
    assert files["src/lib.rs"] == "OWN_TRANSLATION"              # non-contract files untouched
    assert snapshot.calls[0]["cwd"] != str(target)               # ran in an EPHEMERAL copy, never run_dir itself
    # And the run's own (mutated) target tree must be left completely alone.
    assert (target / "Cargo.toml").read_text(encoding="utf-8") == "MUTATED_CARGO_TOML"


def test_crust_validated_tests_eval_unavailable_when_scaffold_has_no_contract(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "scaffold").mkdir(parents=True)
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=SnapshotRunner())
    assert all(v.status == Status.UNAVAILABLE for v in result.values())


def test_crust_validated_tests_eval_not_applicable_without_scaffold(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=SnapshotRunner())
    # 7 of the 8 keys are NOT_APPLICABLE (no scaffold at all for this dataset).
    for key in ("expected", "executed", "passed", "failed", "not_executed", "expected_native", "expected_source"):
        assert result[key].status == Status.NOT_APPLICABLE, key
    # expected_paper is a genuinely SEPARATE, optional input (a paper-aligned
    # lookup is conceptually independent of whether a scaffold exists on
    # disk) -- honestly UNAVAILABLE (no --crust-paper-expected-tests was
    # supplied), never silently coerced to match the other 7 keys' status.
    assert result["expected_paper"].status == Status.UNAVAILABLE


# --------------------------------------------------------------------------- #
# CRUST "binary assertion harness" detection (e.g. the real libfor project's
# src/bin/test.rs: a plain fn main() with NO #[test] attribute at all, whose
# own process exit code IS the test verdict -- never discovered/run by plain
# `cargo test`).
# --------------------------------------------------------------------------- #
def test_crust_binary_test_harnesses_empty_when_no_bin_dir(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    scaffold.mkdir()
    assert COL.crust_binary_test_harnesses(scaffold) == []


def test_crust_binary_test_harnesses_detects_file_with_no_test_attribute(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "src" / "bin" / "test.rs", "fn main() { std::process::exit(0); }\n")
    assert COL.crust_binary_test_harnesses(scaffold) == ["test"]


def test_crust_binary_test_harnesses_excludes_file_with_test_attribute(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "src" / "bin" / "harness.rs", "#[test]\nfn a() {}\n")
    assert COL.crust_binary_test_harnesses(scaffold) == []


def test_crust_binary_test_harnesses_mixed_bin_dir_only_returns_no_test_files(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "src" / "bin" / "harness.rs", "#[test]\nfn a() {}\n")
    _write(scaffold / "src" / "bin" / "test.rs", "fn main() {}\n")
    assert COL.crust_binary_test_harnesses(scaffold) == ["test"]


def test_crust_validated_tests_expected_native_includes_binary_harness_count(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "src" / "bin" / "test.rs", "fn main() {}\n")   # binary harness, 0 #[test]
    _write(scaffold / "tests" / "other_test.rs", "#[test]\nfn a() {}\n#[test]\nfn b() {}\n")
    result = COL.crust_validated_tests_expected_native(scaffold)
    assert result.is_measured
    assert result.value == 3   # 2 regex-discovered #[test] + 1 binary harness, never double-counted


# --------------------------------------------------------------------------- #
# crust_run_binary_test_harnesses: actually EXECUTING the detected binary(ies)
# -- exit code 0/nonzero is the verdict, never anything `cargo test` reports.
# --------------------------------------------------------------------------- #
def test_crust_run_binary_test_harnesses_empty_names_is_measured_zero():
    result = COL.crust_run_binary_test_harnesses(Path("/nonexistent"), [], {}, timeout=None, runner=FakeRunner())
    assert result["total"].is_measured and result["total"].value == 0
    assert result["passed"].is_measured and result["passed"].value == 0
    assert result["failed"].is_measured and result["failed"].value == 0


def test_crust_run_binary_test_harnesses_counts_pass_and_fail_by_exit_code(tmp_path: Path):
    runner = PerFileRunner(results={"a": _exec(returncode=0), "b": _exec(returncode=1)})
    result = COL.crust_run_binary_test_harnesses(tmp_path, ["a", "b"], {}, timeout=None, runner=runner)
    assert result["total"].value == 2
    assert result["passed"].value == 1
    assert result["failed"].value == 1
    # invoked via the default `cargo run --quiet --manifest-path Cargo.toml --bin <name>` template
    assert all("run" in call["argv"] for call in runner.calls)
    assert [call["argv"][-1] for call in runner.calls] == ["a", "b"]


def test_crust_run_binary_test_harnesses_uses_dataset_spec_override_template(tmp_path: Path):
    runner = PerFileRunner(results={"a": _exec(returncode=0)})
    spec = {"binary_test_cmd_template": ["custom-runner", "{bin_name}"]}
    result = COL.crust_run_binary_test_harnesses(tmp_path, ["a"], spec, timeout=None, runner=runner)
    assert result["passed"].value == 1
    assert runner.calls[0]["argv"] == ["custom-runner", "a"]


def test_crust_run_binary_test_harnesses_records_reason_on_partial_spawn_failure(tmp_path: Path):
    runner = PerFileRunner(results={
        "a": _exec(returncode=0),
        "b": _exec(timed_out=True, error="timed out after 5.0s"),
    })
    result = COL.crust_run_binary_test_harnesses(tmp_path, ["a", "b"], {}, timeout=5.0, runner=runner)
    assert result["total"].is_measured
    assert result["total"].value == 1          # only "a" produced a real verdict
    assert result["passed"].value == 1
    assert result["failed"].value == 0
    assert "b" in result["total"].reason


def test_crust_run_binary_test_harnesses_error_when_every_binary_fails_to_run(tmp_path: Path):
    runner = PerFileRunner(results={}, default=_exec(timed_out=True, error="timed out"))
    result = COL.crust_run_binary_test_harnesses(tmp_path, ["a", "b"], {}, timeout=1.0, runner=runner)
    assert result["total"].status == Status.ERROR
    assert result["passed"].status == Status.ERROR
    assert result["failed"].status == Status.ERROR
    assert result["total"].value is None   # never a fabricated 0/0


# --------------------------------------------------------------------------- #
# _merge_test_counts: cargo-test result + binary-harness result reconciliation
# --------------------------------------------------------------------------- #
def test_merge_test_counts_sums_when_both_measured():
    a = {"total": Measurement.ok(2), "passed": Measurement.ok(2), "failed": Measurement.ok(0)}
    b = {"total": Measurement.ok(1), "passed": Measurement.ok(0), "failed": Measurement.ok(1)}
    merged = COL._merge_test_counts(a, b)
    assert merged["total"].value == 3
    assert merged["passed"].value == 2
    assert merged["failed"].value == 1
    assert all(m.is_measured for m in merged.values())


def test_merge_test_counts_inherits_a_when_a_not_measured():
    unavailable = Measurement.unavailable("cargo test output unparseable")
    a = {"total": unavailable, "passed": unavailable, "failed": unavailable}
    b = {"total": Measurement.ok(1), "passed": Measurement.ok(1), "failed": Measurement.ok(0)}
    merged = COL._merge_test_counts(a, b)
    # never masks a real cargo-test failure just because the binary-harness
    # portion happened to be a clean "nothing to run" zero.
    assert merged["total"].status == Status.UNAVAILABLE
    assert merged["total"].reason == "cargo test output unparseable"


def test_merge_test_counts_inherits_b_when_only_b_not_measured():
    a = {"total": Measurement.ok(1), "passed": Measurement.ok(1), "failed": Measurement.ok(0)}
    error = Measurement.error("every binary assertion harness failed to run")
    b = {"total": error, "passed": error, "failed": error}
    merged = COL._merge_test_counts(a, b)
    # never masks a binary-harness spawn error just because cargo test itself passed.
    assert merged["total"].status == Status.ERROR
    assert merged["total"].reason == "every binary assertion harness failed to run"


# --------------------------------------------------------------------------- #
# read_crust_paper_expected_tests_{json,csv,xlsx} / dispatch -- the paper's
# AUTHORITATIVE, hand-curated per-project expected-test-count (e.g. from the
# official results.xlsx's own "sweagent crust - tool test" sheet), or an
# explicit reference-inventory file. Structurally separate from the native
# static count above; never silently presented as equal to it.
# --------------------------------------------------------------------------- #
def test_read_crust_paper_expected_tests_json_flat_object(tmp_path: Path):
    path = tmp_path / "ref.json"
    path.write_text(json.dumps({"bitset": 4, "2dpartint": 6}), encoding="utf-8")
    mapping, reason = COL.read_crust_paper_expected_tests_json(path)
    assert mapping == {"bitset": 4, "2dpartint": 6}
    assert str(path) in reason


def test_read_crust_paper_expected_tests_json_none_when_not_object(tmp_path: Path):
    path = tmp_path / "ref.json"
    path.write_text("[1, 2, 3]", encoding="utf-8")
    mapping, reason = COL.read_crust_paper_expected_tests_json(path)
    assert mapping is None
    assert reason


def test_read_crust_paper_expected_tests_json_none_when_malformed(tmp_path: Path):
    path = tmp_path / "ref.json"
    path.write_text("{not valid json", encoding="utf-8")
    mapping, reason = COL.read_crust_paper_expected_tests_json(path)
    assert mapping is None
    assert reason


def test_read_crust_paper_expected_tests_csv_header_columns(tmp_path: Path):
    path = tmp_path / "ref.csv"
    path.write_text("project,expected_tests\nbitset,4\n2dpartint,6\n", encoding="utf-8")
    mapping, reason = COL.read_crust_paper_expected_tests_csv(path)
    assert mapping == {"bitset": 4, "2dpartint": 6}
    assert str(path) in reason


def test_read_crust_paper_expected_tests_csv_positional_fallback_for_two_columns(tmp_path: Path):
    # Header text that matches none of the known project/count candidate
    # names (not even by substring) -- must still work via the 2-column
    # positional (0, 1) fallback.
    path = tmp_path / "ref.csv"
    path.write_text("alpha,beta\nholdem-odds,22\n", encoding="utf-8")
    mapping, reason = COL.read_crust_paper_expected_tests_csv(path)
    assert mapping == {"holdem-odds": 22}
    assert reason


def test_read_crust_paper_expected_tests_csv_none_when_columns_unidentifiable(tmp_path: Path):
    path = tmp_path / "ref.csv"
    path.write_text("alpha,beta,gamma\nholdem-odds,22,x\n", encoding="utf-8")   # 3 columns: no fallback applies
    mapping, reason = COL.read_crust_paper_expected_tests_csv(path)
    assert mapping is None
    assert reason


def test_read_crust_paper_expected_tests_xlsx_reads_named_sheet(tmp_path: Path):
    if C.optional_import("openpyxl") is None:
        pytest.skip("openpyxl not installed in this environment")
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sweagent crust - tool test"
    ws.append(["project", "expected_tests"])
    ws.append(["bitset", 4])
    ws.append(["2dpartint", 6])
    ws.append(["holdem-odds", 22])
    path = tmp_path / "results.xlsx"
    wb.save(path)
    mapping, reason = COL.read_crust_paper_expected_tests_xlsx(path)
    assert mapping == {"bitset": 4, "2dpartint": 6, "holdem-odds": 22}
    assert "sweagent crust - tool test" in reason


def test_read_crust_paper_expected_tests_xlsx_case_insensitive_sheet_name(tmp_path: Path):
    if C.optional_import("openpyxl") is None:
        pytest.skip("openpyxl not installed in this environment")
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SweAgent CRUST -  Tool Test"   # different case/spacing than the constant
    ws.append(["project", "expected_tests"])
    ws.append(["libfor", 1])
    path = tmp_path / "results.xlsx"
    wb.save(path)
    mapping, _ = COL.read_crust_paper_expected_tests_xlsx(path)
    assert mapping == {"libfor": 1}


def test_read_crust_paper_expected_tests_xlsx_missing_sheet_returns_none(tmp_path: Path):
    if C.optional_import("openpyxl") is None:
        pytest.skip("openpyxl not installed in this environment")
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "unrelated sheet"
    ws.append(["a", "b"])
    path = tmp_path / "results.xlsx"
    wb.save(path)
    mapping, reason = COL.read_crust_paper_expected_tests_xlsx(path)
    assert mapping is None
    assert "sweagent crust - tool test" in reason.lower()


def test_read_crust_paper_expected_tests_xlsx_unavailable_without_openpyxl(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(COL.C, "optional_import", lambda name: None)
    mapping, reason = COL.read_crust_paper_expected_tests_xlsx(tmp_path / "whatever.xlsx")
    assert mapping is None
    assert "openpyxl" in reason


def test_read_crust_paper_expected_tests_dispatches_by_extension(tmp_path: Path):
    json_path = tmp_path / "ref.json"
    json_path.write_text(json.dumps({"bitset": 4}), encoding="utf-8")
    mapping, _ = COL.read_crust_paper_expected_tests(json_path)
    assert mapping == {"bitset": 4}

    csv_path = tmp_path / "ref.csv"
    csv_path.write_text("project,expected_tests\nbitset,4\n", encoding="utf-8")
    mapping, _ = COL.read_crust_paper_expected_tests(csv_path)
    assert mapping == {"bitset": 4}


def test_read_crust_paper_expected_tests_missing_path_returns_none(tmp_path: Path):
    mapping, reason = COL.read_crust_paper_expected_tests(tmp_path / "nope.json")
    assert mapping is None
    assert "does not exist" in reason


# --------------------------------------------------------------------------- #
# crust_paper_expected_lookup / crust_combine_expected
# --------------------------------------------------------------------------- #
def test_crust_paper_expected_lookup_unavailable_when_mapping_none():
    assert COL.crust_paper_expected_lookup(None, "bitset").status == Status.UNAVAILABLE


def test_crust_paper_expected_lookup_unavailable_when_mapping_empty():
    assert COL.crust_paper_expected_lookup({}, "bitset").status == Status.UNAVAILABLE


def test_crust_paper_expected_lookup_unavailable_when_project_missing():
    assert COL.crust_paper_expected_lookup({"bitset": 4}, None).status == Status.UNAVAILABLE


def test_crust_paper_expected_lookup_unavailable_when_project_not_found():
    result = COL.crust_paper_expected_lookup({"bitset": 4}, "unknown-project")
    assert result.status == Status.UNAVAILABLE
    assert "unknown-project" in result.reason


def test_crust_paper_expected_lookup_measured_case_and_whitespace_insensitive():
    result = COL.crust_paper_expected_lookup({"2dpartint": 6}, "  2DPartInt  ")
    assert result.is_measured
    assert result.value == 6


def test_crust_combine_expected_prefers_paper_when_measured():
    native = Measurement.ok(8)
    paper = Measurement.ok(6)
    expected, source = COL.crust_combine_expected(native, paper)
    assert expected.value == 6
    assert source.value == "paper"


def test_crust_combine_expected_falls_back_to_native_when_paper_unavailable():
    native = Measurement.ok(8)
    paper = Measurement.unavailable("no --crust-paper-expected-tests mapping was supplied/parsed")
    expected, source = COL.crust_combine_expected(native, paper)
    assert expected.value == 8
    assert source.value == "native"


def test_crust_combine_expected_inherits_native_status_when_neither_measured():
    native = Measurement.unavailable("scaffold has no .rs contract paths to count #[test] attributes in")
    paper = Measurement.unavailable("no --crust-paper-expected-tests mapping was supplied/parsed")
    expected, source = COL.crust_combine_expected(native, paper)
    assert expected.status == Status.UNAVAILABLE
    assert expected.value is None
    assert source.status == Status.UNAVAILABLE
    assert "no --crust-paper-expected-tests mapping was supplied/parsed" in expected.reason


def test_crust_combine_expected_never_claims_native_equals_paper_when_both_measured_and_differ():
    # The real 2dpartint case: paper counts 6, naive scaffold regex counts 8.
    native = Measurement.ok(8)
    paper = Measurement.ok(6)
    expected, _source = COL.crust_combine_expected(native, paper)
    assert expected.value != native.value
    assert expected.value == paper.value


# --------------------------------------------------------------------------- #
# crust_validated_tests_eval: end-to-end native-vs-paper reconciliation and
# binary-assertion-harness execution (the real libfor/2dpartint/holdem-odds
# discrepancies named by the paper).
# --------------------------------------------------------------------------- #
def test_crust_validated_tests_eval_runs_binary_harness_for_libfor_like_project(tmp_path: Path):
    """Regression: the real CRUST libfor project's sole oracle is
    src/bin/test.rs, a binary assertion harness with NO #[test] attribute at
    all -- plain `cargo test` alone would silently report a measured 0 for
    both executed AND expected. After this fix, the binary harness is
    separately detected, counted into expected_native, and actually EXECUTED
    via `cargo run --bin test`, merged into the same executed/passed counts."""
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "src" / "bin" / "test.rs", "fn main() { std::process::exit(0); }\n")
    target.mkdir(parents=True)

    runner = CrustCargoRunner(
        cargo_test_result=_exec(stdout="test result: ok. 0 passed; 0 failed; 0 ignored; 0 measured\n"),
        binary_results={"test": _exec(returncode=0)},
    )
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=runner, project="libfor")

    assert result["expected_native"].is_measured
    assert result["expected_native"].value == 1          # NOT the old silent 0
    assert result["expected"].value == 1                 # no paper mapping supplied -> falls back to native
    assert result["expected_source"].value == "native"
    assert result["executed"].is_measured
    assert result["executed"].value == 1                 # merged: 0 (cargo test) + 1 (binary harness)
    assert result["passed"].value == 1
    assert result["failed"].value == 0
    assert result["not_executed"].value == 0
    # both the plain `cargo test` and the binary harness were actually invoked.
    assert any("run" in call for call in runner.calls)
    assert any("run" not in call for call in runner.calls)


def test_crust_validated_tests_eval_skips_binary_harness_when_cargo_test_not_measured(tmp_path: Path):
    """The binary-harness run must be SKIPPED (never attempted) whenever the
    plain `cargo test` result itself isn't measured (e.g. a build/compile
    failure) -- there is no "clean" build to run a binary against, and
    attempting one would just duplicate the same underlying failure."""
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "src" / "bin" / "test.rs", "fn main() {}\n")
    target.mkdir(parents=True)

    runner = CrustCargoRunner(
        cargo_test_result=_exec(returncode=101, stdout="error[E0433]: failed to resolve\n"),
        binary_results={"test": _exec(returncode=0)},
    )
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=runner, project="libfor")

    assert result["executed"].status == Status.UNAVAILABLE   # unrecognized cargo test output
    assert len(runner.calls) == 1                            # the binary harness was NEVER attempted
    assert "run" not in runner.calls[0]


def test_crust_validated_tests_eval_prefers_paper_aligned_count_over_native(tmp_path: Path):
    """Regression: the real 2dpartint project's scaffold has 8 naive-regex
    #[test] functions but the paper's own bookkeeping counts only 6 -- the
    paper-aligned figure must win and must never be silently presented as
    equal to the native count."""
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "tests" / "a_test.rs", "\n".join(f"#[test]\nfn t{i}() {{}}" for i in range(8)))
    target.mkdir(parents=True)

    runner = SnapshotRunner(stdout="test result: ok. 6 passed; 0 failed; 0 ignored; 0 measured\n")
    result = COL.crust_validated_tests_eval(
        run_dir, CRUST_SPEC, timeout=None, runner=runner, project="2dpartint",
        crust_paper_expected_tests={"2dpartint": 6},
    )
    assert result["expected_native"].value == 8
    assert result["expected_paper"].value == 6
    assert result["expected"].value == 6              # paper wins, never claimed equal to native's 8
    assert result["expected_source"].value == "paper"
    assert result["not_executed"].value == 0          # 6 expected, 6 executed -> nothing missed


def test_crust_validated_tests_eval_falls_back_to_native_when_no_paper_mapping(tmp_path: Path):
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "tests" / "a_test.rs", "\n".join(f"#[test]\nfn t{i}() {{}}" for i in range(8)))
    target.mkdir(parents=True)

    runner = SnapshotRunner(stdout="test result: ok. 8 passed; 0 failed; 0 ignored; 0 measured\n")
    result = COL.crust_validated_tests_eval(run_dir, CRUST_SPEC, timeout=None, runner=runner, project="2dpartint")
    assert result["expected"].value == 8
    assert result["expected_source"].value == "native"
    assert result["expected_paper"].status == Status.UNAVAILABLE


def test_evaluate_independent_oracle_crust_threads_paper_expected_tests(tmp_path: Path):
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "tests" / "a_test.rs", "\n".join(f"#[test]\nfn t{i}() {{}}" for i in range(8)))
    target.mkdir(parents=True)
    manifest_row = {"id": "crust__2dpartint", "tool": "crust", "project": "2dpartint"}
    runner = SnapshotRunner(stdout="test result: ok. 6 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL.evaluate_independent_oracle(
        "crust", run_dir, manifest_row, CRUST_SPEC, None, timeout=None, runner=runner,
        crust_paper_expected_tests={"2dpartint": 6},
    )
    assert result.validated["expected"].value == 6
    assert result.validated["expected_source"].value == "paper"


def test_collect_run_threads_crust_paper_expected_tests(tmp_path: Path):
    run_dir = tmp_path / "run"
    scaffold = run_dir / "scaffold"
    target = run_dir / "pipeline" / "target"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "tests" / "a_test.rs", "\n".join(f"#[test]\nfn t{i}() {{}}" for i in range(8)))
    target.mkdir(parents=True)
    _write_state(run_dir)
    manifest_row = dict(MANIFEST_ROW, project="2dpartint")

    runner = FakeRunner(default_stdout="test result: ok. 6 passed; 0 failed; 0 ignored; 0 measured\n")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__2dpartint", tool="crust", repetition=0,
                         manifest_row=manifest_row, dataset_spec=CRUST_SPEC, runner=runner,
                         crust_paper_expected_tests={"2dpartint": 6})

    assert row["validated_tests_expected"] == 6
    assert row["validated_tests_expected_native"] == 8
    assert row["validated_tests_expected_paper"] == 6
    assert row["validated_tests_expected_source"] == "paper"
    assert row["validated_tests_expected_source_status"] == Status.MEASURED


# --------------------------------------------------------------------------- #
# CLI: --crust-paper-expected-tests
# --------------------------------------------------------------------------- #
def test_build_parser_accepts_crust_paper_expected_tests():
    parser = COL.build_parser()
    ns = parser.parse_args(["--manifest", "m.json", "--runs-root", "r", "--output-root", "o",
                            "--crust-paper-expected-tests", "results.xlsx"])
    assert ns.crust_paper_expected_tests == "results.xlsx"


def test_build_parser_crust_paper_expected_tests_defaults_to_none():
    parser = COL.build_parser()
    ns = parser.parse_args(["--manifest", "m.json", "--runs-root", "r", "--output-root", "o"])
    assert ns.crust_paper_expected_tests is None


def test_cli_main_crust_paper_expected_tests_smoke(tmp_path: Path):
    manifest_path = tmp_path / "manifest.json"
    manifest = _manifest(["crust__a"])
    C.atomic_write_json(manifest_path, manifest)
    ref_path = tmp_path / "ref.json"
    ref_path.write_text(json.dumps({"a": 4}), encoding="utf-8")
    output_root = tmp_path / "out"
    rc = COL.main([
        "--manifest", str(manifest_path), "--runs-root", str(tmp_path / "runs"),
        "--output-root", str(output_root), "--variant", "full", "--repetitions", "1",
        "--crust-paper-expected-tests", str(ref_path),
    ])
    assert rc == 0
    assert (output_root / "raw_runs.csv").exists()


def test_cli_main_warns_on_unparseable_crust_paper_expected_tests(tmp_path: Path, capsys):
    manifest_path = tmp_path / "manifest.json"
    manifest = _manifest(["crust__a"])
    C.atomic_write_json(manifest_path, manifest)
    output_root = tmp_path / "out"
    rc = COL.main([
        "--manifest", str(manifest_path), "--runs-root", str(tmp_path / "runs"),
        "--output-root", str(output_root), "--variant", "full", "--repetitions", "1",
        "--crust-paper-expected-tests", str(tmp_path / "does-not-exist.json"),
    ])
    assert rc == 0
    captured = capsys.readouterr()
    assert "WARNING" in captured.err


# --------------------------------------------------------------------------- #
# Oxidizer: rust/tests/*.rs classification (developer-test oracle vs.
# per-function validation harness vs. excluded "*generated*" build artifacts)
# --------------------------------------------------------------------------- #
def _make_oxidizer_reference(root: Path, project: str) -> Path:
    """Creates .../tool_projects/oxidizer/<project>/rust/{Cargo.toml,src,tests}
    and returns the PROJECT directory (NOT the rust/ subdir) -- exactly the
    shape ``reference_project_dir``/``oxidizer_reference_test_files`` expect
    as their own ``ref_project_dir`` argument (which appends "rust/tests"
    itself). ``foo_test.rs``/``Bar_Test.rs`` each carry one real ``#[test]``
    function (2 total) so ``oxidizer_validated_tests_expected`` has a
    meaningful, non-zero static count to compare against a mocked test run."""
    project_dir = root / "recodeagent_translations" / "data" / "tool_projects" / "oxidizer" / project
    rust_dir = project_dir / "rust"
    _write(rust_dir / "Cargo.toml", "REF_CARGO_TOML")
    _write(rust_dir / "src" / "main.rs", "REF_IMPL")
    _write(
        rust_dir / "tests" / "foo_test.rs",
        "mod test_data;\nORACLE_FOO\n#[test]\nfn foo() {}\n",
    )
    _write(rust_dir / "tests" / "Bar_Test.rs", "ORACLE_BAR\n#[test]\nfn bar() {}\n")   # case-insensitive match
    _write(rust_dir / "tests" / "baz_generated_test.rs", "EXCLUDED_GENERATED_TEST")
    _write(rust_dir / "tests" / "harness_helper.rs", "HARNESS_HELPER")
    _write(rust_dir / "tests" / "test_data.rs", "pub const testData: &[i32] = &[1];")
    _write(rust_dir / "tests" / "mod_generated.rs", "EXCLUDED_GENERATED_MOD")
    return project_dir


def test_oxidizer_reference_test_files_classifies_oracle_vs_harness_vs_excluded(tmp_path: Path):
    ref = _make_oxidizer_reference(tmp_path, "oxi_proj")
    oracle, harness = COL.oxidizer_reference_test_files(ref)
    assert sorted(p.name for p in oracle) == ["Bar_Test.rs", "foo_test.rs"]
    assert sorted(p.name for p in harness) == ["harness_helper.rs"]
    assert [p.name for p in COL.oxidizer_reference_support_files(ref)] == ["test_data.rs"]
    all_names = {p.name for p in oracle} | {p.name for p in harness}
    assert "baz_generated_test.rs" not in all_names   # "generated" excludes it despite "_test.rs" suffix
    assert "mod_generated.rs" not in all_names


def test_oxidizer_reference_test_files_empty_when_ref_dir_none():
    assert COL.oxidizer_reference_test_files(None) == ([], [])


def test_oxidizer_reference_test_files_empty_when_no_tests_dir(tmp_path: Path):
    project_dir = tmp_path / "oxi_proj"
    (project_dir / "rust").mkdir(parents=True)   # rust/ exists but rust/tests/ does not
    assert COL.oxidizer_reference_test_files(project_dir) == ([], [])


# --------------------------------------------------------------------------- #
# Oracle identifier-rewrite (Oxidizer only): CodeWeaver's own idiomatic
# renaming (e.g. exposing ``new_luhn`` for the source language's
# ``NewLuhn``) must not make a PRISTINE reference oracle test's ordinary
# "cannot find function" compile error look like a BEHAVIORAL test failure.
# read_name_mapping / build_identifier_rewrite_index / rust_source_code_mask
# / rewrite_identifiers_with_name_mapping are pure/file-read-only helpers,
# tested directly here with no toolchain/network involved.
# --------------------------------------------------------------------------- #
def test_read_name_mapping_reads_authoritative_plan_json_key(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json",
          json.dumps({"fragments": [], "name_mapping": {"NewLuhn": "new_luhn"}, "skeleton": None}))
    assert COL.read_name_mapping(run_dir) == {"NewLuhn": "new_luhn"}


def test_read_name_mapping_accepts_name_map_fallback_alias(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json", json.dumps({"name_map": {"Foo": "foo"}}))
    assert COL.read_name_mapping(run_dir) == {"Foo": "foo"}


def test_read_name_mapping_prefers_name_mapping_over_name_map_when_both_present(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json",
          json.dumps({"name_mapping": {"A": "a"}, "name_map": {"B": "b"}}))
    assert COL.read_name_mapping(run_dir) == {"A": "a"}


def test_read_name_mapping_empty_when_plan_json_missing(tmp_path: Path):
    assert COL.read_name_mapping(tmp_path / "no-such-run") == {}


def test_read_name_mapping_empty_when_plan_json_malformed(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json", "{not valid json")
    assert COL.read_name_mapping(run_dir) == {}


def test_read_name_mapping_empty_when_top_level_not_a_dict(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json", json.dumps(["not", "a", "dict"]))
    assert COL.read_name_mapping(run_dir) == {}


def test_read_name_mapping_empty_when_name_mapping_key_not_a_dict(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json", json.dumps({"name_mapping": ["A", "a"]}))
    assert COL.read_name_mapping(run_dir) == {}


def test_read_name_mapping_filters_non_string_values(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "plan.json",
          '{"name_mapping": {"Foo": "foo", "Bad": 123, "Null": null, "Nested": {"x": 1}, "Lst": ["a"]}}')
    assert COL.read_name_mapping(run_dir) == {"Foo": "foo"}


def test_build_identifier_rewrite_index_exact_keys_kept_verbatim():
    exact, normalized = COL.build_identifier_rewrite_index({"NewLuhn": "new_luhn", "Foo": "Foo"})
    assert exact == {"NewLuhn": "new_luhn"}   # identity entry ("Foo": "Foo") dropped -- never needs rewriting


def test_build_identifier_rewrite_index_normalized_fallback_added():
    exact, normalized = COL.build_identifier_rewrite_index({"NewLuhn": "new_luhn"})
    assert normalized.get("newluhn") == "new_luhn"


def test_build_identifier_rewrite_index_ambiguous_normalized_collision_dropped():
    # "NewLuhn" and "New_Luhn" both normalize to "newluhn" but map to
    # DIFFERENT targets -- must never guess; drop the normalized entry
    # entirely (exact matches for either source spelling are unaffected).
    exact, normalized = COL.build_identifier_rewrite_index({
        "NewLuhn": "new_luhn", "New_Luhn": "new_luhn_v2",
    })
    assert "newluhn" not in normalized
    assert exact["NewLuhn"] == "new_luhn"
    assert exact["New_Luhn"] == "new_luhn_v2"


def test_build_identifier_rewrite_index_skips_non_string_entries():
    exact, normalized = COL.build_identifier_rewrite_index({"Foo": "foo", "Bad": 123, "AlsoBad": None})
    assert exact == {"Foo": "foo"}


def test_build_identifier_rewrite_index_empty_mapping_yields_empty_indexes():
    exact, normalized = COL.build_identifier_rewrite_index({})
    assert exact == {} and normalized == {}


def _code_positions(text: str, token: str) -> list[bool]:
    """Every mask value at each occurrence of ``token`` in ``text`` (helper
    for the mask tests below -- most fixtures have exactly one occurrence)."""
    mask = COL.rust_source_code_mask(text)
    return [mask[i] for i in range(len(text)) if text.startswith(token, i)]


def test_rust_source_code_mask_plain_call_site_is_code():
    text = "fn f() { NewLuhn(); }"
    assert all(_code_positions(text, "NewLuhn"))


def test_rust_source_code_mask_excludes_line_comment():
    text = "// NewLuhn appears only in this comment\nfn f() {}"
    assert not any(_code_positions(text, "NewLuhn"))


def test_rust_source_code_mask_excludes_block_comment():
    text = "/* NewLuhn */ fn f() {}"
    assert not any(_code_positions(text, "NewLuhn"))


def test_rust_source_code_mask_handles_nested_block_comments():
    text = "/* outer /* NewLuhn */ still-comment */ fn f() {}"
    mask = COL.rust_source_code_mask(text)
    idx = text.index("still-comment")
    assert mask[idx] is False   # still inside the OUTER comment after the inner one closed
    after = text.index("fn f()")
    assert mask[after] is True


def test_rust_source_code_mask_excludes_string_literal():
    text = 'assert_eq!(registry_names(), vec!["NewLuhn"]);'
    assert not any(_code_positions(text, "NewLuhn"))
    # but the real call-site identifiers around it remain code:
    assert all(_code_positions(text, "registry_names"))


def test_rust_source_code_mask_excludes_byte_string_literal():
    text = 'let x = b"NewLuhn";'
    assert not any(_code_positions(text, "NewLuhn"))


def test_rust_source_code_mask_excludes_raw_string_literal_with_hashes():
    text = 'let x = r#"NewLuhn"#; NewLuhn();'
    positions = _code_positions(text, "NewLuhn")
    assert positions == [False, True]   # first occurrence (raw string) excluded, second (call) is code


def test_rust_source_code_mask_excludes_char_literal():
    text = "let c = 'x'; let f = NewLuhn();"
    mask = COL.rust_source_code_mask(text)
    assert mask[text.index("'x'") + 1] is False   # the 'x' char literal's content
    assert all(_code_positions(text, "NewLuhn"))


def test_rust_source_code_mask_excludes_escaped_char_literals():
    text = r"let a = '\n'; let b = '\''; let c = '\x41'; let d = '\u{1F600}';"
    mask = COL.rust_source_code_mask(text)
    # every character strictly between each literal's own delimiting quotes must be masked out
    for token in (r"'\n'", r"'\''", r"'\x41'", r"'\u{1F600}'"):
        start = text.index(token)
        for offset in range(1, len(token) - 1):
            assert mask[start + offset] is False, f"expected False inside {token!r}"


def test_rust_source_code_mask_excludes_lifetime_entirely():
    """Regression guard for a real bug found while implementing this
    feature: a lifetime's NAME (e.g. the 'a' in impl<'a>) sits in an
    identifier-shaped position immediately after an apostrophe. If a
    name_mapping entry's SOURCE key happens to equal a common lifetime name
    (e.g. a single-letter symbol), the mask must still exclude the ENTIRE
    lifetime (apostrophe + name), never just the apostrophe, or every
    ``'a`` in the file would be corrupted by rewrite_identifiers_with_name_
    mapping."""
    text = "impl<'a> Foo<'a> for Bar<'a> { fn f(&'a self) -> &'a str { \"x\" } }"
    mask = COL.rust_source_code_mask(text)
    for i, ch in enumerate(text):
        if ch == "'":
            # the apostrophe itself, and the 'a' immediately after it, are BOTH excluded
            assert mask[i] is False
            assert mask[i + 1] is False


def test_rust_source_code_mask_never_raises_on_unterminated_string():
    text = 'let x = "unterminated'
    mask = COL.rust_source_code_mask(text)   # must not raise
    assert len(mask) == len(text)
    assert all(v is False for v in mask[text.index('"'):])   # safe direction: masks out the remainder


def test_rewrite_identifiers_with_name_mapping_rewrites_call_site():
    text = 'fn t() { assert_eq!(NewLuhn().checksum("1"), true); }'
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"NewLuhn": "new_luhn"})
    assert "new_luhn()" in new_text
    assert applied == ["NewLuhn"]


def test_rewrite_identifiers_with_name_mapping_never_touches_string_literal_or_comment():
    text = ('fn t() {\n'
           '    // NewLuhn is mentioned here too\n'
           '    assert_eq!(NewLuhn(), true);\n'
           '    assert_eq!(registry_names(), vec!["NewLuhn"]);\n'
           '}\n')
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"NewLuhn": "new_luhn"})
    assert applied == ["NewLuhn"]
    assert "new_luhn()" in new_text                       # the real call site WAS rewritten
    assert '// NewLuhn is mentioned here too' in new_text  # the comment was NOT touched
    assert 'vec!["NewLuhn"]' in new_text                   # the string literal was NOT touched


def test_rewrite_identifiers_with_name_mapping_never_corrupts_a_colliding_lifetime():
    text = "impl<'a> Foo<'a> { fn f(&'a self) -> &'a str { a() } }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"a": "z"})
    assert "'a" in new_text and "'z" not in new_text   # every lifetime survives untouched
    assert "z()" in new_text                            # only the genuine call-site identifier is rewritten
    assert applied == ["a"]


def test_rewrite_identifiers_with_name_mapping_never_touches_char_literal_content():
    text = "fn f() { let c = 'x'; y(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"x": "w"})
    assert "'x'" in new_text            # the char literal survives untouched
    assert applied == []                # 'x' never occurs as a genuine identifier token in this text


def test_rewrite_identifiers_with_name_mapping_normalized_fallback_case_insensitive():
    text = "fn f() { NEW_LUHN(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"NewLuhn": "new_luhn"})
    assert "new_luhn()" in new_text
    assert applied == ["NEW_LUHN"]


def test_rewrite_identifiers_with_name_mapping_protects_fixture_symbols_and_use_lines():
    text = (
        "use test_data::testData;\n"
        "fn f() { for value in testData { NewLuhn(); } }\n"
    )
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(
        text,
        {"test_data": "TEST_DATA", "testData": "TEST_DATA", "NewLuhn": "new_luhn"},
        protected_identifiers={"test_data", "testData"},
    )
    assert "use test_data::testData;" in new_text
    assert "value in testData" in new_text
    assert "new_luhn()" in new_text
    assert applied == ["NewLuhn"]


def test_rewrite_identifiers_with_name_mapping_ignores_non_rust_target_description():
    text = "fn f() { value.String(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(
        text, {"String": "Display::fmt (or to_string)"},
    )
    assert new_text == text
    assert applied == []


def test_rewrite_rust_use_paths_imports_owner_for_associated_function():
    text = "use gohistogram::{NewHistogram, NewWeightedHistogram};\n"
    rewritten = COL.rewrite_rust_use_paths(
        text,
        {
            "NewHistogram": "NumericHistogram::new",
            "NewWeightedHistogram": "WeightedHistogram::new",
        },
        {"NumericHistogram": "", "WeightedHistogram": ""},
    )
    assert rewritten == (
        "use gohistogram::NumericHistogram;\n"
        "use gohistogram::WeightedHistogram;\n"
    )


def test_rewrite_rust_use_paths_honors_public_reexport_for_identity_name():
    rewritten = COL.rewrite_rust_use_paths(
        "use textrank::rank::Rank;\n",
        {"Rank": "Rank"},
        {"Rank": ""},
    )
    assert rewritten == "use textrank::Rank;\n"


def test_add_rust_trait_imports_adds_unambiguous_method_trait():
    text = "use gohistogram::NumericHistogram;\nfn f() { value.add(1.0); }\n"
    rewritten = COL.add_rust_trait_imports(
        text,
        {"Add": "add"},
        ["Add"],
        {"add": "Histogram"},
    )
    assert rewritten.startswith("use gohistogram::Histogram;\n")


def test_add_rust_trait_imports_does_not_duplicate_braced_import():
    text = "use textrank::algorithm::{Algorithm, AlgorithmDefault};\n"
    rewritten = COL.add_rust_trait_imports(
        text,
        {"Populate": "populate"},
        ["Populate"],
        {"populate": "Algorithm"},
    )
    assert rewritten == text


def test_rewrite_rust_field_accesses_is_context_specific():
    text = (
        "let Word = value.Word; let type_name = Word::default(); "
        "value.Word();"
    )
    rewritten, applied = COL.rewrite_rust_field_accesses(
        text, {"Word": "word"},
    )
    assert rewritten == (
        "let Word = value.word; let type_name = Word::default(); "
        "value.Word();"
    )
    assert applied == ["Word"]


def test_derive_rust_field_mapping_uses_public_target_fields(tmp_path: Path):
    target = tmp_path / "target"
    oracle = tmp_path / "oracle.rs"
    _write(
        target / "src" / "lib.rs",
        "pub struct SingleWord { pub word: String }\n",
    )
    _write(oracle, "fn check(x: SingleWord) { assert_eq!(x.Word, \"x\"); }\n")
    assert COL.derive_rust_field_mapping(target, [oracle]) == {"Word": "word"}


def test_rewrite_identifiers_with_name_mapping_noop_when_mapping_empty():
    text = "fn t() { NewLuhn(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {})
    assert new_text == text
    assert applied == []


def test_rewrite_identifiers_with_name_mapping_noop_when_no_token_matches():
    text = "fn t() { unrelated_call(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"NewLuhn": "new_luhn"})
    assert new_text == text
    assert applied == []


def test_rewrite_identifiers_with_name_mapping_returns_sorted_applied_list():
    text = "fn t() { Alpha(); Beta(); }"
    new_text, applied = COL.rewrite_identifiers_with_name_mapping(text, {"Beta": "beta", "Alpha": "alpha"})
    assert applied == ["Alpha", "Beta"]   # sorted, not insertion/mapping order
    assert "alpha();" in new_text and "beta();" in new_text


# --------------------------------------------------------------------------- #
# AlphaTrans: verified_test/ discovery
# --------------------------------------------------------------------------- #
def _make_alphatrans_reference(root: Path, project: str) -> Path:
    ref = root / "recodeagent_translations" / "data" / "tool_projects" / "alphatrans" / project
    _write(ref / "python" / "impl.py", "REF_IMPL")
    _write(ref / "verified_test" / "test_foo.py", "def test_foo():\n    assert True\n")
    _write(ref / "verified_test" / "conftest.py", "REF_CONFTEST")
    return ref


def test_alphatrans_verified_test_dir_found(tmp_path: Path):
    ref = _make_alphatrans_reference(tmp_path, "alpha_proj")
    assert COL.alphatrans_verified_test_dir(ref) == ref / "verified_test"


def test_alphatrans_verified_test_dir_none_when_absent(tmp_path: Path):
    ref = tmp_path / "alpha_proj"
    ref.mkdir()
    assert COL.alphatrans_verified_test_dir(ref) is None


def test_alphatrans_verified_test_dir_none_when_ref_dir_none():
    assert COL.alphatrans_verified_test_dir(None) is None


# --------------------------------------------------------------------------- #
# AlphaTrans: agent_test/ GENERATED function-harness discovery (structurally
# separate from verified_test/ above -- see collect.py's "POST-HOC
# INDEPENDENT EVALUATOR" docstring section).
# --------------------------------------------------------------------------- #
def _make_alphatrans_agent_test(ref_project_dir: Path, *, nested_python_subdir: bool) -> None:
    """Adds an ``agent_test/`` GENERATED function-harness fixture under an
    already-created AlphaTrans reference project dir (see
    ``_make_alphatrans_reference``), matching the two real shapes verified
    directly against the official RESULTS artifact: NESTED
    (commons-cli/commons-csv/commons-validator: a ``python/`` subdir
    ALONGSIDE a SIBLING ``resources/``, plus a top-level ``__init__.py``) or
    FLAT (commons-fileupload: no ``python/`` subdir, no ``resources/`` at
    all). Both shapes mix "generated" files (the ones this harness must
    select) with plain, non-generated ``XxxTest.py`` files (the official
    system's own translated developer tests, which must be excluded)."""
    agent_test = ref_project_dir / "agent_test"
    _write(agent_test / "__init__.py", "")
    if nested_python_subdir:
        base = agent_test / "python"
        _write(base / "__init__.py", "")
        _write(base / "conftest.py", "AGENT_TEST_CONFTEST")
        _write(base / "org" / "apache" / "commons" / "OptionBuilderTest_generated.py", "GENERATED_SUFFIX_STYLE")
        _write(base / "org" / "apache" / "commons" / "OptionBuilderGeneratedTest.py", "GENERATED_EMBEDDED_STYLE")
        _write(base / "org" / "apache" / "commons" / "OptionBuilderTest.py", "OFFICIAL_NON_GENERATED")
        _write(agent_test / "resources" / "org" / "apache" / "commons" / "data.txt", "FIXTURE_DATA")
    else:
        _write(agent_test / "org" / "apache" / "commons" / "fileupload" / "FooTest_generated.py",
              "GENERATED_SUFFIX_STYLE")
        _write(agent_test / "org" / "apache" / "commons" / "fileupload" / "FooTest.py", "OFFICIAL_NON_GENERATED")


def test_alphatrans_agent_test_dir_found(tmp_path: Path):
    ref = _make_alphatrans_reference(tmp_path, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=True)
    assert COL.alphatrans_agent_test_dir(ref) == ref / "agent_test"


def test_alphatrans_agent_test_dir_none_when_absent(tmp_path: Path):
    ref = tmp_path / "alpha_proj"
    ref.mkdir()
    assert COL.alphatrans_agent_test_dir(ref) is None


def test_alphatrans_agent_test_dir_none_when_ref_dir_none():
    assert COL.alphatrans_agent_test_dir(None) is None


def test_alphatrans_function_harness_files_nested_python_subdir_selects_generated_and_support(tmp_path: Path):
    ref = _make_alphatrans_reference(tmp_path, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=True)
    agent_test_dir = COL.alphatrans_agent_test_dir(ref)
    files = COL.alphatrans_function_harness_files(agent_test_dir)
    rels = sorted(p.relative_to(agent_test_dir).as_posix() for p in files)
    assert "__init__.py" in rels
    assert "python/__init__.py" in rels
    assert "python/conftest.py" in rels
    assert "python/org/apache/commons/OptionBuilderTest_generated.py" in rels
    assert "python/org/apache/commons/OptionBuilderGeneratedTest.py" in rels
    assert "resources/org/apache/commons/data.txt" in rels
    # the official system's OWN plain, non-generated translated test -- a DIFFERENT
    # metric from the generated function-harness this adapter measures -- excluded:
    assert "python/org/apache/commons/OptionBuilderTest.py" not in rels


def test_alphatrans_function_harness_files_flat_layout_selects_generated_excludes_official(tmp_path: Path):
    ref = _make_alphatrans_reference(tmp_path, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=False)
    agent_test_dir = COL.alphatrans_agent_test_dir(ref)
    files = COL.alphatrans_function_harness_files(agent_test_dir)
    rels = sorted(p.relative_to(agent_test_dir).as_posix() for p in files)
    assert "org/apache/commons/fileupload/FooTest_generated.py" in rels
    assert "org/apache/commons/fileupload/FooTest.py" not in rels


def test_alphatrans_function_harness_files_empty_when_agent_test_dir_none():
    assert COL.alphatrans_function_harness_files(None) == []


def test_alphatrans_function_harness_eval_copies_relative_structure_and_excludes_official_impl(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    ref = _make_alphatrans_reference(ref_root, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=True)
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "impl.py", "OWN_IMPL")
    _write(target / "agent_test" / "stale.py", "OWN_STALE")   # must be wiped, not merged with the reference copy
    snapshot = SnapshotRunner(stdout="4 passed in 0.02s\n")

    result = COL.alphatrans_function_harness_eval(target, ref, timeout=None, runner=snapshot)

    assert result["total"].is_measured
    assert result["total"].value == 4
    assert len(snapshot.calls) == 1
    assert snapshot.calls[0]["argv"] == list(COL.ALPHATRANS_FUNCTION_HARNESS_TEST_CMD)
    files = snapshot.calls[0]["cwd_files"]
    assert files["impl.py"] == "OWN_IMPL"                                 # CodeWeaver's own impl, untouched
    assert "agent_test/python/impl.py" not in files                       # reference's PRODUCTION impl NEVER copied
    assert "python/impl.py" not in files
    assert files["agent_test/python/org/apache/commons/OptionBuilderTest_generated.py"] == "GENERATED_SUFFIX_STYLE"
    assert files["agent_test/python/org/apache/commons/OptionBuilderGeneratedTest.py"] == "GENERATED_EMBEDDED_STYLE"
    assert files["agent_test/resources/org/apache/commons/data.txt"] == "FIXTURE_DATA"
    assert files["agent_test/python/conftest.py"] == "AGENT_TEST_CONFTEST"
    assert files["agent_test/__init__.py"] == ""
    # official, non-generated translated test -- excluded (a different metric):
    assert "agent_test/python/org/apache/commons/OptionBuilderTest.py" not in files
    assert "agent_test/stale.py" not in files                             # wiped, never merged


def test_alphatrans_function_harness_eval_unavailable_when_no_agent_test_dir(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    ref = _make_alphatrans_reference(ref_root, "alpha_proj")   # no agent_test/ built at all
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "impl.py", "OWN_IMPL")
    snapshot = SnapshotRunner()
    result = COL.alphatrans_function_harness_eval(target, ref, timeout=None, runner=snapshot)
    assert result["total"].status == Status.UNAVAILABLE
    assert result["total"].value is None
    assert snapshot.calls == []


def test_alphatrans_function_harness_eval_unavailable_when_ref_project_dir_none(tmp_path: Path):
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "impl.py", "OWN_IMPL")
    snapshot = SnapshotRunner()
    result = COL.alphatrans_function_harness_eval(target, None, timeout=None, runner=snapshot)
    assert result["total"].status == Status.UNAVAILABLE
    assert snapshot.calls == []


# --------------------------------------------------------------------------- #
# SKEL: javascript/*generated*.js GENERATED function-harness discovery and
# per-file, exit-code-based execution (SKEL has no independent DEVELOPER-test
# oracle at all -- see the always-unavailable dispatch test below -- but IS
# newly given this separate function_harness_tests_* execution evidence).
# --------------------------------------------------------------------------- #
def _make_skel_reference(root: Path, project: str) -> Path:
    ref = root / "recodeagent_translations" / "data" / "tool_projects" / "skel" / project
    _write(ref / "javascript" / "source.js", "// REFERENCE PRODUCTION IMPLEMENTATION")
    _write(ref / "javascript" / "tracer_skip.js", "// REFERENCE-ONLY INTERNAL HELPER")
    _write(ref / "javascript" / "SKELTest_generated.js", "GENERATED_MAIN")
    _write(ref / "javascript" / "FooFunctionsTest_generated.js", "GENERATED_EXTRA")
    return ref


def test_skel_reference_javascript_dir_found(tmp_path: Path):
    ref = _make_skel_reference(tmp_path, "bst")
    assert COL.skel_reference_javascript_dir(ref) == ref / "javascript"


def test_skel_reference_javascript_dir_none_when_absent(tmp_path: Path):
    ref = tmp_path / "bst"
    ref.mkdir()
    assert COL.skel_reference_javascript_dir(ref) is None


def test_skel_reference_javascript_dir_none_when_ref_dir_none():
    assert COL.skel_reference_javascript_dir(None) is None


def test_skel_function_harness_files_selects_generated_excludes_source_and_helpers(tmp_path: Path):
    ref = _make_skel_reference(tmp_path, "bst")
    javascript_dir = COL.skel_reference_javascript_dir(ref)
    files = COL.skel_function_harness_files(javascript_dir)
    names = sorted(p.name for p in files)
    assert names == ["FooFunctionsTest_generated.js", "SKELTest_generated.js"]
    assert "source.js" not in names           # reference production implementation -- NEVER selected
    assert "tracer_skip.js" not in names      # reference-only internal helper -- NEVER selected


def test_skel_function_harness_files_empty_when_javascript_dir_none():
    assert COL.skel_function_harness_files(None) == []


def test_skel_function_harness_eval_aliases_entry_file_and_aggregates_per_file_exit_codes(tmp_path: Path):
    ref = _make_skel_reference(tmp_path, "bst")
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", "// CODEWEAVER OWN TRANSLATION")
    runner = PerFileRunner(results={
        "SKELTest_generated.js": _exec(returncode=0),
        "FooFunctionsTest_generated.js": _exec(returncode=1, stderr="AssertionError"),
    })

    result = COL.skel_function_harness_eval(target, ref, timeout=None, runner=runner)

    assert result["total"].is_measured
    assert result["total"].value == 2          # a FILE count, not an assertion count
    assert result["passed"].value == 1
    assert result["failed"].value == 1
    assert len(runner.calls) == 2               # one node invocation PER selected file
    argvs = sorted(c["argv"] for c in runner.calls)
    assert argvs == [["node", "FooFunctionsTest_generated.js"], ["node", "SKELTest_generated.js"]]
    files_at_call_time = runner.calls[0]["cwd_files"]
    assert "index.js" in files_at_call_time               # CodeWeaver's own entry file, still present
    assert "source.js" in files_at_call_time              # ADDITIONAL alias of index.js, now present too
    assert "SKELTest_generated.js" in files_at_call_time
    assert "FooFunctionsTest_generated.js" in files_at_call_time
    assert "tracer_skip.js" not in files_at_call_time     # reference-only internal helper NEVER copied


def test_skel_function_harness_eval_entry_alias_is_a_copy_never_a_rename(tmp_path: Path):
    """CodeWeaver's own index.js must never be renamed/removed -- only
    ADDITIONALLY copied to source.js -- and only inside the temporary
    evaluation copy; the original target tree is never touched at all."""
    ref = _make_skel_reference(tmp_path, "bst")
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", "// CODEWEAVER OWN TRANSLATION")
    runner = PerFileRunner(results={}, default=_exec(returncode=0))
    COL.skel_function_harness_eval(target, ref, timeout=None, runner=runner)
    assert (target / "index.js").read_text(encoding="utf-8") == "// CODEWEAVER OWN TRANSLATION"
    assert not (target / "source.js").exists()


def test_skel_function_harness_removes_inline_reference_implementation(tmp_path: Path):
    ref = tmp_path / "refroot" / "recodeagent_translations" / "data" / "tool_projects" / "skel" / "colorsys"
    _write(ref / "test_comparison_report.json", "{}")
    generated = ref / "javascript" / "_vTest_generated.js"
    _write(generated, """
function _v() { return 1; }
if (_v() !== 2) {
    process.exit(1);
}
""")
    _write(ref / "javascript" / "SkelHeadTest_generated.js", """
function user_get_type() { return "reference"; }
function user_check_type() { return false; }
function SkelClass() { return {source: "reference"}; }
if (user_get_type() !== "target" || !user_check_type() || SkelClass().source !== "target") {
    process.exit(1);
}
""")
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", """
function _v() { return 2; }
function user_get_type() { return "target"; }
function user_check_type() { return true; }
function SkelClass() { return {source: "target"}; }
""")

    result = COL.skel_function_harness_eval(target, ref, timeout=30)

    assert result["total"].value == 46
    assert result["passed"].value == 46
    assert result["failed"].value == 0
    # The official artifact remains immutable; rewriting happens only in the
    # temporary evaluation copy.
    assert "return 1" in generated.read_text(encoding="utf-8")


def test_skel_function_harness_case_counts_unavailable_after_script_abort(tmp_path: Path):
    ref = tmp_path / "refroot" / "recodeagent_translations" / "data" / "tool_projects" / "skel" / "colorsys"
    _write(ref / "test_comparison_report.json", "{}")
    _write(ref / "javascript" / "_vTest_generated.js", """
function _v() { return 1; }
if (_v() !== 2) {
    process.exit(1);
}
""")
    _write(ref / "javascript" / "SkelHeadTest_generated.js", """
function user_get_type() { return "reference"; }
function user_check_type() { return false; }
function SkelClass() { return {source: "reference"}; }
""")
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", """
function _v() { return 3; }
function user_get_type() { return "target"; }
function user_check_type() { return true; }
function SkelClass() { return {source: "target"}; }
""")

    result = COL.skel_function_harness_eval(target, ref, timeout=30)

    assert result["total"].status == Status.UNAVAILABLE
    assert result["passed"].status == Status.UNAVAILABLE
    assert "fixed 46-case inventory" in result["total"].reason


def test_skel_function_harness_eval_unavailable_when_no_generated_files_in_reference(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    ref = ref_root / "recodeagent_translations" / "data" / "tool_projects" / "skel" / "bst"
    _write(ref / "javascript" / "source.js", "// REFERENCE IMPL")   # exists, but ships no *generated* files
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", "// OWN")
    runner = PerFileRunner(results={})
    result = COL.skel_function_harness_eval(target, ref, timeout=None, runner=runner)
    assert result["total"].status == Status.UNAVAILABLE
    assert result["total"].value is None
    assert runner.calls == []


def test_skel_function_harness_eval_unavailable_when_ref_project_dir_none(tmp_path: Path):
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", "// OWN")
    runner = PerFileRunner(results={})
    result = COL.skel_function_harness_eval(target, None, timeout=None, runner=runner)
    assert result["total"].status == Status.UNAVAILABLE
    assert runner.calls == []


# --------------------------------------------------------------------------- #
# SKEL: AST-extracted independent VALIDATED developer tests
# (test_name_mapping.csv's "verified test" column + javascript/source.js,
# which embeds BOTH the reference implementation AND its own translated
# tests together -- unlike CRUST/Oxidizer/AlphaTrans, SKEL ships no separate
# oracle FILE TREE at all). Low-level AST-walker unit tests below need a
# REAL tree-sitter-javascript parser (there is no reasonable way to fake an
# AST without reimplementing a JS parser) and are skipped -- never failed --
# when tree-sitter/tree-sitter-javascript are not installed, mirroring how
# this codebase already treats every other optional dependency (e.g. Qwen
# embeddings in test_test_compare.py, which are exercised only via a fully
# fake module). The tree-sitter-UNAVAILABLE degradation path itself, and
# skel_build_validated_harness_source (a pure string-assembly function),
# always run regardless of environment.
# --------------------------------------------------------------------------- #
SKEL_CSV_HEADER = ("project,python test path,python test name,javascript test path,"
                  "javascript test name,verified test,agent test,mismatch")

# A small, realistic, self-contained source.js: one exported production
# function (`add`), one private/non-exported helper (`_sub`), and several
# verified-test-shaped functions exercising every extraction/blocking rule
# this adapter implements (safe builtins, exported-target-identifier
# resolution, private-helper blocking, the shorthand-property-identifier
# regression, and a safe Node-core `require`).
SKEL_SAMPLE_SOURCE_JS = """\
function add(a, b) {
  return a + b;
}

function _sub(a, b) {
  return a - b;
}

function test_trivial() {
  const x = 1 + 1;
  console.assert(x === 2);
  return true;
}

function test_uses_export() {
  return add(2, 3) === 5;
}

function test_uses_private_helper() {
  return _sub(5, 2) === 3;
}

function test_shorthand() {
  const bar = 42;
  return { add, bar };
}

const assert = require('assert');
function test_with_assert() {
  assert(1 + 1 === 2);
  return true;
}

module.exports = { add };
"""


def _write_skel_source_js(root: Path, text: str = SKEL_SAMPLE_SOURCE_JS) -> Path:
    path = root / "javascript" / "source.js"
    _write(path, text)
    return path


def _make_skel_reference_with_csv(root: Path, project: str, *, verified_js_names: list[str],
                                  all_js_names: list[str] | None = None,
                                  source_text: str = SKEL_SAMPLE_SOURCE_JS) -> Path:
    """Builds a real ``<root>/recodeagent_translations/data/tool_projects/
    skel/<project>/`` reference tree: a genuine ``test_name_mapping.csv``
    (one row per name in ``all_js_names`` -- defaulting to just
    ``verified_js_names`` -- with ``verified test`` = "1" iff the name is
    also in ``verified_js_names``) plus the given ``javascript/source.js``
    text."""
    ref = root / "recodeagent_translations" / "data" / "tool_projects" / "skel" / project
    names = all_js_names if all_js_names is not None else verified_js_names
    lines = [SKEL_CSV_HEADER]
    for n in names:
        verified = "1" if n in verified_js_names else "0"
        lines.append(f"{project},p/{n}.py,{n},j/{n}.js,{n},{verified},1,0")
    _write(ref / "test_name_mapping.csv", "\n".join(lines) + "\n")
    _write(ref / "javascript" / "source.js", source_text)
    return ref


def _skel_parse_function(js_snippet: str):
    """Parses ``js_snippet`` (expected to be a single top-level function
    declaration) with a REAL tree-sitter parser and returns
    ``(function_node, src_bytes)``. Skips (never fails) the calling test if
    tree-sitter/tree-sitter-javascript are not installed in this
    environment."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = js_snippet.encode("utf-8")
    tree = parser.parse(src)
    return tree.root_node.children[0], src


def _skel_parse_top_level_const_value(js_snippet: str):
    """Parses ``js_snippet`` (expected to be exactly ``const NAME =
    <expr>;``) with a REAL tree-sitter parser and returns just the VALUE
    expression's own AST node (``_skel_is_pure_literal_expression`` only
    ever inspects node types/structure, never raw source text, so no
    ``src`` bytes are needed by callers). Skips (never fails) the calling
    test if tree-sitter/tree-sitter-javascript are not installed."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = js_snippet.encode("utf-8")
    tree = parser.parse(src)
    stmt = tree.root_node.children[0]
    declarator = next(c for c in stmt.children if c.type == "variable_declarator")
    return declarator.child_by_field_name("value")


# --- _skel_module_exports_names: object-literal AND member-assignment --- #
# --- export forms (resolution rules (c)/(d) both rely on this). --- #
def test_skel_module_exports_names_recognizes_member_assignment_forms(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = (
        "function add(a, b) { return a + b; }\n"
        "function sub(a, b) { return a - b; }\n"
        "module.exports.add = add;\n"
        "exports.sub = sub;\n"
    ).encode("utf-8")
    tree = parser.parse(src)
    assert COL._skel_module_exports_names(tree.root_node, src) == {"add", "sub"}


def test_skel_module_exports_names_object_literal_and_member_assignment_combine(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = (
        "function add(a, b) { return a + b; }\n"
        "function sub(a, b) { return a - b; }\n"
        "module.exports = { add };\n"
        "module.exports.sub = sub;\n"
    ).encode("utf-8")
    tree = parser.parse(src)
    assert COL._skel_module_exports_names(tree.root_node, src) == {"add", "sub"}


# --- _skel_read_module_export_names: best-effort export listing of ANY --- #
# --- JS file -- used to read CodeWeaver's OWN target entry file. --- #
def test_skel_read_module_export_names_empty_when_file_missing(tmp_path: Path):
    assert COL._skel_read_module_export_names(tmp_path / "index.js") == frozenset()


def test_skel_read_module_export_names_empty_when_tree_sitter_unavailable(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(COL.C, "optional_import", lambda name: None)
    _write(tmp_path / "index.js", "module.exports = { add: (a, b) => a + b };")
    assert COL._skel_read_module_export_names(tmp_path / "index.js") == frozenset()


def test_skel_read_module_export_names_reads_object_literal_exports(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    path = tmp_path / "index.js"
    _write(path, "function heapifyMax(a) { return a; }\nmodule.exports = { heapifyMax };\n")
    assert COL._skel_read_module_export_names(path) == frozenset({"heapifyMax"})


def test_skel_read_module_export_names_reads_member_assignment_exports(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    path = tmp_path / "index.js"
    _write(path, "function heapifyMax(a) { return a; }\nmodule.exports.heapifyMax = heapifyMax;\n")
    assert COL._skel_read_module_export_names(path) == frozenset({"heapifyMax"})


def test_skel_read_module_export_names_never_crashes_on_malformed_js(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    path = tmp_path / "index.js"
    _write(path, "this is not { valid javascript at ][ all (((")
    assert COL._skel_read_module_export_names(path) == frozenset()


# --- _skel_is_pure_literal_expression: the safety gate for resolution --- #
# --- rule (e) ("literal-expectation" constants). --- #
@pytest.mark.parametrize("snippet", [
    "const x = 1;",
    "const x = -5;",
    "const x = +5;",
    "const x = 'hi';",
    "const x = true;",
    "const x = false;",
    "const x = null;",
    "const x = `hi`;",
    "const x = [1, 2, [3, 4]];",
    "const x = { a: 1, b: [2, 3] };",
    "const x = {};",
    "const x = [];",
    "const x = [1, , 3];",
])
def test_skel_is_pure_literal_expression_true_for_pure_data(snippet: str):
    node = _skel_parse_top_level_const_value(snippet)
    assert COL._skel_is_pure_literal_expression(node) is True, snippet


@pytest.mark.parametrize("snippet", [
    "const x = someName;",
    "const x = someCall();",
    "const x = `hi ${name}`;",
    "const x = { a: someCall() };",
    "const x = { ...spreadMe };",
    "const x = { foo() { return 1; } };",
    "const x = { [computedKey]: 1 };",
    "const x = { bar };",
    "const x = [someName];",
    "const x = [...spreadArr];",
])
def test_skel_is_pure_literal_expression_false_for_anything_referencing_outside_names(snippet: str):
    node = _skel_parse_top_level_const_value(snippet)
    assert COL._skel_is_pure_literal_expression(node) is False, snippet


# --- _skel_top_level_literal_declarations: safe splice-candidate lookup --- #
def test_skel_top_level_literal_declarations_single_declarator_pure(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = "const EXPECTED = [1, 2, 3];\nfunction noop() {}\n".encode("utf-8")
    tree = parser.parse(src)
    out = COL._skel_top_level_literal_declarations(tree.root_node, src)
    assert set(out) == {"EXPECTED"}
    _decl, stmt = out["EXPECTED"]
    assert src[stmt.start_byte:stmt.end_byte].decode("utf-8") == "const EXPECTED = [1, 2, 3];"


def test_skel_top_level_literal_declarations_excludes_multi_declarator_statement_entirely(tmp_path: Path):
    """A pure literal sharing a statement with a NON-literal sibling
    declarator (e.g. ``const a = 1, b = someExternalCall();``) must be
    excluded ENTIRELY -- copying the whole statement's verbatim text for
    ``a`` would otherwise also copy ``b``'s own non-pure initializer."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = "const a = 1, b = someExternalCall();\n".encode("utf-8")
    tree = parser.parse(src)
    out = COL._skel_top_level_literal_declarations(tree.root_node, src)
    assert out == {}


def test_skel_top_level_literal_declarations_excludes_non_literal_declarator(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    src = "const x = someCall();\n".encode("utf-8")
    tree = parser.parse(src)
    out = COL._skel_top_level_literal_declarations(tree.root_node, src)
    assert out == {}



# --- _skel_collect_free_identifiers: the exact grammar-edge-case regressions
# --- found while prototyping this feature against real SKEL artifacts. --- #
def test_skel_free_identifiers_shorthand_value_position_is_a_reference():
    """Regression: ``return {foo};`` is shorthand for ``{foo: foo}`` -- a
    SHORTHAND_PROPERTY_IDENTIFIER node, a grammar type DISTINCT from plain
    IDENTIFIER, that is just as much a read of the outer ``foo`` binding.
    An earlier prototype silently missed this (treated it as neither bound
    nor free), which would have wrongly classified a test that references a
    target export ONLY via object-literal shorthand as needing zero target
    identifiers -- see the end-to-end 'test_shorthand' case in
    test_skel_extract_verified_tests_classifies_extractable_vs_blocked."""
    node, src = _skel_parse_function("function test_shorthand() { const bar = 1; return { add, bar }; }")
    free = COL._skel_collect_free_identifiers(node, src)
    assert "add" in free
    assert "bar" not in free   # locally bound, despite ALSO appearing in shorthand position


def test_skel_free_identifiers_destructured_for_of_bindings_are_local():
    node, src = _skel_parse_function(
        "function test_x() { const t = [[1, 0, 5]]; "
        "for (const [val, floor, ceil] of t) { console.assert(val >= floor); } return true; }"
    )
    free = COL._skel_collect_free_identifiers(node, src)
    assert free.isdisjoint({"val", "floor", "ceil"})


def test_skel_free_identifiers_value_position_of_declarator_is_not_a_binding():
    """Regression: the exact real-artifact bug found in strsim/heapq-style
    fixtures -- a value-position identifier on the RIGHT of a
    ``variable_declarator`` (e.g. ``const s2 = someImportedThing;``) is a
    genuine reference to ``someImportedThing``, never a new binding; only
    ``s2`` (the LEFT/name side) is bound."""
    node, src = _skel_parse_function("function test_x() { const s2 = someImportedThing; return s2 === 'a'; }")
    free = COL._skel_collect_free_identifiers(node, src)
    assert "someImportedThing" in free
    assert "s2" not in free


def test_skel_free_identifiers_nested_helper_function_name_is_local():
    node, src = _skel_parse_function(
        "function test_x() { function helper() { return 42; } return helper() === 42; }"
    )
    free = COL._skel_collect_free_identifiers(node, src)
    assert "helper" not in free


def test_skel_free_identifiers_object_assignment_pattern_binding_vs_default_value():
    """``{ a = someDefaultFn() } = {}``'s ``object_assignment_pattern`` has
    explicit ``left`` (the binding, ``a``) / ``right`` (the default VALUE
    expression) fields; only ``left`` may ever be treated as a binding --
    recursing into ``right`` as if it were a binding would wrongly hide a
    genuine free reference (here ``someDefaultFn``) from classification,
    which could silently under-block a test that actually depends on a
    private/unresolvable helper only reachable through a default value."""
    node, src = _skel_parse_function("function test_x({ a = someDefaultFn() } = {}) { return a; }")
    free = COL._skel_collect_free_identifiers(node, src)
    assert "a" not in free
    assert "someDefaultFn" in free


# --- skel_extract_verified_tests: end-to-end static classification --- #
def test_skel_extract_verified_tests_classifies_extractable_vs_blocked(tmp_path: Path):
    source_js = _write_skel_source_js(tmp_path)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    verified = ["test_trivial", "test_uses_export", "test_uses_private_helper", "test_shorthand",
               "test_with_assert", "test_missing_in_source"]

    outcome = COL.skel_extract_verified_tests(source_js, verified)

    assert outcome is not None
    extractable_names = {t.name for t in outcome.extractable}
    blocked_names = {n for n, _ in outcome.blocked}
    assert extractable_names == {"test_trivial", "test_uses_export", "test_shorthand", "test_with_assert"}
    assert blocked_names == {"test_uses_private_helper", "test_missing_in_source"}

    by_name = {t.name: t for t in outcome.extractable}
    assert by_name["test_uses_export"].target_identifiers == ("add",)
    assert by_name["test_shorthand"].target_identifiers == ("add",)   # the shorthand-position regression
    assert by_name["test_trivial"].target_identifiers == ()
    assert by_name["test_with_assert"].target_identifiers == ()
    assert outcome.safe_require_lines == ('const assert = require("assert");',)

    reasons = dict(outcome.blocked)
    assert "_sub" in reasons["test_uses_private_helper"]
    assert "no matching top-level" in reasons["test_missing_in_source"]


def test_skel_extract_verified_tests_never_inlines_private_helper_body(tmp_path: Path):
    """'Never copies the whole source.js' invariant, at the finest grain:
    even a single blocked private helper's own body text must never leak
    into any EXTRACTABLE test's source_text (which would smuggle reference
    implementation logic into the synthetic harness)."""
    source_js = _write_skel_source_js(tmp_path)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_trivial", "test_uses_export", "test_shorthand", "test_with_assert"]
    )

    assert outcome is not None
    all_extracted_text = "\n".join(t.source_text for t in outcome.extractable)
    assert "a - b" not in all_extracted_text   # _sub's OWN body -- never extracted/inlined
    assert "_sub" not in all_extracted_text
    by_name = {t.name: t for t in outcome.extractable}
    assert by_name["test_trivial"].source_text.strip().startswith("function test_trivial()")


def test_skel_extract_verified_tests_zero_verified_names_yields_empty_outcome(tmp_path: Path):
    source_js = _write_skel_source_js(tmp_path)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    outcome = COL.skel_extract_verified_tests(source_js, [])
    assert outcome is not None
    assert outcome.extractable == []
    assert outcome.blocked == []


def test_skel_extract_verified_tests_none_when_tree_sitter_unavailable(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(COL.C, "optional_import", lambda name: None)
    source_js = _write_skel_source_js(tmp_path)
    assert COL.skel_extract_verified_tests(source_js, ["test_trivial"]) is None


def test_skel_extract_verified_tests_none_when_source_js_missing(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    assert COL.skel_extract_verified_tests(tmp_path / "javascript" / "source.js", ["test_trivial"]) is None


# --- skel_extract_verified_tests: resolution rule (d) -- a source.js --- #
# --- top-level production declaration NOT in ITS OWN module.exports, --- #
# --- but that CodeWeaver's OWN target independently ALSO exports. --- #
# Real heapq-style SKEL fixtures ship source.js WITHOUT any module.exports
# assignment at all: the file is a self-contained script whose own tests
# call the sibling top-level implementation function DIRECTLY by name.
SKEL_NO_EXPORTS_SOURCE_JS = """\
function heapifyMax(arr) {
  return arr.slice().sort((a, b) => b - a);
}

function _heappop_max(arr) {
  return arr.slice(1);
}

function test_heapify_direct() {
  return heapifyMax([3, 1, 2])[0] === 3;
}

function test_uses_private_helper_direct() {
  return _heappop_max([3, 1, 2]).length === 2;
}
"""


def test_skel_extract_verified_tests_target_export_unblocks_non_exported_production_function(tmp_path: Path):
    """Extension/rule (d): if CodeWeaver's OWN target independently ALSO
    exports a symbol with the exact same name as a source.js top-level
    production declaration that source.js itself never exports via
    module.exports, that identifier becomes safely resolvable -- never
    copies source.js's own body, only widens which names may be BOUND from
    CodeWeaver's target, exactly like the existing module.exports path."""
    source_js = _write_skel_source_js(tmp_path, SKEL_NO_EXPORTS_SOURCE_JS)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_heapify_direct", "test_uses_private_helper_direct"],
        target_export_names=frozenset({"heapifyMax"}),
    )

    assert outcome is not None
    assert {t.name for t in outcome.extractable} == {"test_heapify_direct"}
    by_name = {t.name: t for t in outcome.extractable}
    assert by_name["test_heapify_direct"].target_identifiers == ("heapifyMax",)
    blocked = dict(outcome.blocked)
    assert set(blocked) == {"test_uses_private_helper_direct"}
    assert "_heappop_max" in blocked["test_uses_private_helper_direct"]


def test_skel_extract_verified_tests_target_export_names_default_empty_matches_prior_behavior(tmp_path: Path):
    """Regression: omitting ``target_export_names`` (default: empty
    frozenset) must reproduce IDENTICAL blocking behavior to before rule
    (d) existed -- an existing caller that has not been updated never
    silently becomes MORE permissive."""
    source_js = _write_skel_source_js(tmp_path, SKEL_NO_EXPORTS_SOURCE_JS)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(source_js, ["test_heapify_direct"])

    assert outcome is not None
    assert outcome.extractable == []
    assert "heapifyMax" in dict(outcome.blocked)["test_heapify_direct"]


def test_skel_extract_verified_tests_target_export_names_does_not_unblock_unexported_private_helper(tmp_path: Path):
    """Non-regression: rule (d) requires BOTH a matching source.js
    top-level declaration AND a matching CodeWeaver target export --
    ``target_export_names`` not containing the private helper's own name
    must still block it exactly as before, even while OTHER identifiers
    now resolve via rule (d)."""
    source_js = _write_skel_source_js(tmp_path, SKEL_NO_EXPORTS_SOURCE_JS)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_uses_private_helper_direct"],
        target_export_names=frozenset({"heapifyMax"}),   # deliberately excludes _heappop_max
    )

    assert outcome is not None
    assert outcome.extractable == []
    assert "_heappop_max" in dict(outcome.blocked)["test_uses_private_helper_direct"]


def test_skel_extract_verified_tests_target_export_requires_source_js_declaration_too(tmp_path: Path):
    """Rule (d) never resolves an identifier purely because it happens to
    appear in ``target_export_names`` -- it must ALSO be an actual
    top-level function/class declaration in source.js itself (otherwise
    this is just the pre-existing 'no matching declaration'/undefined-
    reference case, unrelated to rule (d))."""
    source_js = _write_skel_source_js(
        tmp_path, "function test_uses_unknown() { return totallyUnknownName(); }\n",
    )
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_uses_unknown"], target_export_names=frozenset({"totallyUnknownName"}),
    )

    assert outcome is not None
    assert outcome.extractable == []
    assert "totallyUnknownName" in dict(outcome.blocked)["test_uses_unknown"]


# --- skel_extract_verified_tests: resolution rule (e) -- "literal --- #
# --- expectation" top-level pure-data constants. --- #
SKEL_LITERAL_EXPECTATION_SOURCE_JS = """\
function add(a, b) {
  return a + b;
}

const EXPECTED_SUM = 5;
const EXPECTED_LIST = [1, 2, 3];

function test_uses_literal_expectation() {
  return add(2, 3) === EXPECTED_SUM;
}

function test_uses_literal_list() {
  return JSON.stringify([1, 2, 3]) === JSON.stringify(EXPECTED_LIST);
}

module.exports = { add };
"""


def test_skel_extract_verified_tests_literal_expectation_constant_is_extractable(tmp_path: Path):
    source_js = _write_skel_source_js(tmp_path, SKEL_LITERAL_EXPECTATION_SOURCE_JS)
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_uses_literal_expectation", "test_uses_literal_list"],
    )

    assert outcome is not None
    assert {t.name for t in outcome.extractable} == {"test_uses_literal_expectation", "test_uses_literal_list"}
    by_name = {t.name: t for t in outcome.extractable}
    assert by_name["test_uses_literal_expectation"].literal_support_names == ("EXPECTED_SUM",)
    assert by_name["test_uses_literal_expectation"].target_identifiers == ("add",)
    assert by_name["test_uses_literal_list"].literal_support_names == ("EXPECTED_LIST",)
    assert set(outcome.literal_support_lines) == {"const EXPECTED_SUM = 5;", "const EXPECTED_LIST = [1, 2, 3];"}


def test_skel_extract_verified_tests_never_leaks_non_pure_top_level_constant(tmp_path: Path):
    """A top-level const whose value is NOT a provably pure literal (e.g.
    it calls a function) must never be treated as a safe literal-support
    reference -- referencing it still blocks the test, never silently
    inlines a computed value that could smuggle reference logic."""
    source_js = _write_skel_source_js(
        tmp_path,
        "function computeDefault() { return 42; }\n"
        "const COMPUTED = computeDefault();\n"
        "function test_uses_computed() { return COMPUTED === 42; }\n",
    )
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(source_js, ["test_uses_computed"])

    assert outcome is not None
    assert outcome.extractable == []
    assert "COMPUTED" in dict(outcome.blocked)["test_uses_computed"]


def test_skel_extract_verified_tests_combines_target_export_and_literal_and_builtin_in_one_test(tmp_path: Path):
    """Resolution rules (a)/(d)/(e) must compose correctly within a SINGLE
    test's own free-identifier set."""
    source_js = _write_skel_source_js(
        tmp_path,
        "function heapifyMax(arr) { return arr.slice().sort((a, b) => b - a); }\n"
        "const EXPECTED_TOP = 3;\n"
        "function test_mixed_resolution() {\n"
        "  const result = heapifyMax([3, 1, 2]);\n"
        "  return result[0] === EXPECTED_TOP && typeof console === 'object';\n"
        "}\n",
    )
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")

    outcome = COL.skel_extract_verified_tests(
        source_js, ["test_mixed_resolution"], target_export_names=frozenset({"heapifyMax"}),
    )

    assert outcome is not None
    assert {t.name for t in outcome.extractable} == {"test_mixed_resolution"}
    test = outcome.extractable[0]
    assert test.target_identifiers == ("heapifyMax",)
    assert test.literal_support_names == ("EXPECTED_TOP",)


# --- skel_build_validated_harness_source: pure string assembly, no --- #
# --- tree-sitter dependency -- always runs regardless of environment --- #
def test_skel_build_validated_harness_source_contains_expected_pieces():
    outcome = COL.SkelExtractionOutcome(
        extractable=[
            COL.SkelExtractedTest(name="test_a", source_text="function test_a() { return add(1, 2) === 3; }",
                                  target_identifiers=("add",)),
            COL.SkelExtractedTest(name="test_b", source_text="function test_b() { return true; }",
                                  target_identifiers=()),
        ],
        blocked=[("test_c", "references identifier(s) ['_priv'] that are private/non-exported ...")],
        safe_require_lines=('const assert = require("assert");',),
    )
    src = COL.skel_build_validated_harness_source(outcome)
    assert f'await import("./{COL.SKEL_TARGET_ENTRY_FILENAME}")' in src
    assert "const { add } = __target;" in src
    assert 'const assert = require("assert");' in src
    assert "function test_a() { return add(1, 2) === 3; }" in src
    assert "function test_b() { return true; }" in src
    assert "console.assert = function" in src
    assert "__result === false" in src
    assert "# pass " in src and "# fail " in src
    assert '"test_a"' in src and '"test_b"' in src
    assert "_priv" not in src   # a BLOCKED test's own reason text never leaks into the harness itself


def test_skel_build_validated_harness_source_zero_extractable_is_still_valid():
    outcome = COL.SkelExtractionOutcome(extractable=[], blocked=[("t", "some reason")], safe_require_lines=())
    src = COL.skel_build_validated_harness_source(outcome)
    assert "const __tests = [];" in src
    assert "# pass " in src and "# fail " in src


def test_skel_build_validated_harness_source_guards_target_load_failure():
    outcome = COL.SkelExtractionOutcome(
        extractable=[COL.SkelExtractedTest(name="t", source_text="function t() { return true; }",
                                           target_identifiers=())],
        blocked=[], safe_require_lines=(),
    )
    src = COL.skel_build_validated_harness_source(outcome)
    assert "__targetLoadError" in src
    assert "catch" in src


def test_skel_build_validated_harness_source_includes_literal_support_lines_before_test_bodies():
    """Resolution rule (e): a "literal-expectation" constant's own verbatim
    declaration text must be spliced into the harness BEFORE the extracted
    test bodies that reference it."""
    outcome = COL.SkelExtractionOutcome(
        extractable=[
            COL.SkelExtractedTest(name="test_a", source_text="function test_a() { return EXPECTED === 5; }",
                                  target_identifiers=(), literal_support_names=("EXPECTED",)),
        ],
        blocked=[], safe_require_lines=(), literal_support_lines=("const EXPECTED = 5;",),
    )
    src = COL.skel_build_validated_harness_source(outcome)
    assert "const EXPECTED = 5;" in src
    assert src.index("const EXPECTED = 5;") < src.index("function test_a()")


# --- skel_parse_verified_test_names: pure stdlib CSV parsing --- #
def test_skel_parse_verified_test_names_filters_by_verified_column(tmp_path: Path):
    csv_path = tmp_path / "test_name_mapping.csv"
    _write(csv_path, SKEL_CSV_HEADER + "\n"
          "bst,p,test_put,j,test_put,1,1,0\n"
          "bst,p,test_search,j,test_search,0,1,1\n")
    assert COL.skel_parse_verified_test_names(csv_path) == ["test_put"]


def test_skel_parse_verified_test_names_none_when_csv_missing(tmp_path: Path):
    assert COL.skel_parse_verified_test_names(tmp_path / "nope.csv") is None


def test_skel_parse_verified_test_names_none_when_expected_columns_missing(tmp_path: Path):
    csv_path = tmp_path / "test_name_mapping.csv"
    _write(csv_path, "project,other_column\nbst,x\n")
    assert COL.skel_parse_verified_test_names(csv_path) is None


def test_skel_parse_verified_test_names_empty_list_distinct_from_none(tmp_path: Path):
    """CSV present but genuinely zero verified rows (e.g. real toml
    fixtures) must parse as a real empty list -- never confused with a
    missing/malformed CSV (None), since callers treat these two cases
    completely differently (a real MEASURED 0 vs. Status.UNAVAILABLE)."""
    csv_path = tmp_path / "test_name_mapping.csv"
    _write(csv_path, SKEL_CSV_HEADER + "\ntoml,p,t1,j,t1,0,1,1\n")
    names = COL.skel_parse_verified_test_names(csv_path)
    assert names == []
    assert names is not None


def test_skel_parse_validated_test_names_includes_flag_zero_rows(tmp_path: Path):
    csv_path = tmp_path / "test_name_mapping.csv"
    _write(csv_path, SKEL_CSV_HEADER + "\n"
          "heapq,p,test_push,j,test_push,1,1,0\n"
          "heapq,p,test_pop,j,test_pop,0,1,1\n")
    assert COL.skel_parse_validated_test_names(csv_path) == ["test_push", "test_pop"]


# --- skel_validated_tests_eval: end-to-end orchestration/aggregation --- #
def test_skel_validated_tests_eval_unavailable_when_ref_project_dir_none(tmp_path: Path):
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, None, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert result["executed"].value is None
    assert snapshot.calls == []


def test_skel_validated_tests_eval_unavailable_when_csv_missing(tmp_path: Path):
    ref = tmp_path / "ref"
    _write(ref / "javascript" / "source.js", SKEL_SAMPLE_SOURCE_JS)
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert "test_name_mapping.csv" in result["executed"].reason
    assert snapshot.calls == []


def test_skel_validated_tests_eval_flag_zero_rows_stay_in_fixed_denominator(tmp_path: Path):
    """The verified flag is a prior-tool outcome, never a row selector."""
    ref = _make_skel_reference_with_csv(tmp_path / "refroot", "toml", verified_js_names=[],
                                        all_js_names=["test_a", "test_b"])
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = {};")
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert result["expected"].is_measured
    assert result["expected"].value == 2
    assert result["not_executed"].value == 2
    assert snapshot.calls == []


def test_skel_validated_tests_eval_unavailable_when_source_js_missing(tmp_path: Path):
    ref = tmp_path / "refroot" / "recodeagent_translations" / "data" / "tool_projects" / "skel" / "bst"
    _write(ref / "test_name_mapping.csv", SKEL_CSV_HEADER + "\nbst,p,test_put,j,test_put,1,1,0\n")
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = {};")
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert "source.js" in result["executed"].reason
    # `expected` stays measured (the CSV's own known count) even though the
    # LATER source.js-missing step prevented anything from executing.
    assert result["expected"].is_measured
    assert result["expected"].value == 1
    assert snapshot.calls == []


def test_skel_validated_tests_eval_unavailable_when_tree_sitter_missing(tmp_path: Path, monkeypatch):
    ref = _make_skel_reference_with_csv(tmp_path / "refroot", "bst", verified_js_names=["test_trivial"])
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    monkeypatch.setattr(COL.C, "optional_import", lambda name: None)
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert "tree-sitter" in result["executed"].reason
    assert result["expected"].is_measured and result["expected"].value == 1
    assert snapshot.calls == []


def test_skel_validated_tests_eval_unavailable_when_all_verified_tests_blocked(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(tmp_path / "refroot", "bst",
                                        verified_js_names=["test_uses_private_helper"])
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.UNAVAILABLE
    assert "blocked" in result["executed"].reason
    assert result["expected"].is_measured and result["expected"].value == 1
    assert snapshot.calls == []


def test_skel_validated_tests_eval_missing_when_target_dir_absent(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(tmp_path / "refroot", "bst", verified_js_names=["test_trivial"])
    target = tmp_path / "target"   # never created
    snapshot = SnapshotRunner()
    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)
    assert result["executed"].status == Status.MISSING
    assert result["expected"].is_measured and result["expected"].value == 1
    assert snapshot.calls == []


def test_skel_validated_tests_eval_measured_and_never_leaks_source_js_or_mutates_target(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "bst",
        verified_js_names=["test_trivial", "test_uses_export", "test_shorthand", "test_with_assert"],
    )
    target = tmp_path / "run" / "pipeline" / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    snapshot = SnapshotRunner(stdout="# pass 4\n# fail 0\n")

    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)

    assert result["executed"].is_measured
    assert result["executed"].value == 4
    assert result["passed"].value == 4
    assert result["failed"].value == 0
    assert result["expected"].is_measured
    assert result["expected"].value == 4
    assert result["not_executed"].value == 0
    assert len(snapshot.calls) == 1
    call = snapshot.calls[0]
    assert call["argv"] == ["node", COL.SKEL_VALIDATED_HARNESS_FILENAME]
    files = call["cwd_files"]
    assert "source.js" not in files                # the reference's own file NEVER copied in
    assert COL.SKEL_VALIDATED_HARNESS_FILENAME in files
    assert "index.js" in files                     # CodeWeaver's own target, present in the temp copy
    assert files["index.js"] == "module.exports = { add: (a, b) => a + b };"
    assert "function test_trivial()" in files[COL.SKEL_VALIDATED_HARNESS_FILENAME]
    assert "_sub" not in files[COL.SKEL_VALIDATED_HARNESS_FILENAME]   # private helper never leaked
    # the ORIGINAL target tree is never mutated -- only a TEMPORARY copy is:
    assert (target / "index.js").read_text(encoding="utf-8") == "module.exports = { add: (a, b) => a + b };"
    assert not (target / COL.SKEL_VALIDATED_HARNESS_FILENAME).exists()


def test_skel_validated_tests_eval_partial_extraction_total_excludes_blocked_and_notes_reason(tmp_path: Path):
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "bst",
        verified_js_names=["test_trivial", "test_uses_export", "test_uses_private_helper"],
    )
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    snapshot = SnapshotRunner(stdout="# pass 2\n# fail 0\n")

    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)

    assert result["executed"].is_measured
    assert result["executed"].value == 2   # NOT 3 -- the blocked test is excluded from the executed count
    assert "1 of 3" in result["executed"].reason
    assert "test_uses_private_helper" in result["executed"].reason
    # expected stays the FULL CSV-known count of 3 regardless of the blocked
    # extraction -- the paper's own TPR denominator is not reduced by a
    # harness-side extraction limitation.
    assert result["expected"].is_measured
    assert result["expected"].value == 3
    assert result["not_executed"].value == 1


# --- skel_validated_tests_eval: resolution rule (d) end-to-end -- the --- #
# --- centerpiece regression for a real no-module.exports SKEL project. --- #
def test_skel_validated_tests_eval_measured_for_no_module_exports_project_via_target_export_names(tmp_path: Path):
    """Centerpiece regression: a real heapq-style SKEL project ships
    source.js with NO module.exports assignment at all, so verified tests
    call the sibling top-level implementation function DIRECTLY. Before
    resolution rule (d) existed, EVERY such test was blocked, so this
    project's independently-validated developer tests were always
    Status.UNAVAILABLE end-to-end -- with CodeWeaver's OWN target
    independently exporting a same-named symbol, they are now genuinely
    MEASURED (never copying any of source.js's own text into the harness)."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "heapq", verified_js_names=["test_heapify_direct"],
        source_text=SKEL_NO_EXPORTS_SOURCE_JS,
    )
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { heapifyMax: (a) => a };")
    snapshot = SnapshotRunner(stdout="# pass 1\n# fail 0\n")

    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)

    assert result["executed"].is_measured
    assert result["executed"].value == 1
    assert result["passed"].value == 1
    assert result["expected"].is_measured
    assert result["expected"].value == 1
    assert result["not_executed"].value == 0
    assert len(snapshot.calls) == 1
    files = snapshot.calls[0]["cwd_files"]
    assert "source.js" not in files
    harness_text = files[COL.SKEL_VALIDATED_HARNESS_FILENAME]
    assert "heapifyMax" in harness_text
    assert "_heappop_max" not in harness_text   # private helper's own name never leaked either


def test_skel_validated_tests_eval_still_unavailable_when_target_does_not_export_matching_name(tmp_path: Path):
    """Non-regression: if CodeWeaver's OWN target does NOT independently
    export a same-named symbol, a no-module.exports project's tests remain
    exactly as UNAVAILABLE as before rule (d) existed -- this extension is
    purely additive, never a blanket assumption that any target export
    will do."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref = _make_skel_reference_with_csv(
        tmp_path / "refroot", "heapq", verified_js_names=["test_heapify_direct"],
        source_text=SKEL_NO_EXPORTS_SOURCE_JS,
    )
    target = tmp_path / "target"
    _write(target / "index.js", "module.exports = { somethingElseEntirely: (a) => a };")
    snapshot = SnapshotRunner()

    result = COL.skel_validated_tests_eval(target, ref, timeout=None, runner=snapshot)

    assert result["executed"].status == Status.UNAVAILABLE
    assert "heapifyMax" in result["executed"].reason
    assert result["expected"].is_measured
    assert result["expected"].value == 1
    assert snapshot.calls == []


# --------------------------------------------------------------------------- #
# compute_not_executed / compute_paper_pass_rate: the core expected-vs-
# executed helpers (paper TPR = passed / expected, a FIXED oracle-known
# denominator -- e.g. the paper's own worked example of 1,822/2,107 despite
# only TE=1,970 tests actually executing -- NEVER passed / executed).
# --------------------------------------------------------------------------- #
def test_compute_not_executed_normal_subtraction_when_both_measured():
    result = COL.compute_not_executed(Measurement.ok(10), Measurement.ok(7))
    assert result.is_measured
    assert result.value == 3


def test_compute_not_executed_clamped_to_zero_never_negative():
    """A static 'expected' regex/AST counter can legitimately UNDER-count
    relative to what a real test runner reports as executed (e.g. CRUST's
    whole-crate `cargo test` may also run the target's own embedded
    #[test]s beyond the restored scaffold contract) -- not_executed must
    clamp to 0 rather than go negative in that case."""
    result = COL.compute_not_executed(Measurement.ok(2), Measurement.ok(5))
    assert result.is_measured
    assert result.value == 0


def test_compute_not_executed_propagates_expected_status_when_expected_not_measured():
    unavailable = Measurement.unavailable("no scaffold for this dataset")
    result = COL.compute_not_executed(unavailable, Measurement.ok(3))
    assert result.status == Status.UNAVAILABLE
    assert result.value is None
    assert "no scaffold for this dataset" in result.reason


def test_compute_not_executed_full_expected_count_when_executed_not_measured():
    """A build failure must never present not_executed as a fabricated 0 --
    when the run never executed at all, EVERY expected test is honestly
    not_executed, with a reason naming the real underlying failure."""
    expected = Measurement.ok(6)
    executed = Measurement(value=None, status=Status.ERROR, reason="cargo build failed: exit code 101")
    result = COL.compute_not_executed(expected, executed)
    assert result.status == Status.MEASURED
    assert result.value == 6
    assert "cargo build failed" in result.reason
    assert "'error'" in result.reason


def test_compute_paper_pass_rate_normal_ratio_when_both_measured():
    result = COL.compute_paper_pass_rate(Measurement.ok(2107), Measurement.ok(1822))
    assert result.is_measured
    assert result.value == pytest.approx(1822 / 2107)


def test_compute_paper_pass_rate_propagates_expected_status_when_not_measured():
    unavailable = Measurement.unavailable("--reference-results-root not supplied")
    result = COL.compute_paper_pass_rate(unavailable, Measurement.ok(5))
    assert result.status == Status.UNAVAILABLE
    assert result.value is None


def test_compute_paper_pass_rate_not_applicable_when_expected_is_zero():
    result = COL.compute_paper_pass_rate(Measurement.ok(0), Measurement.ok(0))
    assert result.status == Status.NOT_APPLICABLE
    assert result.value is None


def test_compute_paper_pass_rate_zero_substituted_numerator_when_passed_not_measured():
    """A build failure (passed is Status.ERROR, never a fabricated measured
    0) must still produce a real, measured pass rate of 0/expected -- exactly
    mirroring the paper's own methodology that a never-built project
    contributes zero passing tests, not an excluded/undefined row -- while
    the reason names the real underlying failure so it is never mistaken for
    a genuine all-failing execution."""
    expected = Measurement.ok(6)
    passed = Measurement(value=None, status=Status.ERROR, reason="cargo build failed: exit code 101")
    result = COL.compute_paper_pass_rate(expected, passed)
    assert result.status == Status.MEASURED
    assert result.value == 0.0
    assert "cargo build failed" in result.reason


def test_compute_paper_pass_rate_never_clamped_above_one():
    """Deliberately NOT clamped: a >100% rate (passed > expected, e.g. a
    whole-crate `cargo test` running extra tests beyond the static contract
    count) is left visible as an honest signal of a data-quality mismatch,
    never silently masked to 1.0."""
    result = COL.compute_paper_pass_rate(Measurement.ok(2), Measurement.ok(5))
    assert result.is_measured
    assert result.value == pytest.approx(2.5)


# --------------------------------------------------------------------------- #
# count_rust_test_attributes / count_python_test_functions: the static,
# best-effort "expected" counters shared across CRUST/Oxidizer (#[test]) and
# AlphaTrans (pytest-/unittest-style Python test functions).
# --------------------------------------------------------------------------- #
def test_count_rust_test_attributes_counts_across_multiple_files(tmp_path: Path):
    f1 = tmp_path / "a.rs"
    f2 = tmp_path / "b.rs"
    _write(f1, "#[test]\nfn one() {}\n#[test]\nfn two() {}\n")
    _write(f2, "fn not_a_test() {}\n#[test]\nfn three() {}\n")
    assert COL.count_rust_test_attributes([f1, f2]) == 3


def test_count_rust_test_attributes_tolerates_whitespace_variants(tmp_path: Path):
    f = tmp_path / "a.rs"
    _write(f, "# [ test ]\nfn odd_spacing() {}\n#[test]\nfn normal() {}\n")
    assert COL.count_rust_test_attributes([f]) == 2


def test_count_rust_test_attributes_skips_unreadable_file(tmp_path: Path):
    missing = tmp_path / "does_not_exist.rs"
    present = tmp_path / "present.rs"
    _write(present, "#[test]\nfn one() {}\n")
    assert COL.count_rust_test_attributes([missing, present]) == 1


def test_count_python_test_functions_counts_pytest_and_unittest_style(tmp_path: Path):
    f = tmp_path / "test_foo.py"
    _write(f, (
        "def test_free_function():\n    assert True\n\n"
        "class FooTest:\n"
        "    def test_method(self):\n        assert True\n\n"
        "    async def test_async_method(self):\n        assert True\n\n"
        "def helper_not_a_test():\n    pass\n"
    ))
    assert COL.count_python_test_functions([f]) == 3


def test_count_python_test_functions_skips_file_with_syntax_error(tmp_path: Path):
    bad = tmp_path / "bad.py"
    good = tmp_path / "good.py"
    _write(bad, "def test_broken(:\n    this is not valid python\n")
    _write(good, "def test_ok():\n    assert True\n")
    assert COL.count_python_test_functions([bad, good]) == 1


def test_count_python_test_functions_empty_for_no_paths():
    assert COL.count_python_test_functions([]) == 0


# --------------------------------------------------------------------------- #
# Per-tool validated_tests_expected computation (the paper's FIXED,
# oracle-known TPR denominator) -- available BEFORE and INDEPENDENTLY of any
# CodeWeaver translation/build attempt.
# --------------------------------------------------------------------------- #
def test_crust_validated_tests_expected_native_na_when_scaffold_missing(tmp_path: Path):
    result = COL.crust_validated_tests_expected_native(tmp_path / "no-such-scaffold")
    assert result.status == Status.NOT_APPLICABLE


def test_crust_validated_tests_expected_native_unavailable_when_no_rs_contract_paths(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "Cargo.toml", "PRISTINE")   # a contract file, but not a .rs one
    result = COL.crust_validated_tests_expected_native(scaffold)
    assert result.status == Status.UNAVAILABLE


def test_crust_validated_tests_expected_native_counts_test_attributes_in_contract_dirs(tmp_path: Path):
    scaffold = tmp_path / "scaffold"
    _write(scaffold / "Cargo.toml", "PRISTINE")
    _write(scaffold / "src" / "bin" / "harness.rs", "#[test]\nfn a() {}\n#[test]\nfn b() {}\n")
    _write(scaffold / "tests" / "extra_test.rs", "#[test]\nfn c() {}\n")
    result = COL.crust_validated_tests_expected_native(scaffold)
    assert result.is_measured
    assert result.value == 3



def test_oxidizer_validated_tests_expected_unavailable_when_no_oracle_files():
    result = COL.oxidizer_validated_tests_expected([])
    assert result.status == Status.UNAVAILABLE
    assert result.value is None


def test_oxidizer_validated_tests_expected_counts_test_attributes(tmp_path: Path):
    f1 = tmp_path / "foo_test.rs"
    f2 = tmp_path / "Bar_Test.rs"
    _write(f1, "#[test]\nfn a() {}\n")
    _write(f2, "#[test]\nfn b() {}\n#[test]\nfn c() {}\n")
    result = COL.oxidizer_validated_tests_expected([f1, f2])
    assert result.is_measured
    assert result.value == 3


def test_read_name_mapping_flattens_real_planner_categories(tmp_path: Path):
    _write(
        tmp_path / "pipeline" / "plan.json",
        json.dumps({
            "name_mapping": {
                "functions": {"NewLuhn": "new_luhn"},
                "methods": {"Verify": "verify"},
                "metadata": {"ignored": 1},
            }
        }),
    )
    assert COL.read_name_mapping(tmp_path) == {
        "NewLuhn": "new_luhn",
        "Verify": "verify",
    }


def test_oxidizer_expected_prefers_paper_runtime_denominator(tmp_path: Path):
    oracle = tmp_path / "checkdigit_test.rs"
    _write(oracle, "#[test]\nfn a() {}\n#[test]\nfn b() {}\n")
    result = COL.oxidizer_validated_tests_expected(
        [oracle], "checkdigit", official_artifact_verified=True
    )
    assert result.value == 36
    assert "runtime-case" in result.reason


def test_retain_named_rust_tests_removes_only_uninventoried_tests():
    source = """
fn helper() { let _x = "{ not a brace }"; }
#[test]
fn keep_me() { assert!(helper_call({ 1 })); }
#[test]
fn remove_me() { assert_eq!(1, 1); }
"""
    result = COL.retain_named_rust_tests(source, {"keep_me"})
    assert "fn helper" in result
    assert "fn keep_me" in result
    assert "fn remove_me" not in result


def test_rewrite_rust_use_paths_uses_target_public_module_index(tmp_path: Path):
    target = tmp_path / "target"
    _write(
        target / "src" / "lib.rs",
        "pub mod utils;\npub use algorithms::cosine::cosine_similarity;\n",
    )
    _write(target / "src" / "utils.rs", "pub fn equal() {}\n")
    _write(target / "src" / "algorithms" / "cosine.rs",
           "pub fn cosine_similarity() {}\n")
    paths = COL.rust_target_symbol_paths(target)
    source = (
        "use edlib::cosine::CosineSimilarity;\n"
        "use edlib::internal::utils::utils::Equal;\n"
    )
    result = COL.rewrite_rust_use_paths(
        source,
        {"CosineSimilarity": "cosine_similarity", "Equal": "equal"},
        paths,
    )
    assert "use edlib::cosine_similarity;" in result
    assert "use edlib::utils::equal;" in result


def test_oxidizer_reference_inventory_reads_curated_mapping(tmp_path: Path):
    project = tmp_path / "checkdigit"
    _write(
        project / "test_name_mapping.csv",
        "rust test path,rust test name\n"
        "tests/checkdigit_test.rs,TestNewLuhn\n"
        "tests/checkdigit_test.rs,TestNewDamm\n",
    )
    assert COL.oxidizer_reference_test_inventory(project) == {
        "checkdigit_test.rs": {"TestNewLuhn", "TestNewDamm"}
    }


def test_alphatrans_validated_tests_expected_unavailable_when_verified_dir_none():
    result = COL.alphatrans_validated_tests_expected(None)
    assert result.status == Status.UNAVAILABLE
    assert result.value is None


def test_alphatrans_validated_tests_expected_unavailable_when_no_py_files(tmp_path: Path):
    verified = tmp_path / "verified_test"
    verified.mkdir(parents=True)
    result = COL.alphatrans_validated_tests_expected(verified)
    assert result.status == Status.UNAVAILABLE


def test_alphatrans_validated_tests_expected_counts_test_functions(tmp_path: Path):
    verified = tmp_path / "verified_test"
    _write(verified / "test_foo.py", "def test_a():\n    assert True\n\ndef test_b():\n    assert True\n")
    _write(verified / "test_bar.py", "def test_c():\n    assert True\n")
    result = COL.alphatrans_validated_tests_expected(verified)
    assert result.is_measured
    assert result.value == 3


# --------------------------------------------------------------------------- #
# evaluate_independent_oracle: full per-tool dispatch, no-leakage, and
# unavailable-vs-zero (never a fabricated 0/pass when there is no oracle)
# --------------------------------------------------------------------------- #
def test_evaluate_independent_oracle_oxidizer_never_leaks_reference_impl_or_cargo(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    _make_oxidizer_reference(ref_root, "oxi_proj")
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "Cargo.toml", "OWN_CARGO_TOML")
    _write(target / "src" / "lib.rs", "OWN_IMPL")
    _write(target / "tests" / "existing_dev_test.rs", "OWN_DEV_TEST")   # must be wiped, not merged
    manifest_row = {"project": "oxi_proj"}
    snapshot = SnapshotRunner(stdout="test result: ok. 2 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL.evaluate_independent_oracle("oxidizer", run_dir, manifest_row, OXIDIZER_SPEC,
                                             ref_root, timeout=None, runner=snapshot)

    assert result.validated["executed"].is_measured
    assert result.validated["executed"].value == 2
    # expected is a static #[test] count over the reference oracle files
    # ALONE (foo_test.rs + Bar_Test.rs, one #[test] each) -- read straight
    # from --reference-results-root, independent of the mocked test run.
    assert result.validated["expected"].is_measured
    assert result.validated["expected"].value == 2
    assert result.validated["not_executed"].value == 0
    assert result.function_validation["total"].is_measured
    assert result.function_harness_tests["total"].is_measured
    assert result.oracle_integrity.status == Status.NOT_APPLICABLE   # only CRUST gets a real value
    assert len(snapshot.calls) == 4
    validated_files, funcval_files, generated_files_a, generated_files_b = (
        snapshot.calls[0]["cwd_files"],
        snapshot.calls[1]["cwd_files"],
        snapshot.calls[2]["cwd_files"],
        snapshot.calls[3]["cwd_files"],
    )

    for files in (validated_files, funcval_files, generated_files_a, generated_files_b):
        assert files["Cargo.toml"] == "OWN_CARGO_TOML"        # CodeWeaver's own manifest, untouched
        assert files["src/lib.rs"] == "OWN_IMPL"               # CodeWeaver's own source, untouched
        assert "src/main.rs" not in files                      # the reference's own impl NEVER copied in
        assert files.get("Cargo.toml") != "REF_CARGO_TOML"     # never overwritten by the reference's manifest

    assert "tests/foo_test.rs" in validated_files
    assert "tests/Bar_Test.rs" in validated_files
    assert "tests/existing_dev_test.rs" not in validated_files    # wiped, never merged with CodeWeaver's own
    assert "tests/harness_helper.rs" not in validated_files       # function-validation-only file
    assert "tests/baz_generated_test.rs" not in validated_files   # excluded ("generated")

    assert "tests/harness_helper.rs" in funcval_files
    assert "tests/foo_test.rs" not in funcval_files
    assert "tests/existing_dev_test.rs" not in funcval_files
    assert "tests/baz_generated_test.rs" in generated_files_a
    assert "tests/mod_generated.rs" in generated_files_b


def test_evaluate_independent_oracle_oxidizer_unavailable_without_reference_root(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "target" / "Cargo.toml", "OWN_CARGO_TOML")
    manifest_row = {"project": "oxi_proj"}
    snapshot = SnapshotRunner()
    result = COL.evaluate_independent_oracle("oxidizer", run_dir, manifest_row, OXIDIZER_SPEC,
                                             None, timeout=None, runner=snapshot)
    assert result.validated["executed"].status == Status.UNAVAILABLE
    assert result.validated["executed"].value is None            # never a fabricated 0
    assert result.validated["expected"].status == Status.UNAVAILABLE
    assert result.validated["expected"].value is None
    assert result.validated["not_executed"].status == Status.UNAVAILABLE
    assert result.validated["not_executed"].value is None
    assert result.function_validation["total"].status == Status.UNAVAILABLE
    assert result.function_validation["total"].value is None
    assert result.function_harness_tests["total"].status == Status.UNAVAILABLE
    assert snapshot.calls == []                                # no subprocess when there is nothing to evaluate


# --------------------------------------------------------------------------- #
# Oxidizer identifier-rewrite: end-to-end proof of the concrete verified
# ``oxidizer__checkdigit`` scenario (the reference oracle calls ``NewLuhn``,
# CodeWeaver's own idiomatic translation only ever exposes ``new_luhn``)
# through the PUBLIC evaluate_independent_oracle dispatch function -- not
# just the private rewrite helpers tested in isolation above.
# --------------------------------------------------------------------------- #
def _make_oxidizer_checkdigit_reference(root: Path, project: str) -> Path:
    """A minimal, focused Oxidizer reference fixture modeling the concrete
    verified ``oxidizer__checkdigit`` case: the oracle developer test calls
    ``NewLuhn()`` (the SOURCE-language spelling) -- see collect.py's "Oracle
    identifier-rewrite (Oxidizer only)" section."""
    project_dir = root / "recodeagent_translations" / "data" / "tool_projects" / "oxidizer" / project
    rust_dir = project_dir / "rust"
    _write(rust_dir / "Cargo.toml", "REF_CARGO_TOML")
    _write(rust_dir / "src" / "main.rs", "REF_IMPL")
    _write(rust_dir / "tests" / "checkdigit_test.rs",
          "#[test]\nfn test_luhn() {\n    let l = NewLuhn();\n    assert!(l.checksum(\"1\"));\n}\n")
    return project_dir


class IdentifierAwareCargoRunner:
    """Fake command runner simulating a real ``cargo test``: FAILS with a
    realistic rustc "cannot find function" compile error (no parseable
    test-result summary at all -- exactly what a real compile failure looks
    like) if ANY copied ``.rs`` file under ``cwd`` still contains
    ``missing_symbol`` (modeling CodeWeaver's target never defining that
    symbol), else SUCCEEDS with ``pass_stdout``. Proves the identifier-
    rewrite mechanism end-to-end: the SAME reference oracle file transitions
    from an unavailable compile failure to a measured, passing run purely
    because ``name_mapping`` was supplied to the run's own ``plan.json``."""
    def __init__(self, *, missing_symbol: str, pass_stdout: str):
        self.missing_symbol = missing_symbol
        self.pass_stdout = pass_stdout
        self.calls: list[dict] = []

    def __call__(self, argv, *, cwd, timeout=None):
        cwd = Path(cwd)
        texts = [p.read_text(encoding="utf-8", errors="replace") for p in cwd.rglob("*.rs") if p.is_file()]
        self.calls.append({"argv": list(argv), "cwd": str(cwd)})
        if any(self.missing_symbol in t for t in texts):
            return ExecResult(
                argv=list(argv), returncode=101, stdout="",
                stderr=f"error[E0425]: cannot find function `{self.missing_symbol}` in this scope\n"
                       " --> tests/checkdigit_test.rs:3:17\n",
                duration_s=0.01, timed_out=False, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), cwd=str(cwd))
        return ExecResult(argv=list(argv), returncode=0, stdout=self.pass_stdout, stderr="", duration_s=0.01,
                          timed_out=False, started_at=C.utcnow_iso(), ended_at=C.utcnow_iso(), cwd=str(cwd))


def test_evaluate_independent_oracle_oxidizer_name_mapping_fixes_idiomatic_naming_mismatch(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    _make_oxidizer_checkdigit_reference(ref_root, "checkdigit")
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "target" / "Cargo.toml", "OWN_CARGO_TOML")
    _write(run_dir / "pipeline" / "target" / "src" / "lib.rs", "pub fn new_luhn() {}")   # idiomatic Rust name
    _write(run_dir / "pipeline" / "plan.json",
          json.dumps({"name_mapping": {"NewLuhn": "new_luhn"}}))   # the real Planner's own artifact
    manifest_row = {"project": "checkdigit"}
    runner = IdentifierAwareCargoRunner(
        missing_symbol="NewLuhn", pass_stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL.evaluate_independent_oracle("oxidizer", run_dir, manifest_row, OXIDIZER_SPEC,
                                             ref_root, timeout=None, runner=runner)

    assert result.validated["executed"].is_measured
    assert result.validated["executed"].value == 1
    assert result.validated["passed"].value == 1
    assert result.validated["failed"].value == 0
    assert "identifier rewrite applied" in result.validated["executed"].reason
    assert "NewLuhn" in result.validated["executed"].reason


def test_evaluate_independent_oracle_oxidizer_derives_unambiguous_name_mapping(tmp_path: Path):
    """A case/underscore-only API rename is safely derived from target code
    even for variants that do not emit a Planner artifact."""
    ref_root = tmp_path / "refroot"
    _make_oxidizer_checkdigit_reference(ref_root, "checkdigit")
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "target" / "Cargo.toml", "OWN_CARGO_TOML")
    _write(run_dir / "pipeline" / "target" / "src" / "lib.rs", "pub fn new_luhn() {}")
    # deliberately no pipeline/plan.json at all -> read_name_mapping returns {}
    manifest_row = {"project": "checkdigit"}
    runner = IdentifierAwareCargoRunner(
        missing_symbol="NewLuhn", pass_stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL.evaluate_independent_oracle("oxidizer", run_dir, manifest_row, OXIDIZER_SPEC,
                                             ref_root, timeout=None, runner=runner)

    assert result.validated["expected"].is_measured
    assert result.validated["expected"].value == 1
    assert result.validated["executed"].is_measured
    assert result.validated["executed"].value == 1
    assert result.validated["passed"].value == 1
    assert result.validated["failed"].value == 0
    assert result.validated["not_executed"].value == 0
    assert "identifier rewrite applied" in result.validated["executed"].reason
    assert "NewLuhn" in result.validated["executed"].reason


def test_evaluate_with_replaced_subdir_name_mapping_none_is_byte_identical_to_before(tmp_path: Path):
    target = tmp_path / "target"
    _write(target / "Cargo.toml", "X")
    oracle = tmp_path / "checkdigit_test.rs"
    original = "#[test]\nfn t() { NewLuhn(); }\n"
    _write(oracle, original)
    snapshot = SnapshotRunner(stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL._evaluate_with_replaced_subdir(target, [oracle], "tests", "oxidizer", ["cargo", "test"],
                                                timeout=None, runner=snapshot, name_mapping=None)

    copied = snapshot.calls[0]["cwd_files"]["tests/checkdigit_test.rs"]
    assert copied == original                     # byte-identical -- zero behavior change without name_mapping
    assert result["total"].reason == ""            # no rewrite note ever attached when nothing was rewritten
    assert oracle.read_text(encoding="utf-8") == original   # the reference file ON DISK was never touched


def test_evaluate_with_replaced_subdir_name_mapping_empty_dict_is_also_byte_identical(tmp_path: Path):
    target = tmp_path / "target"
    _write(target / "Cargo.toml", "X")
    oracle = tmp_path / "checkdigit_test.rs"
    original = "#[test]\nfn t() { NewLuhn(); }\n"
    _write(oracle, original)
    snapshot = SnapshotRunner(stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")

    COL._evaluate_with_replaced_subdir(target, [oracle], "tests", "oxidizer", ["cargo", "test"],
                                       timeout=None, runner=snapshot, name_mapping={})

    assert snapshot.calls[0]["cwd_files"]["tests/checkdigit_test.rs"] == original


def test_evaluate_with_replaced_subdir_name_mapping_noop_file_falls_back_to_byte_copy(tmp_path: Path):
    """A name_mapping IS supplied, but this particular file has no eligible
    substitution -- must still be a plain, byte-identical copy (the
    ``shutil.copy2`` fallback path), not a re-encoded/rewritten write."""
    target = tmp_path / "target"
    _write(target / "Cargo.toml", "X")
    oracle = tmp_path / "other_test.rs"
    original = "#[test]\nfn t() { unrelated_call(); }\n"
    _write(oracle, original)
    snapshot = SnapshotRunner(stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")

    result = COL._evaluate_with_replaced_subdir(target, [oracle], "tests", "oxidizer", ["cargo", "test"],
                                                timeout=None, runner=snapshot,
                                                name_mapping={"NewLuhn": "new_luhn"})

    assert snapshot.calls[0]["cwd_files"]["tests/other_test.rs"] == original
    assert result["total"].reason == ""


def test_evaluate_with_replaced_subdir_stages_and_protects_rust_fixture_module(tmp_path: Path):
    target = tmp_path / "target"
    _write(target / "Cargo.toml", "X")
    _write(target / "src" / "lib.rs", "pub const TEST_DATA: &[i32] = &[2];")
    oracle = tmp_path / "histogram_test.rs"
    support = tmp_path / "test_data.rs"
    _write(
        oracle,
        "mod test_data;\nuse test_data::testData;\n"
        "#[test]\nfn test_histogram() { assert_eq!(testData[0], 1); }\n",
    )
    _write(support, "pub const testData: &[i32] = &[1];\n")
    snapshot = SnapshotRunner(
        stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n",
    )

    result = COL._evaluate_with_replaced_subdir(
        target,
        [oracle],
        "tests",
        "oxidizer",
        ["cargo", "test"],
        timeout=None,
        runner=snapshot,
        support_files=[support],
    )

    copied = snapshot.calls[0]["cwd_files"]
    assert copied["tests/histogram_test.rs"] == oracle.read_text(encoding="utf-8")
    assert copied["tests/test_data.rs"] == support.read_text(encoding="utf-8")
    assert result["passed"].value == 1


def test_evaluate_independent_oracle_alphatrans_never_leaks_reference_python_impl(tmp_path: Path):
    ref_root = tmp_path / "refroot"
    _make_alphatrans_reference(ref_root, "alpha_proj")
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "impl.py", "OWN_IMPL")
    _write(target / "verified_test" / "stale_test.py", "OWN_STALE_TEST")   # must be wiped, not merged
    manifest_row = {"project": "alpha_proj"}
    snapshot = SnapshotRunner(stdout="1 passed in 0.02s\n")

    result = COL.evaluate_independent_oracle("alphatrans", run_dir, manifest_row, ALPHATRANS_SPEC,
                                             ref_root, timeout=None, runner=snapshot)

    assert result.validated["executed"].is_measured
    assert result.validated["executed"].value == 1
    # expected is a static test-function count over verified_test/ ALONE,
    # read straight from --reference-results-root -- here it matches the
    # (mocked) executed count exactly, so not_executed is 0.
    assert result.validated["expected"].is_measured
    assert result.validated["expected"].value == 1
    assert result.validated["not_executed"].value == 0
    # No reusable per-function harness is known for AlphaTrans -- regardless of the flag.
    assert result.function_validation["total"].status == Status.UNAVAILABLE
    assert result.function_validation["total"].value is None
    # No agent_test/ was built in this fixture -- function_harness_tests_* is
    # unavailable (structurally distinct from function_validation_* above),
    # never a fabricated 0.
    assert result.function_harness_tests["total"].status == Status.UNAVAILABLE
    assert result.function_harness_tests["total"].value is None
    assert result.oracle_integrity.status == Status.NOT_APPLICABLE

    assert len(snapshot.calls) == 1
    assert snapshot.calls[0]["argv"] == list(COL.ALPHATRANS_VERIFIED_TEST_CMD)
    files = snapshot.calls[0]["cwd_files"]
    assert files["impl.py"] == "OWN_IMPL"                       # CodeWeaver's own impl, untouched
    assert "python/impl.py" not in files                        # reference's own impl tree NEVER copied in
    assert files["verified_test/test_foo.py"] == "def test_foo():\n    assert True\n"
    assert files["verified_test/conftest.py"] == "REF_CONFTEST"
    assert "verified_test/stale_test.py" not in files           # wiped, never merged with the reference copy


def test_evaluate_independent_oracle_alphatrans_function_harness_tests_measured_when_agent_test_present(
    tmp_path: Path,
):
    """Both AlphaTrans oracles run TOGETHER for one project: verified_test/
    (validated_tests_*, unchanged) AND agent_test/ (the NEW
    function_harness_tests_*) -- structurally separate, never conflated, and
    function_validation_* remains unavailable regardless."""
    ref_root = tmp_path / "refroot"
    ref = _make_alphatrans_reference(ref_root, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=True)
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "impl.py", "OWN_IMPL")
    manifest_row = {"project": "alpha_proj"}
    snapshot = SnapshotRunner(stdout="1 passed in 0.02s\n")

    result = COL.evaluate_independent_oracle("alphatrans", run_dir, manifest_row, ALPHATRANS_SPEC,
                                             ref_root, timeout=None, runner=snapshot)

    assert result.validated["executed"].is_measured             # from verified_test/
    assert result.function_validation["total"].status == Status.UNAVAILABLE   # never conflated with the below
    assert result.function_harness_tests["total"].is_measured                 # from agent_test/
    assert len(snapshot.calls) == 2
    argvs = [c["argv"] for c in snapshot.calls]
    assert list(COL.ALPHATRANS_VERIFIED_TEST_CMD) in argvs
    assert list(COL.ALPHATRANS_FUNCTION_HARNESS_TEST_CMD) in argvs


def test_evaluate_independent_oracle_alphatrans_unavailable_without_reference_root(tmp_path: Path):
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "target" / "impl.py", "OWN_IMPL")
    manifest_row = {"project": "alpha_proj"}
    snapshot = SnapshotRunner()
    result = COL.evaluate_independent_oracle("alphatrans", run_dir, manifest_row, ALPHATRANS_SPEC,
                                             None, timeout=None, runner=snapshot)
    assert result.validated["executed"].status == Status.UNAVAILABLE
    assert result.validated["executed"].value is None
    assert result.validated["expected"].status == Status.UNAVAILABLE
    assert result.validated["expected"].value is None
    assert result.validated["not_executed"].status == Status.UNAVAILABLE
    assert result.validated["not_executed"].value is None
    assert result.function_validation["total"].status == Status.UNAVAILABLE
    assert result.function_validation["total"].value is None
    assert result.function_harness_tests["total"].status == Status.UNAVAILABLE
    assert result.function_harness_tests["total"].value is None
    assert snapshot.calls == []


def test_evaluate_independent_oracle_skel_always_unavailable_regardless_of_reference_root(tmp_path: Path):
    """function_validation_* stays Status.UNAVAILABLE for SKEL unconditionally
    (no reusable per-function harness is known -- see
    function_harness_tests_* instead). validated_tests_* is ALSO
    Status.UNAVAILABLE here specifically because neither ref_root below ever
    resolves a test_name_mapping.csv (the populated ``refroot`` has no CSV at
    all) -- it is NOT unconditionally unavailable for SKEL in general; see
    test_skel_validated_tests_eval_measured_and_never_leaks_source_js_or_mutates_target
    and test_evaluate_independent_oracle_skel_validated_tests_measured_with_reference_csv_and_source
    for the AST-extraction "measured" case once a real CSV + source.js are
    supplied. function_harness_tests_* is ALSO unavailable here specifically
    because neither ref_root below resolves to a populated
    javascript/*generated*.js reference tree (see the dedicated "measured"
    dispatch test for the populated case)."""
    run_dir = tmp_path / "run"
    _write(run_dir / "pipeline" / "target" / "index.js", "// own translation")
    manifest_row = {"project": "skel_proj"}
    snapshot = SnapshotRunner()

    for ref_root in (None, tmp_path / "refroot"):
        result = COL.evaluate_independent_oracle("skel", run_dir, manifest_row, {}, ref_root,
                                                 timeout=None, runner=snapshot)
        assert result.validated["executed"].status == Status.UNAVAILABLE
        assert result.validated["executed"].value is None
        assert result.validated["expected"].status == Status.UNAVAILABLE
        assert result.validated["expected"].value is None
        assert result.validated["not_executed"].status == Status.UNAVAILABLE
        assert result.validated["not_executed"].value is None
        assert result.function_validation["total"].status == Status.UNAVAILABLE
        assert result.function_validation["total"].value is None
        assert result.function_harness_tests["total"].status == Status.UNAVAILABLE
        assert result.function_harness_tests["total"].value is None
        assert result.oracle_integrity.status == Status.NOT_APPLICABLE
    assert snapshot.calls == []   # never spawns a subprocess for SKEL


def test_evaluate_independent_oracle_skel_function_harness_tests_measured_when_javascript_generated_present(
    tmp_path: Path,
):
    """SKEL's per-function validation stays unavailable (unconditionally --
    no reusable per-function harness is known at all), and validated_tests_*
    stays unavailable too because THIS fixture's reference tree
    (_make_skel_reference) ships no test_name_mapping.csv -- but
    function_harness_tests_* becomes measured once a populated
    javascript/*generated*.js reference tree is supplied. The three must
    never be conflated."""
    ref_root = tmp_path / "refroot"
    ref = _make_skel_reference(ref_root, "bst")
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "index.js", "// own translation")
    manifest_row = {"project": "bst"}
    runner = PerFileRunner(results={}, default=_exec(returncode=0))

    result = COL.evaluate_independent_oracle("skel", run_dir, manifest_row, {}, ref_root,
                                             timeout=None, runner=runner)

    assert result.validated["executed"].status == Status.UNAVAILABLE          # no CSV in this fixture
    assert result.validated["expected"].status == Status.UNAVAILABLE
    assert result.function_validation["total"].status == Status.UNAVAILABLE   # unchanged -- no reliable mapping
    assert result.function_harness_tests["total"].is_measured
    assert result.function_harness_tests["total"].value == 2               # 2 *generated*.js files in the fixture
    assert result.function_harness_tests["passed"].value == 2
    assert len(runner.calls) == 2


def test_evaluate_independent_oracle_skel_validated_tests_measured_with_reference_csv_and_source(tmp_path: Path):
    """SKEL's NEW validated_tests_* becomes measured once a real
    test_name_mapping.csv + javascript/source.js are supplied via
    --reference-results-root, while function_validation_* stays unavailable
    (unchanged -- SKEL still has no reliable PER-FUNCTION harness) and
    function_harness_tests_* is independently evaluated from the SAME
    reference tree's *generated*.js files -- all three structurally
    separate, never conflated."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref_root = tmp_path / "refroot"
    ref = _make_skel_reference_with_csv(
        ref_root, "bst", verified_js_names=["test_trivial", "test_uses_export"],
    )
    _write(ref / "javascript" / "SKELTest_generated.js", "GENERATED_MAIN")
    run_dir = tmp_path / "run"
    target = run_dir / "pipeline" / "target"
    _write(target / "index.js", "module.exports = { add: (a, b) => a + b };")
    manifest_row = {"project": "bst"}
    runner = PerFileRunner(
        results={COL.SKEL_VALIDATED_HARNESS_FILENAME: _exec(returncode=0, stdout="# pass 2\n# fail 0\n")},
        default=_exec(returncode=0),
    )

    result = COL.evaluate_independent_oracle("skel", run_dir, manifest_row, {}, ref_root,
                                             timeout=None, runner=runner)

    assert result.validated["executed"].is_measured
    assert result.validated["executed"].value == 2
    assert result.validated["passed"].value == 2
    # expected == len(verified test names) from test_name_mapping.csv, read
    # BEFORE any tree-sitter/source.js extraction is attempted -- here it
    # matches the executed count exactly (both verified names extracted OK).
    assert result.validated["expected"].is_measured
    assert result.validated["expected"].value == 2
    assert result.validated["not_executed"].value == 0
    assert result.function_validation["total"].status == Status.UNAVAILABLE   # unchanged
    assert result.function_harness_tests["total"].is_measured                 # from the SAME ref tree, separately
    assert result.function_harness_tests["total"].value == 1
    assert result.oracle_integrity.status == Status.NOT_APPLICABLE


def test_evaluate_independent_oracle_crust_uses_scaffold_regardless_of_reference_results_root(tmp_path: Path):
    """CRUST's own scaffold IS its independent oracle; reference_results_root
    (the OTHER three tools' mechanism) is irrelevant to it. This minimal
    fixture ships no ``.rs`` scaffold contract file (only ``Cargo.toml``), so
    ``expected`` is honestly UNAVAILABLE here -- see
    test_crust_validated_tests_eval_uses_pristine_content_despite_mutated_target
    for the populated-scaffold "measured" case."""
    run_dir = tmp_path / "run"
    _write(run_dir / "scaffold" / "Cargo.toml", "PRISTINE")
    _write(run_dir / "pipeline" / "target" / "Cargo.toml", "PRISTINE")
    snapshot = SnapshotRunner(stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")
    result = COL.evaluate_independent_oracle("crust", run_dir, {"project": "bitset"}, CRUST_SPEC,
                                             tmp_path / "nonexistent-results-root",
                                             timeout=None, runner=snapshot)
    assert result.validated["executed"].is_measured
    assert result.validated["expected"].status == Status.UNAVAILABLE   # no .rs contract file in this fixture
    assert result.validated["not_executed"].status == Status.UNAVAILABLE
    assert result.oracle_integrity.value == "pristine"
    assert result.function_validation["total"].status == Status.NOT_APPLICABLE
    assert result.function_harness_tests["total"].status == Status.NOT_APPLICABLE   # no such concept for CRUST


# --------------------------------------------------------------------------- #
# collect_run integration: translated vs. validated fields end-to-end
# --------------------------------------------------------------------------- #
OXIDIZER_MANIFEST_ROW = {"id": "oxidizer__proj", "tool": "oxidizer", "target_language": "Rust",
                        "function_count_source": 4, "project": "oxi_proj"}


def test_collect_run_validated_tests_unavailable_not_zero_without_reference_root(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write_state(run_dir)
    runner = FakeRunner(default_stdout="test result: ok. 3 passed; 0 failed; 0 ignored; 0 measured\n")
    row = COL.collect_run(run_dir, variant="full", project_id="oxidizer__proj", tool="oxidizer", repetition=0,
                          manifest_row=OXIDIZER_MANIFEST_ROW, dataset_spec=OXIDIZER_SPEC, runner=runner)
    # translated_tests_* (CodeWeaver's own self-graded tests) ARE measured...
    assert row["translated_tests_total"] == 3
    assert row["translated_tests_total_status"] == Status.MEASURED
    assert row["translated_tests_pass_rate"] == pytest.approx(1.0)
    # ...but validated_tests_*/function_validation_* (the independent oracle)
    # are explicitly unavailable -- never a fabricated 0 -- because
    # --reference-results-root was never supplied to collect_run().
    assert row["validated_tests_expected"] is None
    assert row["validated_tests_expected_status"] == Status.UNAVAILABLE
    assert row["validated_tests_executed"] is None
    assert row["validated_tests_executed_status"] == Status.UNAVAILABLE
    assert row["validated_tests_not_executed"] is None
    assert row["validated_tests_not_executed_status"] == Status.UNAVAILABLE
    assert row["validated_tests_pass_rate"] is None
    assert row["validated_tests_pass_rate_status"] == Status.UNAVAILABLE
    assert row["function_validation_total"] is None
    assert row["function_validation_total_status"] == Status.UNAVAILABLE
    assert row["oracle_integrity"] is None
    assert row["oracle_integrity_status"] == Status.NOT_APPLICABLE


def test_collect_run_threads_reference_results_root_into_crust_validated_tests(tmp_path: Path):
    """End-to-end: collect_run's own reference_results_root parameter must
    reach the CRUST oracle evaluation (which does not even need the
    parameter's VALUE, only proves the plumbing doesn't drop/break it) and
    produce a real measured validated_tests_* result from the pristine
    scaffold, independent of dev_tests_*/translated_tests_*."""
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write(run_dir / "scaffold" / "Cargo.toml", "PRISTINE")
    # 6 #[test] functions in the pristine scaffold's own contract dir -- the
    # paper's FIXED, oracle-known validated_tests_expected denominator,
    # matching the mocked "5 passed; 1 failed" == 6 executed below exactly.
    harness_rs = "\n".join(f"#[test]\nfn t{i}() {{ assert!(true); }}" for i in range(6))
    _write(run_dir / "scaffold" / "src" / "bin" / "harness.rs", harness_rs)
    _write(run_dir / "pipeline" / "target" / "Cargo.toml", "PRISTINE")
    # target's own copy is IDENTICAL to the scaffold's, so oracle_integrity
    # reports "pristine" here (a separate concern from validated_tests_*,
    # which never reads this copy anyway -- see crust_validated_tests_eval).
    _write(run_dir / "pipeline" / "target" / "src" / "bin" / "harness.rs", harness_rs)
    _write_state(run_dir)
    runner = FakeRunner(default_stdout="test result: ok. 5 passed; 1 failed; 0 ignored; 0 measured\n")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust", repetition=0,
                          manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=runner,
                          reference_results_root=tmp_path / "unused-for-crust")
    assert row["validated_tests_expected"] == 6
    assert row["validated_tests_expected_status"] == Status.MEASURED
    assert row["validated_tests_executed"] == 6
    assert row["validated_tests_passed"] == 5
    assert row["validated_tests_failed"] == 1
    assert row["validated_tests_not_executed"] == 0
    # paper-equivalent TPR is passed/expected (5/6), NOT passed/executed --
    # here they coincide numerically only because executed == expected, but
    # the formula used is unambiguously the expected-denominator one (see
    # test_compute_paper_pass_rate_* in this module for a case where they differ).
    assert row["validated_tests_pass_rate"] == pytest.approx(5 / 6)
    assert row["validated_tests_executed_status"] == Status.MEASURED
    assert row["oracle_integrity"] == "pristine"


def test_collect_run_function_harness_tests_measured_for_alphatrans_agent_test(tmp_path: Path):
    """End-to-end: collect_run flattens AlphaTrans's NEW
    function_harness_tests_* (agent_test/ GENERATED function-harness
    execution) into the row, structurally separate from BOTH
    validated_tests_* (verified_test/, unaffected) and the
    (still-unavailable) function_validation_*."""
    ref_root = tmp_path / "refroot"
    ref = _make_alphatrans_reference(ref_root, "alpha_proj")
    _make_alphatrans_agent_test(ref, nested_python_subdir=True)
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write(run_dir / "pipeline" / "target" / "impl.py", "OWN_IMPL")
    _write_state(run_dir)
    manifest_row = {"id": "alphatrans__alpha_proj", "tool": "alphatrans", "target_language": "Python",
                    "function_count_source": 4, "project": "alpha_proj"}
    runner = FakeRunner(default_stdout="1 passed in 0.02s\n")

    row = COL.collect_run(run_dir, variant="full", project_id="alphatrans__alpha_proj", tool="alphatrans",
                          repetition=0, manifest_row=manifest_row, dataset_spec=ALPHATRANS_SPEC, runner=runner,
                          reference_results_root=ref_root)

    assert row["validated_tests_expected"] == 1             # from verified_test/, unaffected by the below
    assert row["validated_tests_executed"] == 1
    assert row["validated_tests_executed_status"] == Status.MEASURED
    assert row["validated_tests_not_executed"] == 0
    assert row["function_validation_total"] is None         # still unavailable -- never conflated
    assert row["function_validation_total_status"] == Status.UNAVAILABLE
    assert row["function_harness_tests_total"] == 1         # from agent_test/ -- the NEW field
    assert row["function_harness_tests_total_status"] == Status.MEASURED
    assert row["function_harness_tests_passed"] == 1
    assert row["function_harness_tests_pass_rate"] == pytest.approx(1.0)


def test_collect_run_function_harness_tests_measured_for_skel_generated_js(tmp_path: Path):
    """End-to-end: collect_run flattens SKEL's NEW function_harness_tests_*
    (javascript/*generated*.js GENERATED function-harness execution) into
    the row, while validated_tests_*/function_validation_* stay unavailable
    for THIS fixture specifically (_make_skel_reference ships no
    test_name_mapping.csv; see the dedicated validated_tests_* collect_run
    test below for the populated-CSV "measured" case)."""
    ref_root = tmp_path / "refroot"
    _make_skel_reference(ref_root, "bst")
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write(run_dir / "pipeline" / "target" / "index.js", "// own translation")
    _write_state(run_dir)
    manifest_row = {"id": "skel__bst", "tool": "skel", "target_language": "JavaScript",
                    "function_count_source": 4, "project": "bst"}
    skel_spec = {"label": "SKEL", "target_language": "JavaScript", "build_cmd": [], "unit_test_cmd": []}
    runner = FakeRunner(default_returncode=0, default_stdout="")

    row = COL.collect_run(run_dir, variant="full", project_id="skel__bst", tool="skel", repetition=0,
                          manifest_row=manifest_row, dataset_spec=skel_spec, runner=runner,
                          reference_results_root=ref_root)

    assert row["validated_tests_expected"] is None
    assert row["validated_tests_expected_status"] == Status.UNAVAILABLE
    assert row["validated_tests_executed"] is None
    assert row["validated_tests_executed_status"] == Status.UNAVAILABLE
    assert row["function_validation_total"] is None
    assert row["function_validation_total_status"] == Status.UNAVAILABLE
    assert row["function_harness_tests_total"] == 2          # 2 *generated*.js files in the fixture
    assert row["function_harness_tests_total_status"] == Status.MEASURED
    assert row["function_harness_tests_passed"] == 2
    assert row["function_harness_tests_pass_rate"] == pytest.approx(1.0)


def test_collect_run_validated_tests_measured_for_skel_with_reference_csv(tmp_path: Path):
    """End-to-end: collect_run flattens SKEL's NEW validated_tests_* (AST-
    extracted, independently-executed developer tests) into the row once a
    real test_name_mapping.csv + javascript/source.js are available under
    --reference-results-root -- structurally separate from
    function_harness_tests_*/function_validation_*, never conflated."""
    parser = COL._skel_js_parser()
    if parser is None:
        pytest.skip("tree-sitter/tree-sitter-javascript not installed in this environment")
    ref_root = tmp_path / "refroot"
    _make_skel_reference_with_csv(ref_root, "bst", verified_js_names=["test_trivial", "test_uses_export"])
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write(run_dir / "pipeline" / "target" / "index.js", "module.exports = { add: (a, b) => a + b };")
    _write_state(run_dir)
    manifest_row = {"id": "skel__bst", "tool": "skel", "target_language": "JavaScript",
                    "function_count_source": 4, "project": "bst"}
    skel_spec = {"label": "SKEL", "target_language": "JavaScript", "build_cmd": [], "unit_test_cmd": []}
    runner = PerFileRunner(
        results={COL.SKEL_VALIDATED_HARNESS_FILENAME: _exec(returncode=0, stdout="# pass 2\n# fail 0\n")},
        default=_exec(returncode=0),
    )

    row = COL.collect_run(run_dir, variant="full", project_id="skel__bst", tool="skel", repetition=0,
                          manifest_row=manifest_row, dataset_spec=skel_spec, runner=runner,
                          reference_results_root=ref_root)

    assert row["validated_tests_expected"] == 2
    assert row["validated_tests_expected_status"] == Status.MEASURED
    assert row["validated_tests_executed"] == 2
    assert row["validated_tests_passed"] == 2
    assert row["validated_tests_not_executed"] == 0
    assert row["validated_tests_pass_rate"] == pytest.approx(1.0)
    assert row["function_validation_total"] is None
    assert row["function_validation_total_status"] == Status.UNAVAILABLE
    assert row["function_harness_tests_total"] is None        # no *generated*.js files in THIS fixture
    assert row["function_harness_tests_total_status"] == Status.UNAVAILABLE


# --------------------------------------------------------------------------- #
# CLI: --reference-results-root
# --------------------------------------------------------------------------- #
def test_build_parser_accepts_reference_results_root():
    parser = COL.build_parser()
    ns = parser.parse_args(["--manifest", "m.json", "--runs-root", "r", "--output-root", "o",
                            "--reference-results-root", "/data/results"])
    assert ns.reference_results_root == "/data/results"


def test_build_parser_reference_results_root_defaults_to_none():
    parser = COL.build_parser()
    ns = parser.parse_args(["--manifest", "m.json", "--runs-root", "r", "--output-root", "o"])
    assert ns.reference_results_root is None


# --------------------------------------------------------------------------- #
# collect_run: full integration over a fixture run directory
# --------------------------------------------------------------------------- #
CRUST_SPEC = {
    "label": "CRUST", "target_language": "Rust",
    "build_cmd": ["cargo", "build"], "unit_test_cmd": ["cargo", "test"],
    "coverage_cmd": [], "coverage_format": "",
}
MANIFEST_ROW = {
    "id": "crust__bitset", "tool": "crust", "target_language": "Rust", "function_count_source": 4,
}


def _write_state(run_dir: Path, *, status="completed", error="", app_id="app1", ablation=None) -> None:
    state = {
        "variant": "full", "project_id": "crust__bitset", "repetition": 0, "status": status,
        "app_id": app_id, "workspace_dir": str(run_dir), "argv": None, "returncode": 0, "attempt": 1,
        "created_at": "2024-01-01T00:00:00.000000Z", "updated_at": "2024-01-01T00:05:00.000000Z",
        "started_at": "2024-01-01T00:00:00.000000Z", "ended_at": "2024-01-01T00:05:00.000000Z",
        "timeout_seconds": None, "error": error,
        "provenance": {"model": {"value": "claude-opus-4.8", "status": "measured", "reason": ""},
                      "git_sha": {"value": "deadbeef", "status": "measured", "reason": ""}},
    }
    if ablation is not None:
        state["ablation"] = ablation
    C.atomic_write_json(run_dir / R.STATE_FILENAME, state)


def test_collect_run_raises_skip_when_run_dir_missing(tmp_path: Path):
    with pytest.raises(COL.CollectionSkip, match="not_attempted"):
        COL.collect_run(tmp_path / "nope", variant="full", project_id="crust__bitset", tool="crust",
                        repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC)


def test_collect_run_raises_skip_when_no_state_file(tmp_path: Path):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    with pytest.raises(COL.CollectionSkip, match="no_state_file"):
        COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust",
                        repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC)


def test_collect_run_raises_skip_when_not_terminal(tmp_path: Path):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    _write_state(run_dir, status="running")
    with pytest.raises(COL.CollectionSkip, match="not_terminal"):
        COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust",
                        repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC)


def test_collect_run_measures_build_and_tests_for_full_variant(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "pipeline" / "logs").mkdir(parents=True)
    _write_state(run_dir)
    (run_dir / "cli.stdout.log").write_text(CLI_STDOUT_SAMPLE, encoding="utf-8")

    runner = FakeRunner(default_stdout="test result: ok. 3 passed; 0 failed; 0 ignored; 0 measured\n")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust", repetition=0,
                          manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=runner)

    assert row["run_status"] == "completed"
    assert row["build"] is True
    assert row["build_status"] == Status.MEASURED
    assert row["dev_tests_total"] == 3
    assert row["dev_tests_passed"] == 3
    assert row["dev_test_pass_rate"] == pytest.approx(1.0)
    assert row["baseline_build_status"] == Status.NOT_APPLICABLE   # no scaffold for this fixture
    assert row["trajectory_precision"] == "lower_bound"
    assert row["nc"] is not None and row["tec"] is not None
    assert row["milestones_total"] == 2
    assert row["milestone_granularity"] == "real"
    assert row["model"] == "claude-opus-4.8"
    assert row["git_sha"] == "deadbeef"


def test_collect_run_reports_missing_not_zero_when_nothing_produced(tmp_path: Path):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    _write_state(run_dir, status="failed", error="agent crashed")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust", repetition=0,
                          manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["run_status"] == "failed"
    assert row["build_status"] == Status.MISSING     # NOT False/0 -- nothing was produced
    assert row["build"] is None
    assert row["dev_tests_total_status"] == Status.MISSING
    assert row["dev_tests_total"] is None


def test_collect_run_uses_scaffold_as_baseline_when_present(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "scaffold").mkdir(parents=True)
    _write_state(run_dir)
    runner = FakeRunner(default_stdout="test result: ok. 1 passed; 0 failed; 0 ignored; 0 measured\n")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust", repetition=0,
                          manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=runner)
    assert row["baseline_build_status"] == Status.MEASURED
    assert row["baseline_tests_total_status"] == Status.MEASURED


def test_collect_run_baseagent_variant_uses_calls_jsonl_trajectory(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    _write_state(run_dir, status="completed")
    calls = _ablation_calls("plan")
    calls_path = run_dir / R.CALLS_FILENAME
    for c in calls:
        C.append_jsonl(calls_path, c)
    row = COL.collect_run(run_dir, variant="baseagent-condensed", project_id="crust__bitset", tool="crust",
                          repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["trajectory_precision"] == "exact"
    # Regression for review finding #3: the skipped "plan" stage's placeholder
    # call must not inflate nc -- it never executed, so nc must equal tec (4),
    # NOT count all 5 stage slots as if every one of them had actually run.
    assert row["nc"] == 4
    assert row["tec"] == 4
    assert row["milestone_granularity"] == "single-synthetic"
    assert row["ablation_skipped_stage"] is None    # baseagent-* never sets CODEWEAVER_SKIP_STAGES


def test_collect_run_stage_skip_ablation_uses_real_burr_trajectory_not_calls_jsonl(tmp_path: Path):
    """Regression: noanalyzer/noplanning/novalidator now run the identical
    real `python -m codeweaver run` CLI subprocess as `full`
    (CodeWeaver core's CODEWEAVER_SKIP_STAGES instrumentation -- see
    run.py.STAGE_SKIP_VARIANTS), so collect_run must reconstruct their
    trajectory/milestones from cli.stdout.log exactly like `full`, NOT from
    the degenerate single "full_pipeline"/"cli" entry run.py now writes to
    recodeagent_calls.jsonl for these variants (which would otherwise
    wrongly collapse the whole run into a single synthetic node/milestone,
    discarding all the real per-milestone/per-stage evidence that is
    actually available)."""
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "pipeline" / "logs").mkdir(parents=True)
    _write_state(run_dir, status="completed", ablation={"skipped_stage": "plan", "execution_mode": "full_burr_graph"})
    (run_dir / "cli.stdout.log").write_text(CLI_STDOUT_SAMPLE, encoding="utf-8")
    # run.py's CURRENT shape for a stage-skip ablation: one degenerate
    # "full_pipeline"/"cli" call -- must NOT be what trajectory/milestones
    # are derived from for this variant.
    C.append_jsonl(run_dir / R.CALLS_FILENAME,
                   {"stage": "full_pipeline", "kind": "cli", "agent": None, "ok": True})
    row = COL.collect_run(run_dir, variant="noplanning", project_id="crust__bitset", tool="crust",
                          repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["trajectory_precision"] == "lower_bound"     # real reconstruction, not calls-based "exact"
    assert row["milestone_granularity"] == "real"
    assert row["milestones_total"] == 2       # m1, m2 -- genuinely reconstructed from history
    assert row["milestones_passed"] == 2      # noplanning's validator is real
    assert row["ablation_skipped_stage"] == "plan"
    sec = json.loads(row["sec_json"])
    assert "plan" not in sec                  # deliberately skipped -- excluded, not zero-masked
    assert sec["translate"] == 3
    assert row["nc"] == sum(1 for v in sec.values() if v > 0)
    assert row["tec"] == sum(sec.values())


def test_collect_run_novalidator_real_history_never_yields_fabricated_milestone_pass(tmp_path: Path):
    """Regression for review finding #3, updated for the current
    CODEWEAVER_SKIP_STAGES-based mechanism: novalidator now runs the real
    Burr CLI subprocess, whose validate() skip branch appends passed=None to
    EVERY milestone's history entry (no genuine validator attestation
    exists) -- collect_run() must never turn that None into either a
    fabricated pass (True/1) or a fabricated confirmed-failure (0). total
    must remain genuinely measured (the milestones DID run); passed must be
    reported missing."""
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "pipeline" / "logs").mkdir(parents=True)
    _write_state(run_dir, status="completed",
                ablation={"skipped_stage": "validate", "execution_mode": "full_burr_graph"})
    (run_dir / "cli.stdout.log").write_text(CLI_STDOUT_NOVALIDATOR_SAMPLE, encoding="utf-8")
    C.append_jsonl(run_dir / R.CALLS_FILENAME,
                   {"stage": "full_pipeline", "kind": "cli", "agent": None, "ok": True})
    row = COL.collect_run(run_dir, variant="novalidator", project_id="crust__bitset", tool="crust",
                          repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["milestone_granularity"] == "real"
    assert row["milestones_total_status"] == Status.MEASURED    # the milestones themselves DID run
    assert row["milestones_total"] == 2
    assert row["milestones_passed_status"] == Status.MISSING    # NOT a fabricated "2 passed" or "0 passed"
    assert row["milestones_passed"] is None
    assert row["ablation_skipped_stage"] == "validate"
    # The skipped "validate" stage itself must not read as an executed node either.
    sec = json.loads(row["sec_json"])
    assert "validate" not in sec


def test_collect_run_ablation_skipped_stage_none_for_full_variant(tmp_path: Path):
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "pipeline" / "logs").mkdir(parents=True)
    _write_state(run_dir, status="completed")
    (run_dir / "cli.stdout.log").write_text(CLI_STDOUT_SAMPLE, encoding="utf-8")
    row = COL.collect_run(run_dir, variant="full", project_id="crust__bitset", tool="crust",
                          repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["ablation_skipped_stage"] is None


def test_collect_run_ablation_skipped_stage_falls_back_to_static_mapping(tmp_path: Path):
    """When run_state.json predates run.py's own ``state["ablation"]``
    bookkeeping (e.g. an older run directory), collect_run must still
    correctly identify the skipped stage from the static
    run.py.STAGE_SKIP_VARIANTS variant->stage mapping, never silently
    treat the run as if nothing were skipped."""
    run_dir = tmp_path / "run"
    (run_dir / "pipeline" / "target").mkdir(parents=True)
    (run_dir / "pipeline" / "logs").mkdir(parents=True)
    _write_state(run_dir, status="completed")   # no "ablation" key at all
    (run_dir / "cli.stdout.log").write_text(CLI_STDOUT_SAMPLE, encoding="utf-8")
    row = COL.collect_run(run_dir, variant="noanalyzer", project_id="crust__bitset", tool="crust",
                          repetition=0, manifest_row=MANIFEST_ROW, dataset_spec=CRUST_SPEC, runner=FakeRunner())
    assert row["ablation_skipped_stage"] == "analyze"
    sec = json.loads(row["sec_json"])
    assert "analyze" not in sec


# --------------------------------------------------------------------------- #
# collect_all: full expected matrix, never-attempted vs. measured
# --------------------------------------------------------------------------- #
def _manifest(project_ids: list[str]) -> dict:
    return {"projects": [
        {"id": pid, "tool": "crust", "project": pid, "source_language": "C", "target_language": "Rust",
        "source_rel_path": f"{pid}/CBench", "oracle_rel_path": None, "scaffold_rel_path": None,
        "ground_truth_target_rel_path": None, "loc_source": 1, "test_count_source": 0,
        "function_count_source": 1, "status": "ok", "notes": "", "discovered_at": C.utcnow_iso()}
        for pid in project_ids
    ]}


def test_collect_all_reports_not_attempted_for_missing_run_dirs(tmp_path: Path):
    manifest = _manifest(["crust__a", "crust__b"])
    rows, failures = COL.collect_all(tmp_path / "runs", manifest, variants=["full"], repetitions=1,
                                    dataset_specs={"crust": CRUST_SPEC}, runner=FakeRunner())
    assert rows == []
    assert len(failures) == 2
    assert all("not_attempted" in f["reason"] for f in failures)


def test_collect_all_partitions_measured_and_unattempted(tmp_path: Path):
    manifest = _manifest(["crust__a", "crust__b"])
    runs_root = tmp_path / "runs"
    run_dir_a = R.run_dir_for(runs_root, "full", "crust__a", 0)
    (run_dir_a / "pipeline" / "target").mkdir(parents=True)
    _write_state(run_dir_a)

    rows, failures = COL.collect_all(runs_root, manifest, variants=["full"], repetitions=1,
                                    dataset_specs={"crust": CRUST_SPEC}, runner=FakeRunner())
    assert len(rows) == 1
    assert rows[0]["project_id"] == "crust__a"
    assert len(failures) == 1
    assert failures[0]["project_id"] == "crust__b"


def test_collect_all_never_raises_on_a_single_bad_run(tmp_path: Path, monkeypatch):
    manifest = _manifest(["crust__a"])
    runs_root = tmp_path / "runs"
    run_dir_a = R.run_dir_for(runs_root, "full", "crust__a", 0)
    (run_dir_a / "pipeline" / "target").mkdir(parents=True)
    _write_state(run_dir_a)

    def boom(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(COL, "evaluate_build", boom)
    rows, failures = COL.collect_all(runs_root, manifest, variants=["full"], repetitions=1,
                                    dataset_specs={"crust": CRUST_SPEC}, runner=FakeRunner())
    assert rows == []
    assert len(failures) == 1
    assert "collection_error" in failures[0]["reason"]


# --------------------------------------------------------------------------- #
# Output writers
# --------------------------------------------------------------------------- #
def test_write_raw_runs_creates_csv_and_jsonl(tmp_path: Path):
    rows = [{
        "variant": "full", "project_id": "crust__a", "tool": "crust", "repetition": 0,
        "build": True, "build_status": Status.MEASURED, "build_reason": "",
        "dev_tests_total": None, "dev_tests_total_status": Status.MISSING, "dev_tests_total_reason": "x",
    }]
    json_path, csv_path = COL.write_raw_runs(rows, tmp_path)
    assert json_path.exists() and csv_path.exists()
    lines = json_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 1
    assert json.loads(lines[0])["project_id"] == "crust__a"
    csv_text = csv_path.read_text(encoding="utf-8")
    assert "crust__a" in csv_text
    assert "build_status" in csv_text.splitlines()[0]


def test_write_failures_creates_csv(tmp_path: Path):
    failures = [{"variant": "full", "project_id": "crust__a", "tool": "crust", "repetition": 0,
                "workspace_dir": "x", "reason": "not_attempted", "detected_at": C.utcnow_iso()}]
    path = COL.write_failures(failures, tmp_path)
    assert path.exists()
    text = path.read_text(encoding="utf-8")
    assert "not_attempted" in text


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def test_parse_variants_all_and_explicit_list():
    assert COL._parse_variants("all") == list(C.RUN_VARIANTS)
    assert COL._parse_variants("full,noanalyzer") == ["full", "noanalyzer"]


def test_parse_variants_rejects_unknown():
    with pytest.raises(ValueError):
        COL._parse_variants("not-a-real-variant")


def test_cli_main_smoke_no_toolchain_needed(tmp_path: Path):
    """No run directories exist at all -- collect_all must short-circuit to
    'not_attempted' failures without ever invoking a real command runner, so
    this is safe to run with no toolchain installed."""
    manifest_path = tmp_path / "manifest.json"
    manifest = _manifest(["crust__a"])
    C.atomic_write_json(manifest_path, manifest)
    output_root = tmp_path / "out"
    rc = COL.main([
        "--manifest", str(manifest_path), "--runs-root", str(tmp_path / "runs"),
        "--output-root", str(output_root), "--variant", "full", "--repetitions", "1",
    ])
    assert rc == 0
    assert (output_root / "raw_runs.csv").exists()
    assert (output_root / "failures.csv").exists()
    failures_text = (output_root / "failures.csv").read_text(encoding="utf-8")
    assert "not_attempted" in failures_text
